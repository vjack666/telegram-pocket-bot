"""
Reporte: ¿Cuánto capital necesito para ganar $60/día de forma sostenible?
=========================================================================
Datos: ejemplo.md — 2,655 señales reales (87.87% WR, 17/03 – 07/05/2026)
Sistema: Masaniello 12/4, cap TOTAL por señal = CAP_PCT del capital
         (entry_max = capital * CAP_PCT / 7.44)

Hojas:
  1. Resumen por Capital     — tabla + gráfica: cuántos días alcanza la meta
  2. Sensibilidad cap%       — qué pasa si cambiamos el cap (5% → 30%) a $300
  3. Mapa de calor           — capital × cap_pct → % días meta cumplida
  4. Análisis matemático     — EV, breakeven WR, tabla de stakes reales
  5. Detalle Capital óptimo  — día a día del mejor escenario identificado
"""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

HISTORY = ROOT / "ejemplo.md"
OUTPUT  = ROOT / "runtime" / "reporte_meta_60.xlsx"

# ── Constantes ────────────────────────────────────────────────────────────
PAYOUT_MULT = 1.92
NET_PAYOUT  = 0.92
_R          = PAYOUT_MULT / NET_PAYOUT   # ≈ 2.0870
_TOTAL_MULT = 1 + _R + _R ** 2           # ≈ 7.442
N_OPS       = 12
W_NEEDED    = 4
META_DIARIA = 60.0

# Capitales a simular
CAPITALES   = [100, 200, 300, 500, 750, 1000, 1500, 2000, 3000, 5000]
# Cap% a simular en sensibilidad
CAP_PCTS    = [0.05, 0.10, 0.15, 0.20, 0.25, 0.30]

# ── Colores ───────────────────────────────────────────────────────────────
AZUL_H   = "1F3864"; AZUL_M = "2E75B6"; AZUL_CL  = "D6E4F7"
VERDE_O  = "1E6F4B"; VERDE  = "D6F0E0"
ROJO_O   = "8B1A1A"; ROJO   = "FDDEDE"
NARANJA  = "FF8C00"; NARAN  = "FFE5B4"
AMARILLO = "FFF2CC"; GRIS   = "F2F2F2"; BLANCO = "FFFFFF"
MORADO   = "7030A0"; MORADO_CL = "EAD1DC"


# ═══════════════════════════════════════════════════════════════════════════
# 1. PARSER
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class Outcome:
    timestamp: datetime
    result: str
    is_win: bool


def parse_outcomes(path: Path) -> list[Outcome]:
    date_pat = re.compile(r"^\[(\d{2}/\d{2}/\d{4}) (\d{2}:\d{2}:\d{2})\]")
    res_pat  = re.compile(
        r"(VICTORIA DIRECTA|VICTORIA EN 1.*?MARTINGALA|VICTORIA EN 2.*?MARTINGALA|P[ÉE]RDIDA)",
        re.IGNORECASE,
    )
    label_map = {
        "victoria directa": "WD", "victoria en 1": "G1",
        "victoria en 2": "G2", "perdida": "L", "pérdida": "L",
    }
    outcomes = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not date_pat.match(line):
            continue
        res_m = res_pat.search(line)
        if not res_m:
            continue
        ts_str = " ".join(date_pat.match(line).groups())
        ts = datetime.strptime(ts_str, "%d/%m/%Y %H:%M:%S")
        raw = res_m.group(1).lower()
        label = next((v for k, v in label_map.items() if raw.startswith(k)), "?")
        outcomes.append(Outcome(ts, label, label != "L"))
    return outcomes


# ═══════════════════════════════════════════════════════════════════════════
# 2. FÓRMULA MASANIELLO
# ═══════════════════════════════════════════════════════════════════════════

def _fwd_prob(ops_left: int, wins_needed: int) -> float:
    pm = PAYOUT_MULT
    if wins_needed <= 0:        return 1.0
    if wins_needed > ops_left:  return 0.0
    if wins_needed == ops_left: return pm ** ops_left
    pw = _fwd_prob(ops_left - 1, wins_needed - 1)
    pl = _fwd_prob(ops_left - 1, wins_needed)
    d  = pw + NET_PAYOUT * pl
    return (pm * pw * pl / d) if d else 0.0


def masaniello_raw(losses: int, wins: int) -> float:
    """Stake puro Masaniello 12/4 sobre base=1 (se escala después)."""
    ops_left  = N_OPS - (losses + wins)
    wins_left = W_NEEDED - wins
    if ops_left <= 0 or wins_left <= 0 or wins_left > ops_left:
        return 0.0
    pw = _fwd_prob(ops_left - 1, wins_left - 1)
    pl = _fwd_prob(ops_left - 1, wins_left)
    d  = pw + NET_PAYOUT * pl
    if not d:
        return 1.0
    s = 1.0 * (1 - PAYOUT_MULT * pw / d)
    return max(0.001, min(s, 1.0))


# Precalcular tabla de stakes normalizados (sobre base=1)
_STAKE_TABLE: dict[tuple[int, int], float] = {
    (l, w): masaniello_raw(l, w)
    for l in range(N_OPS)
    for w in range(N_OPS)
    if l + w < N_OPS
}


# ═══════════════════════════════════════════════════════════════════════════
# 3. SIMULACIÓN
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class DayStats:
    start_bal: float = 0.0
    end_bal:   float = 0.0
    min_bal:   float = 0.0
    pnl:       float = 0.0
    sessions:  int   = 0
    meta_hit:  bool  = False
    sessions_pnl: list = None

    def __post_init__(self):
        if self.sessions_pnl is None:
            self.sessions_pnl = []


@dataclass
class SimResult:
    capital:        float
    cap_pct:        float
    final:          float
    peak:           float
    min_bal:        float
    max_dd:         float
    sessions_total: int
    sessions_won:   int
    total_sigs:     int
    days_meta:      int
    days_total:     int
    daily:          dict
    daily_pnls:     list   # lista de pnl diario


def simulate(outcomes: list[Outcome], capital: float, cap_pct: float) -> SimResult:
    entry_max = capital * cap_pct / _TOTAL_MULT

    balance        = capital
    peak           = capital
    min_bal        = capital
    max_dd         = 0.0
    sessions_total = 0
    sessions_won   = 0
    total_sigs     = 0

    daily: dict[str, DayStats] = {}

    chunks = [outcomes[i:i + N_OPS] for i in range(0, len(outcomes), N_OPS)]

    for chunk in chunks:
        if not chunk:
            continue
        day = chunk[0].timestamp.strftime("%d/%m/%Y")
        if day not in daily:
            daily[day] = DayStats(
                start_bal=round(balance, 2),
                end_bal=round(balance, 2),
                min_bal=round(balance, 2),
            )
        d = daily[day]
        if d.meta_hit:
            continue

        wins = losses = 0
        session_pnl = 0.0

        for outcome in chunk:
            wins_needed = W_NEEDED - wins
            ops_left    = N_OPS - (wins + losses)
            if wins_needed <= 0 or wins_needed > ops_left:
                break

            raw_stake = _STAKE_TABLE.get((losses, wins), 0.0)
            entry     = round(min(raw_stake * capital, entry_max), 2)
            if entry <= 0:
                break

            total_sigs += 1

            if outcome.is_win:
                wins       += 1
                gain        = round(entry * NET_PAYOUT, 2)
                balance     = round(balance + gain, 2)
                session_pnl = round(session_pnl + gain, 2)
                d.pnl       = round(d.pnl + gain, 2)
            else:
                losses     += 1
                total_loss  = round(entry * _TOTAL_MULT, 2)
                balance     = round(balance - total_loss, 2)
                session_pnl = round(session_pnl - total_loss, 2)
                d.pnl       = round(d.pnl - total_loss, 2)

            min_bal   = min(min_bal, balance)
            d.min_bal = min(d.min_bal, balance)
            peak      = max(peak, balance)
            max_dd    = max(max_dd, round(peak - balance, 2))

            if wins >= W_NEEDED:
                sessions_won += 1
                break

        sessions_total += 1
        d.sessions     += 1
        d.end_bal       = round(balance, 2)
        d.sessions_pnl.append(round(session_pnl, 2))

        if d.pnl >= META_DIARIA and not d.meta_hit:
            d.meta_hit = True

    days_meta  = sum(1 for d in daily.values() if d.meta_hit)
    daily_pnls = [round(d.pnl, 2) for d in sorted(daily.values(), key=lambda x: x.start_bal)]

    return SimResult(
        capital=capital, cap_pct=cap_pct,
        final=round(balance, 2), peak=round(peak, 2),
        min_bal=round(min_bal, 2), max_dd=round(max_dd, 2),
        sessions_total=sessions_total, sessions_won=sessions_won,
        total_sigs=total_sigs, days_meta=days_meta,
        days_total=len(daily), daily=daily, daily_pnls=daily_pnls,
    )


# ═══════════════════════════════════════════════════════════════════════════
# 4. HELPERS EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def fill(h: str) -> PatternFill:
    return PatternFill("solid", fgColor=h)

def thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=True) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def write_header(ws, row: int, cols: list[str], bg=AZUL_H, height=28):
    for c, t in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=t)
        cell.fill = fill(bg); cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.alignment = center(); cell.border = thin_border()
    ws.row_dimensions[row].height = height

def style_row(ws, row, ncols, even, bg_override=None):
    bg = bg_override or (GRIS if even else BLANCO)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        if not bg_override:
            cell.fill = fill(bg)
        cell.border = thin_border()
        cell.font   = Font(size=10)
        cell.alignment = center()

def money(v: float) -> str:
    return f"${v:,.2f}"

def pct_str(v: float) -> str:
    return f"{v:.1f}%"

def _color_val(v: float, inverse=False) -> str:
    if inverse:
        return VERDE_O if v < 0 else (ROJO_O if v > 0 else "000000")
    return VERDE_O if v > 0 else (ROJO_O if v < 0 else "000000")


# ═══════════════════════════════════════════════════════════════════════════
# 5. HOJA 1 — RESUMEN POR CAPITAL
# ═══════════════════════════════════════════════════════════════════════════

def sheet_resumen(wb, results: dict[int, SimResult], outcomes: list[Outcome]):
    ws = wb.active
    ws.title = "Resumen por Capital"
    ws.sheet_view.showGridLines = False

    col_widths = [14, 12, 12, 13, 13, 13, 14, 14, 12, 16, 18]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Título
    ncols = len(col_widths)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = "¿CUÁNTO CAPITAL NECESITO PARA GANAR $60/DÍA? — MASANIELLO 12/4 CON CAP TOTAL 10%"
    t.fill = fill(AZUL_H); t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = center(); ws.row_dimensions[1].height = 36

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    sub = ws["A2"]
    sub.value = (f"Datos reales: 2,655 señales — 52 días — WR={87.87:.2f}%  |  "
                 "Sistema: Masaniello 12/4, payout 92%, cap TOTAL ≤10% por señal  |  Meta diaria: $60")
    sub.fill = fill(AZUL_M); sub.font = Font(size=10, color="FFFFFF")
    sub.alignment = center(); ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    cols = [
        "Capital inicial", "Balance final", "ROI total", "Max DrawDown ($)",
        "Max DD (%)", "Entry 0L/0W", "Pérd. máx/señal", "Meta $60/día",
        "Días meta / total", "P&L promedio/día", "Evaluación"
    ]
    write_header(ws, 4, cols)

    for r, cap in enumerate(CAPITALES, 5):
        res = results[cap]
        even = (r % 2 == 0)
        roi  = (res.final - cap) / cap * 100
        dd_pct = res.max_dd / cap * 100
        entry_0 = round(cap * 0.10 / _TOTAL_MULT, 2)
        loss_max = round(cap * 0.10, 2)
        pnl_avg = round((res.final - cap) / res.days_total, 2) if res.days_total else 0
        meta_pct = res.days_meta / res.days_total * 100 if res.days_total else 0

        # Evaluación
        if meta_pct >= 75:
            eval_txt = "✅ EXCELENTE"
            bg_eval  = VERDE
        elif meta_pct >= 50:
            eval_txt = "⚠ BUENO"
            bg_eval  = AMARILLO
        elif meta_pct >= 25:
            eval_txt = "📊 MARGINAL"
            bg_eval  = NARAN
        else:
            eval_txt = "❌ INSUFICIENTE"
            bg_eval  = ROJO

        bg = GRIS if even else BLANCO

        data = [
            money(cap), money(res.final), pct_str(roi), money(res.max_dd),
            pct_str(dd_pct), money(entry_0), money(loss_max),
            f"{res.days_meta}/{res.days_total} días ({meta_pct:.0f}%)",
            f"{meta_pct:.0f}%", money(pnl_avg), eval_txt
        ]

        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border    = thin_border()
            cell.alignment = center()
            cell.font      = Font(size=10)
            if c == 11:
                cell.fill = fill(bg_eval)
                cell.font = Font(bold=True, size=10)
            else:
                cell.fill = fill(bg)
            # colorear ROI y P&L promedio
            if c == 3:
                cell.font = Font(size=10, bold=True, color=_color_val(roi))
            if c == 10:
                cell.font = Font(size=10, bold=True, color=_color_val(pnl_avg))
        ws.row_dimensions[r].height = 20

    # Nota matemática debajo
    last_data = 5 + len(CAPITALES)
    ws.row_dimensions[last_data].height = 10
    ws.merge_cells(f"A{last_data+1}:{get_column_letter(ncols)}{last_data+1}")
    nota = ws[f"A{last_data+1}"]
    nota.value = (
        "📐 MATEMÁTICA: WR real = 87.87% | Breakeven WR = 89.0% | "
        "EV/señal = $-0.38 (con cap 10%, capital $300) → el sistema necesita $60/día de meta por volumen de señales, "
        "no por EV positivo. Con WR=87.87% estás 1.13 puntos por debajo del breakeven matemático. "
        "La clave es aumentar capital para que las ganancias absolutas superen la pérdida esperada."
    )
    nota.fill = fill(AMARILLO); nota.font = Font(italic=True, size=9)
    nota.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[last_data+1].height = 50

    # ── Gráfica: días meta alcanzada por capital ──────────────────────────
    # Crear mini-tabla para la gráfica
    chart_row_start = last_data + 3
    ws.cell(row=chart_row_start, column=1, value="Capital").fill = fill(AZUL_H)
    ws.cell(row=chart_row_start, column=1).font = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=chart_row_start, column=2, value="% Días con Meta $60").fill = fill(AZUL_H)
    ws.cell(row=chart_row_start, column=2).font = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=chart_row_start, column=3, value="P&L diario promedio ($)").fill = fill(AZUL_H)
    ws.cell(row=chart_row_start, column=3).font = Font(bold=True, color="FFFFFF", size=10)

    for i, cap in enumerate(CAPITALES):
        res = results[cap]
        meta_pct = res.days_meta / res.days_total * 100 if res.days_total else 0
        pnl_avg = round((res.final - cap) / res.days_total, 2) if res.days_total else 0
        r = chart_row_start + 1 + i
        ws.cell(row=r, column=1, value=cap)
        ws.cell(row=r, column=2, value=round(meta_pct, 1))
        ws.cell(row=r, column=3, value=pnl_avg)

    n_caps = len(CAPITALES)
    chart = BarChart()
    chart.type  = "col"
    chart.title = "% Días alcanzando Meta $60 según Capital"
    chart.style = 10
    chart.y_axis.title = "% de días"
    chart.x_axis.title = "Capital ($)"
    chart.height = 14; chart.width = 22

    data_ref = Reference(ws, min_col=2, min_row=chart_row_start,
                         max_row=chart_row_start + n_caps)
    cats_ref = Reference(ws, min_col=1, min_row=chart_row_start + 1,
                         max_row=chart_row_start + n_caps)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = AZUL_M
    ws.add_chart(chart, f"E{last_data+3}")


# ═══════════════════════════════════════════════════════════════════════════
# 6. HOJA 2 — SENSIBILIDAD CAP %
# ═══════════════════════════════════════════════════════════════════════════

def sheet_sensibilidad(wb, outcomes: list[Outcome]):
    ws = wb.create_sheet("Sensibilidad Cap%")
    ws.sheet_view.showGridLines = False

    col_widths = [12, 12, 13, 14, 13, 13, 12, 14, 14, 18]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ncols = len(col_widths)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = "SENSIBILIDAD: ¿Qué pasa si cambiamos el CAP%? — Capital fijo $300"
    t.fill = fill(AZUL_H); t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = center(); ws.row_dimensions[1].height = 32

    cols = ["Cap %", "Entry 0L/0W", "Pérd. máx/señal", "Balance final",
            "ROI total", "Max DD ($)", "Max DD (%)", "Días meta $60",
            "P&L prom/día", "Evaluación riesgo"]
    write_header(ws, 2, cols)

    for r, cap_pct in enumerate(CAP_PCTS, 3):
        res = simulate(outcomes, 300.0, cap_pct)
        even = (r % 2 == 0)
        roi  = (res.final - 300) / 300 * 100
        dd_pct = res.max_dd / 300 * 100
        entry_0 = round(300 * cap_pct / _TOTAL_MULT, 2)
        loss_max = round(300 * cap_pct, 2)
        pnl_avg = round((res.final - 300) / res.days_total, 2) if res.days_total else 0
        meta_pct = res.days_meta / res.days_total * 100 if res.days_total else 0

        # Riesgo
        if cap_pct <= 0.10:
            riesgo = "🟢 SEGURO"
            bg_r   = VERDE
        elif cap_pct <= 0.20:
            riesgo = "🟡 MODERADO"
            bg_r   = AMARILLO
        else:
            riesgo = "🔴 ALTO"
            bg_r   = ROJO

        bg = GRIS if even else BLANCO

        data = [
            pct_str(cap_pct * 100), money(entry_0), money(loss_max),
            money(res.final), pct_str(roi), money(res.max_dd),
            pct_str(dd_pct),
            f"{res.days_meta}/{res.days_total} ({meta_pct:.0f}%)",
            money(pnl_avg), riesgo
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border(); cell.alignment = center()
            cell.font   = Font(size=10)
            cell.fill   = fill(bg) if c != ncols else fill(bg_r)
            if c == 5:
                cell.font = Font(size=10, bold=True, color=_color_val(roi))
            if c == 9:
                cell.font = Font(size=10, bold=True, color=_color_val(pnl_avg))
        ws.row_dimensions[r].height = 20

    last = 3 + len(CAP_PCTS)
    ws.row_dimensions[last].height = 8
    ws.merge_cells(f"A{last+1}:{get_column_letter(ncols)}{last+1}")
    n = ws[f"A{last+1}"]
    n.value = ("⚠ Con $300 de capital, NINGÚN cap% garantiza $60/día de forma consistente. "
               "La razón: WR=87.87% está 1.13 puntos por debajo del breakeven matemático (89%). "
               "Cap más alto = más ganancias por señal pero también más riesgo; el sistema se vuelve inestable.")
    n.fill = fill(AMARILLO); n.font = Font(italic=True, size=9); n.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[last+1].height = 40


# ═══════════════════════════════════════════════════════════════════════════
# 7. HOJA 3 — MAPA DE CALOR capital × cap_pct
# ═══════════════════════════════════════════════════════════════════════════

def sheet_mapa_calor(wb, outcomes: list[Outcome]):
    ws = wb.create_sheet("Mapa de Calor")
    ws.sheet_view.showGridLines = False

    # Capitales del eje Y
    caps_heat = [300, 500, 750, 1000, 1500, 2000, 3000, 5000]
    cap_pcts_heat = [0.05, 0.10, 0.15, 0.20, 0.25, 0.30]

    ncols = len(cap_pcts_heat) + 1
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = "MAPA DE CALOR — % días alcanzando Meta $60  (Capital × Cap%)"
    t.fill = fill(AZUL_H); t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = center(); ws.row_dimensions[1].height = 32

    # Cabecera cap_pct
    ws.cell(row=2, column=1, value="Capital \\ Cap%").fill = fill(AZUL_M)
    ws.cell(row=2, column=1).font = Font(bold=True, size=10, color="FFFFFF")
    ws.cell(row=2, column=1).alignment = center()
    ws.cell(row=2, column=1).border = thin_border()
    ws.column_dimensions["A"].width = 12
    for c, cp in enumerate(cap_pcts_heat, 2):
        cell = ws.cell(row=2, column=c, value=f"{cp*100:.0f}%")
        cell.fill = fill(AZUL_M); cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.alignment = center(); cell.border = thin_border()
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.row_dimensions[2].height = 24

    print("  Calculando mapa de calor (esto tarda un momento)...")
    for r, cap in enumerate(caps_heat, 3):
        ws.cell(row=r, column=1, value=f"${cap:,}").fill = fill(AZUL_M)
        ws.cell(row=r, column=1).font = Font(bold=True, size=10, color="FFFFFF")
        ws.cell(row=r, column=1).alignment = center()
        ws.cell(row=r, column=1).border = thin_border()
        for c, cp in enumerate(cap_pcts_heat, 2):
            res = simulate(outcomes, float(cap), cp)
            meta_pct = res.days_meta / res.days_total * 100 if res.days_total else 0
            pnl_avg  = round((res.final - cap) / res.days_total, 2) if res.days_total else 0

            # Color del mapa
            if meta_pct >= 75:
                bg = "1E6F4B"  # verde oscuro
                fg = "FFFFFF"
            elif meta_pct >= 50:
                bg = VERDE
                fg = VERDE_O
            elif meta_pct >= 25:
                bg = AMARILLO
                fg = "806000"
            elif meta_pct >= 10:
                bg = NARAN
                fg = NARANJA
            else:
                bg = ROJO
                fg = ROJO_O

            cell = ws.cell(row=r, column=c, value=f"{meta_pct:.0f}%\n${pnl_avg:+.1f}/d")
            cell.fill = fill(bg)
            cell.font = Font(bold=True, size=9, color=fg)
            cell.alignment = center()
            cell.border = thin_border()
            print(f"    ${cap:,} × {cp*100:.0f}% → {meta_pct:.0f}% días meta | pnl_avg ${pnl_avg:+.2f}")
        ws.row_dimensions[r].height = 28

    # Leyenda
    last = 3 + len(caps_heat)
    ws.row_dimensions[last].height = 8
    for label, bg_c, fg_c in [
        ("≥75% días — Excelente", "1E6F4B", "FFFFFF"),
        ("50–74% días — Bueno", VERDE, VERDE_O),
        ("25–49% días — Marginal", AMARILLO, "806000"),
        ("10–24% días — Bajo", NARAN, NARANJA),
        ("<10% días — Insuficiente", ROJO, ROJO_O),
    ]:
        last += 1
        ws.merge_cells(f"A{last}:{get_column_letter(ncols)}{last}")
        cell = ws[f"A{last}"]
        cell.value = label; cell.fill = fill(bg_c)
        cell.font = Font(bold=True, size=9, color=fg_c)
        cell.alignment = Alignment(horizontal="left")
        ws.row_dimensions[last].height = 18


# ═══════════════════════════════════════════════════════════════════════════
# 8. HOJA 4 — ANÁLISIS MATEMÁTICO
# ═══════════════════════════════════════════════════════════════════════════

def sheet_matematica(wb, outcomes: list[Outcome]):
    ws = wb.create_sheet("Analisis Matematico")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "ANÁLISIS MATEMÁTICO — Masaniello 12/4, WR=87.87%, Payout=92%"
    t.fill = fill(AZUL_H); t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = center(); ws.row_dimensions[1].height = 32

    # Estadísticas del canal
    wd = g1 = g2 = l = 0
    for o in outcomes:
        if o.result == "WD": wd += 1
        elif o.result == "G1": g1 += 1
        elif o.result == "G2": g2 += 1
        else: l += 1
    total = len(outcomes)
    wr = (wd + g1 + g2) / total

    write_header(ws, 3, ["Dato del canal", "Valor", "Porcentaje", "Observación"], bg=AZUL_M)

    canal_data = [
        ("Total señales analizadas", total, "—", "52 días"),
        ("VICTORIA DIRECTA (WD)", wd, f"{wd/total*100:.1f}%", "Gana sin gales"),
        ("VICTORIA en G1", g1, f"{g1/total*100:.1f}%", "Pierde entrada, gana G1"),
        ("VICTORIA en G2", g2, f"{g2/total*100:.1f}%", "Pierde entrada+G1, gana G2"),
        ("PÉRDIDA TOTAL (L)", l, f"{l/total*100:.1f}%", "Pierde los 3 intentos"),
        ("Win Rate REAL", f"{wr*100:.2f}%", "—", "Señales que ganan en algún intento"),
        ("Breakeven WR necesario", "89.00%", "—", "WR mínimo para EV≥0 con cap 10%"),
        ("Diferencia al breakeven", f"{(wr-0.89)*100:.2f} pp", "—", "Negativo = EV negativo por señal"),
    ]

    for r, (label, val, pct_v, obs) in enumerate(canal_data, 4):
        even = (r % 2 == 0)
        bg = GRIS if even else BLANCO
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = fill(bg)
            ws.cell(row=r, column=c).border = thin_border()
            ws.cell(row=r, column=c).alignment = center()
            ws.cell(row=r, column=c).font = Font(size=10)
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=3, value=pct_v)
        ws.cell(row=r, column=4, value=obs)
        if label.startswith("Win Rate"):
            ws.cell(row=r, column=2).font = Font(bold=True, size=10, color=VERDE_O)
        if label.startswith("Breakeven"):
            ws.cell(row=r, column=2).font = Font(bold=True, size=10, color=ROJO_O)
        if label.startswith("Diferencia"):
            ws.cell(row=r, column=2).font = Font(bold=True, size=10, color=ROJO_O)
        ws.row_dimensions[r].height = 20

    # EV por señal a distintos capitales
    ws.row_dimensions[13].height = 12
    write_header(ws, 14, ["Capital", "Entry 0L/0W", "Ganancia/win", "Pérdida/L (total)"], bg=AZUL_M)

    for r, cap in enumerate([100, 200, 300, 500, 1000, 2000, 3000, 5000], 15):
        entry = round(cap * 0.10 / _TOTAL_MULT, 2)
        gain  = round(entry * NET_PAYOUT, 2)
        loss  = round(cap * 0.10, 2)
        ev    = round(wr * gain - (1 - wr) * loss, 2)
        even  = (r % 2 == 0)
        bg    = GRIS if even else BLANCO

        data = [f"${cap:,}", f"${entry:.2f}", f"${gain:.2f}", f"${loss:.2f}"]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill(bg); cell.border = thin_border()
            cell.alignment = center(); cell.font = Font(size=10)
        ws.row_dimensions[r].height = 18

    # Fórmula de cuántas señales necesita para $60/día
    ws.row_dimensions[24].height = 12
    ws.merge_cells("A25:D25")
    f25 = ws["A25"]
    f25.value = "FÓRMULA: Señales necesarias para $60/día promedio"
    f25.fill = fill(AZUL_M); f25.font = Font(bold=True, size=11, color="FFFFFF")
    f25.alignment = center(); ws.row_dimensions[25].height = 24

    ws.merge_cells("A26:D26")
    fm = ws["A26"]
    fm.value = (
        "señales_necesarias = $60 / (WR × ganancia_por_win − (1−WR) × pérdida_por_L)\n"
        f"Con $300 cap 10%: señales = $60 / (0.8787×$3.71 − 0.1213×$30.00) = $60 / ($3.261 − $3.639) = $60 / (−$0.378) → IMPOSIBLE\n"
        f"Con $300 cap 10% necesitas WR≥89% O más capital para que EV positivo permita escalar la ganancia."
    )
    fm.fill = fill(ROJO); fm.font = Font(italic=True, size=10, color=ROJO_O)
    fm.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[26].height = 60


# ═══════════════════════════════════════════════════════════════════════════
# 9. HOJA 5 — DETALLE CAPITAL ÓPTIMO
# ═══════════════════════════════════════════════════════════════════════════

def sheet_detalle_optimo(wb, results: dict[int, SimResult]):
    # Encontrar el capital que más días cumple la meta
    best_cap = max(CAPITALES, key=lambda c: results[c].days_meta)
    res = results[best_cap]

    ws = wb.create_sheet(f"Detalle ${best_cap:,}")
    ws.sheet_view.showGridLines = False

    col_widths = [12, 9, 12, 10, 12, 12, 12, 14]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    meta_pct = res.days_meta / res.days_total * 100
    ws.merge_cells(f"A1:{get_column_letter(len(col_widths))}1")
    t = ws["A1"]
    t.value = (f"DETALLE DÍA A DÍA — Capital ${best_cap:,}  |  "
               f"Meta $60: {res.days_meta}/{res.days_total} días ({meta_pct:.0f}%)")
    t.fill = fill(AZUL_H); t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = center(); ws.row_dimensions[1].height = 32

    cols = ["Fecha", "Sesiones", "P&L día", "Meta $60",
            "Bal. Inicio", "Bal. Fin", "Bal. Min", "DD intradia"]
    write_header(ws, 2, cols)

    dias = sorted(res.daily.keys(), key=lambda d: datetime.strptime(d, "%d/%m/%Y"))
    for r, day in enumerate(dias, 3):
        d    = res.daily[day]
        even = (r % 2 == 0)
        bg   = VERDE if d.meta_hit else (ROJO if d.pnl < 0 else (GRIS if even else BLANCO))
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).fill = fill(bg)
            ws.cell(row=r, column=c).border = thin_border()
            ws.cell(row=r, column=c).alignment = center()
            ws.cell(row=r, column=c).font = Font(size=10)

        dd_int = round((d.start_bal or 0) - (d.min_bal or d.start_bal or 0), 2)
        data = [day, d.sessions, d.pnl, "✔ META" if d.meta_hit else "✖",
                d.start_bal, d.end_bal, d.min_bal, dd_int]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (3, 5, 6, 7, 8):
                cell.number_format = '"$"#,##0.00'
            if c == 3:
                cell.font = Font(size=10, bold=True, color=VERDE_O if d.pnl >= 0 else ROJO_O)
            if c == 4:
                cell.font = Font(size=10, bold=True, color=VERDE_O if d.meta_hit else ROJO_O)
        ws.row_dimensions[r].height = 18

    # Gráfica balance dia a dia
    n_rows = len(dias)
    chart = LineChart()
    chart.title = f"Balance diario — Capital ${best_cap:,}"
    chart.style = 10; chart.height = 14; chart.width = 24
    chart.y_axis.title = "Balance ($)"

    ref_bal = Reference(ws, min_col=6, min_row=2, max_row=2 + n_rows)
    chart.add_data(ref_bal, titles_from_data=True)
    chart.series[0].graphicalProperties.line.solidFill = AZUL_M
    chart.series[0].graphicalProperties.line.width = 20000
    cats = Reference(ws, min_col=1, min_row=3, max_row=2 + n_rows)
    chart.set_categories(cats)
    ws.add_chart(chart, f"J3")
    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main() -> None:
    print("Leyendo histórico Telegram...")
    outcomes = parse_outcomes(HISTORY)
    wd_c = sum(1 for o in outcomes if o.result == "WD")
    g1_c = sum(1 for o in outcomes if o.result == "G1")
    g2_c = sum(1 for o in outcomes if o.result == "G2")
    l_c  = sum(1 for o in outcomes if o.result == "L")
    wr   = (wd_c + g1_c + g2_c) / len(outcomes) * 100
    print(f"  {len(outcomes)} señales  WD={wd_c} G1={g1_c} G2={g2_c} L={l_c}  WR={wr:.2f}%")

    # Simular todos los capitales con cap 10%
    print("\nSimulando por niveles de capital (cap total 10%)...")
    results: dict[int, SimResult] = {}
    for cap in CAPITALES:
        r = simulate(outcomes, float(cap), 0.10)
        meta_pct = r.days_meta / r.days_total * 100 if r.days_total else 0
        pnl_avg  = (r.final - cap) / r.days_total if r.days_total else 0
        print(f"  ${cap:>5,}  balance=${r.final:>9.2f}  meta={r.days_meta:2d}/{r.days_total}d ({meta_pct:3.0f}%)  pnl_avg=${pnl_avg:+.2f}/d")
        results[cap] = r

    print("\nGenerando Excel...")
    wb = openpyxl.Workbook()

    sheet_resumen(wb, results, outcomes)
    sheet_sensibilidad(wb, outcomes)
    sheet_mapa_calor(wb, outcomes)
    sheet_matematica(wb, outcomes)
    sheet_detalle_optimo(wb, results)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\n✔  Reporte guardado en: {OUTPUT}")

    print()
    print("=" * 65)
    print("VEREDICTO FINAL")
    print("=" * 65)
    best = max(results.values(), key=lambda r: r.days_meta)
    best_meta_pct = best.days_meta / best.days_total * 100
    print(f"  Mejor capital simulado: ${best.capital:,.0f}")
    print(f"  Meta $60 alcanzada: {best.days_meta}/{best.days_total} días ({best_meta_pct:.0f}%)")
    print()
    print(f"  WR del canal = {wr:.2f}%  |  Breakeven = 89.00%")
    print(f"  Déficit = {89-wr:.2f} puntos porcentuales")
    print()
    print("  Conclusión: Con el cap total seguro (10%), el sistema es EV-negativo.")
    print("  La meta $60/día se alcanza por volumen y varianza, no por EV positivo.")
    print("  Para garantizar la meta: WR real ≥ 89% O aceptar cap > 10% (más riesgo).")


if __name__ == "__main__":
    main()
