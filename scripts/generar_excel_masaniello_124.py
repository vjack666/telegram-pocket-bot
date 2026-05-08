"""
Informe Excel completo — Masaniello 12/4 sobre datos reales del canal Telegram.
Simula el nuevo sistema con los 2,652 resultados del historico (17/03/2026 - 07/05/2026).

Hojas generadas:
  1. Resumen Ejecutivo       — KPIs globales + tabla comparativa 6/2 vs 12/4
  2. Sesiones Detalladas     — cada sesion: stakes, resultados, P&L
  3. Resumen Diario          — por dia: balance, meta, sessions, DD intradia
  4. Evolucion del Balance   — grafica de balance acumulado dia a dia
  5. Analisis de Rachas      — distribucion de rachas de perdidas y su impacto
  6. Tabla de Stakes 12/4    — referencia de stake por estado L/W
"""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

HISTORY = ROOT / "ejemplo.md"
OUTPUT  = ROOT / "runtime" / "informe_masaniello_124.xlsx"

# ── Configuracion Masaniello 12/4 ──────────────────────────────────────────
N_OPS          = 12
W_NEEDED       = 4
PAYOUT_MULT    = 1.92
BASE_BALANCE   = 300.0
CAPITAL        = 300.0
META_DIARIA    = 60.0
CAP_PCT        = 0.10

# ── Configuracion Masaniello 6/2 (referencia) ──────────────────────────────
N_OPS_62  = 6
W_NEEDED_62 = 2

# ── Colores ────────────────────────────────────────────────────────────────
VERDE_OSCURO  = "1E6F4B"
VERDE_CLARO   = "D6F0E0"
ROJO_OSCURO   = "8B1A1A"
ROJO_CLARO    = "FDDEDE"
AZUL_HEADER   = "1F3864"
AZUL_MEDIO    = "2E75B6"
AZUL_CLARO    = "D6E4F7"
GRIS_CLARO    = "F2F2F2"
GRIS_HEADER   = "D9D9D9"
AMARILLO      = "FFF2CC"
NARANJA       = "FF8C00"
BLANCO        = "FFFFFF"


# ═══════════════════════════════════════════════════════════════════════════
# 1. PARSER DE RESULTADOS
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class Outcome:
    timestamp: datetime
    result: str       # "WD", "G1", "G2", "L"
    is_win: bool


def parse_outcomes(path: Path) -> list[Outcome]:
    date_pat = re.compile(r"^\[(\d{2}/\d{2}/\d{4}) (\d{2}:\d{2}:\d{2})\]")
    res_pat  = re.compile(
        r"(VICTORIA DIRECTA|VICTORIA EN 1.*?MARTINGALA|VICTORIA EN 2.*?MARTINGALA|P[ÉE]RDIDA)",
        re.IGNORECASE,
    )

    label_map = {
        "victoria directa": "WD",
        "victoria en 1": "G1",
        "victoria en 2": "G2",
        "perdida": "L",
        "pérdida": "L",
    }

    outcomes: list[Outcome] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        ts_m  = date_pat.match(line)
        res_m = res_pat.search(line)
        if not ts_m or not res_m:
            continue
        ts = datetime.strptime(f"{ts_m.group(1)} {ts_m.group(2)}", "%d/%m/%Y %H:%M:%S")
        raw = res_m.group(1).lower()
        label = next((v for k, v in label_map.items() if raw.startswith(k)), "?")
        is_win = label != "L"
        outcomes.append(Outcome(ts, label, is_win))
    return outcomes


# ═══════════════════════════════════════════════════════════════════════════
# 2. FORMULA MASANIELLO
# ═══════════════════════════════════════════════════════════════════════════

def _fwd_prob(ops_left: int, wins_needed: int, pm: float) -> float:
    if wins_needed <= 0:   return 1.0
    if wins_needed > ops_left: return 0.0
    if wins_needed == ops_left: return pm ** ops_left
    pw = _fwd_prob(ops_left - 1, wins_needed - 1, pm)
    pl = _fwd_prob(ops_left - 1, wins_needed,     pm)
    d  = pw + (pm - 1) * pl
    return (pm * pw * pl / d) if d else 0.0


def masaniello_stake(balance: float, losses: int, wins: int, n: int, w: int, pm: float) -> float:
    ops_left  = n - (losses + wins)
    wins_left = w - wins
    if ops_left <= 0 or wins_left <= 0 or wins_left > ops_left:
        return 0.0
    pw = _fwd_prob(ops_left - 1, wins_left - 1, pm)
    pl = _fwd_prob(ops_left - 1, wins_left,     pm)
    d  = pw + (pm - 1) * pl
    if not d:
        return balance
    s = balance * (1 - pm * pw / d)
    return round(max(0.01, min(s, balance)), 2)


def gale_amounts(entry: float, pm: float, balance: float, cap: float) -> tuple[float, float, float]:
    net = pm - 1.0          # 0.92
    g1  = round(min(entry * pm / net, cap, balance), 2)
    g2  = round(min(entry * (pm / net) ** 2, cap, balance), 2)
    return entry, g1, g2


# ═══════════════════════════════════════════════════════════════════════════
# 3. SIMULACION
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class SessionRow:
    index:      int
    day:        str
    hora:       str
    wins:       int
    losses:     int
    ops_used:   int
    resultados: list[str]
    stakes:     list[float]
    pnl:        float
    balance_pre:  float
    balance_post: float
    meta_hit:   bool
    acum_pnl_dia: float = 0.0


def simulate(outcomes: list[Outcome], n: int, w: int) -> tuple[list[SessionRow], dict, dict]:
    sessions_raw = [outcomes[i:i + n] for i in range(0, len(outcomes), n)]

    balance   = CAPITAL
    peak      = balance
    max_dd    = 0.0
    min_bal   = balance
    all_sessions: list[SessionRow] = []
    daily: dict[str, dict] = defaultdict(lambda: {
        "start_bal": None, "end_bal": None, "min_bal": None,
        "pnl": 0.0, "sessions": 0, "meta_hit": False, "meta_time": None,
        "session_pnls": [], "ops_used": 0,
    })

    for idx, chunk in enumerate(sessions_raw, 1):
        if not chunk:
            continue
        day  = chunk[0].timestamp.strftime("%d/%m/%Y")
        hora = chunk[0].timestamp.strftime("%H:%M:%S")
        row  = daily[day]
        if row["start_bal"] is None:
            row["start_bal"] = round(balance, 2)
            row["min_bal"]   = round(balance, 2)

        if row["meta_hit"]:
            continue

        wins = losses = 0
        pnl  = 0.0
        results: list[str]  = []
        stakes:  list[float] = []
        bal_pre = round(balance, 2)

        for outcome in chunk:
            wins_needed = w - wins
            ops_left    = n - (wins + losses)
            if wins_needed <= 0 or wins_needed > ops_left:
                break

            # Cap aplicado sobre EXPOSICIÓN TOTAL (entry + G1 + G2)
            # para que el riesgo máximo por señal sea ≤ CAP_PCT del capital.
            # Multiplicador total: 1 + pm/(pm-1) + (pm/(pm-1))^2
            pm = PAYOUT_MULT
            _r  = pm / (pm - 1)                          # ratio de gale ≈ 2.087
            _total_mult = 1 + _r + _r ** 2               # ≈ 7.442
            cap_total = round(balance * CAP_PCT, 2)       # 10% de la cuenta
            entry_max = round(cap_total / _total_mult, 2) # entry que hace total = cap_total

            entry = masaniello_stake(BASE_BALANCE, losses, wins, n, w, PAYOUT_MULT)
            entry = min(entry, entry_max)                 # cap real por señal

            # Gales proporcionales al entry real (no capeados individualmente)
            g1_amount = round(entry * _r, 2)
            g2_amount = round(entry * _r ** 2, 2)
            total_loss = round(entry + g1_amount + g2_amount, 2)

            results.append(outcome.result)
            stakes.append(entry)

            if outcome.is_win:
                wins  += 1
                # La ganancia neta es siempre entry * (pm-1) gracias a Masaniello
                gain   = round(entry * (PAYOUT_MULT - 1), 2)
                pnl   += gain
                balance = round(balance + gain, 2)
            else:
                losses += 1
                # Pérdida real: se agotaron entry + G1 + G2
                pnl    -= total_loss
                balance = round(balance - total_loss, 2)

            min_bal = min(min_bal, balance)
            row["min_bal"] = min(row["min_bal"], balance)
            row["ops_used"] += 1

            peak   = max(peak, balance)
            dd     = round(peak - balance, 2)
            max_dd = max(max_dd, dd)

            if wins >= w:
                break

        pnl = round(pnl, 2)
        row["sessions"] += 1
        row["pnl"]       = round(row["pnl"] + pnl, 2)
        row["session_pnls"].append(pnl)
        row["end_bal"]   = round(balance, 2)

        all_sessions.append(SessionRow(
            index=idx, day=day, hora=hora,
            wins=wins, losses=losses, ops_used=wins + losses,
            resultados=results, stakes=stakes, pnl=pnl,
            balance_pre=bal_pre, balance_post=round(balance, 2),
            meta_hit=(row["pnl"] >= META_DIARIA),
            acum_pnl_dia=row["pnl"],
        ))

        if row["pnl"] >= META_DIARIA and not row["meta_hit"]:
            row["meta_hit"]  = True
            row["meta_time"] = hora

    stats = {
        "final": round(balance, 2),
        "min":   round(min_bal, 2),
        "peak":  round(peak, 2),
        "max_dd": round(max_dd, 2),
        "total_ops": len(outcomes),
    }
    return all_sessions, daily, stats


# ═══════════════════════════════════════════════════════════════════════════
# 4. HELPERS EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def header_font(bold=True, size=11, color="FFFFFF"):
    return Font(bold=bold, size=size, color=color)

def cell_font(bold=False, size=10, color="000000"):
    return Font(bold=bold, size=size, color=color)

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")

def write_header(ws, row: int, cols: list[str], bg=AZUL_HEADER, fg="FFFFFF"):
    for c, title in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=title)
        cell.fill     = fill(bg)
        cell.font     = Font(bold=True, size=10, color=fg)
        cell.alignment = center()
        cell.border   = thin_border()

def style_data_row(ws, row: int, ncols: int, even: bool):
    bg = GRIS_CLARO if even else BLANCO
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill   = fill(bg)
        cell.border = thin_border()
        cell.font   = cell_font()

def money(v: float) -> str:
    return f"${v:,.2f}"

def pct(v: float) -> str:
    return f"{v:.1f}%"


# ═══════════════════════════════════════════════════════════════════════════
# 5. HOJA 1 — RESUMEN EJECUTIVO
# ═══════════════════════════════════════════════════════════════════════════

def sheet_resumen(wb, sessions_124, daily_124, stats_124,
                      sessions_62,  daily_62,  stats_62):
    ws = wb.active
    ws.title = "Resumen Ejecutivo"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14

    # ── Titulo ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "INFORME MASANIELLO 12/4 — DATOS REALES TELEGRAM"
    t.fill  = fill(AZUL_HEADER)
    t.font  = Font(bold=True, size=14, color="FFFFFF")
    t.alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = "Periodo: 17/03/2026 – 07/05/2026  |  Canal: VIP TRADER A + SMART SIGNALS  |  2,652 señales"
    sub.fill  = fill(AZUL_MEDIO)
    sub.font  = Font(bold=False, size=10, color="FFFFFF")
    sub.alignment = center()
    ws.row_dimensions[2].height = 20

    # ── KPI block ──────────────────────────────────────────────────────────
    def kpi_row(row, label, v124, v62, highlight=False):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = thin_border()
        for c, (val, bg) in enumerate([(v124, AZUL_CLARO), (v62, GRIS_CLARO)], 2):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill      = fill(bg)
            cell.alignment = center()
            cell.border    = thin_border()
            cell.font      = Font(bold=highlight, size=10,
                                  color=VERDE_OSCURO if highlight and c == 2 else
                                        ROJO_OSCURO if highlight and c == 3 else "000000")

    ws.row_dimensions[3].height = 6  # espacio

    write_header(ws, 4, ["Métrica", "Masaniello 12/4  (NUEVO)", "Masaniello 6/2  (anterior)", "Diferencia"])
    ws.row_dimensions[4].height = 28

    dias_124    = len(daily_124)
    meta_124    = sum(1 for d in daily_124.values() if d["meta_hit"])
    dias_62     = len(daily_62)
    meta_62     = sum(1 for d in daily_62.values()  if d["meta_hit"])

    wd_124 = sum(1 for s in sessions_124 if s.wins >= W_NEEDED)
    total_sigs_124 = sum(len(s.resultados) for s in sessions_124)
    win_sigs_124   = sum(1 for s in sessions_124 for r in s.resultados if r != "L")
    wr_124 = win_sigs_124 / total_sigs_124 * 100 if total_sigs_124 else 0

    wd_62 = sum(1 for s in sessions_62 if s.wins >= W_NEEDED_62)
    total_sigs_62 = sum(len(s.resultados) for s in sessions_62)
    win_sigs_62   = sum(1 for s in sessions_62 for r in s.resultados if r != "L")
    wr_62 = win_sigs_62 / total_sigs_62 * 100 if total_sigs_62 else 0

    _r = PAYOUT_MULT / (PAYOUT_MULT - 1)
    _total_mult = 1 + _r + _r ** 2   # ≈ 7.442

    # Entry real: masaniello_stake capeado para que (entry+G1+G2) ≤ 10% de cuenta
    cap_total_124 = BASE_BALANCE * CAP_PCT
    entry_max_124 = round(cap_total_124 / _total_mult, 2)
    stake_ini_124 = min(masaniello_stake(BASE_BALANCE, 0, 0, N_OPS, W_NEEDED, PAYOUT_MULT), entry_max_124)
    stake_max_124 = min(masaniello_stake(BASE_BALANCE, 3, 0, N_OPS, W_NEEDED, PAYOUT_MULT), entry_max_124)

    cap_total_62  = BASE_BALANCE * CAP_PCT
    entry_max_62  = round(cap_total_62 / _total_mult, 2)
    stake_ini_62  = min(masaniello_stake(BASE_BALANCE, 0, 0, N_OPS_62, W_NEEDED_62, PAYOUT_MULT), entry_max_62)
    stake_max_62  = min(masaniello_stake(BASE_BALANCE, 3, 0, N_OPS_62, W_NEEDED_62, PAYOUT_MULT), entry_max_62)

    rows_data = [
        ("Capital inicial",                   money(CAPITAL),               money(CAPITAL),               "—"),
        ("Balance final",                     money(stats_124["final"]),     money(stats_62["final"]),      money(stats_124["final"] - stats_62["final"])),
        ("Balance máximo alcanzado",          money(stats_124["peak"]),      money(stats_62["peak"]),       money(stats_124["peak"] - stats_62["peak"])),
        ("Balance mínimo observado",          money(stats_124["min"]),       money(stats_62["min"]),        money(stats_124["min"] - stats_62["min"])),
        ("DrawDown máximo ($)",               money(stats_124["max_dd"]),    money(stats_62["max_dd"]),     money(stats_124["max_dd"] - stats_62["max_dd"])),
        ("DrawDown máximo (%)",               pct(stats_124["max_dd"]/CAPITAL*100), pct(stats_62["max_dd"]/CAPITAL*100), "—"),
        ("Días analizados",                   str(dias_124),                 str(dias_62),                  "—"),
        ("Días con meta $60",                 f"{meta_124}/{dias_124}",      f"{meta_62}/{dias_62}",        f"{meta_124 - meta_62:+d}"),
        ("Sesiones totales operadas",         str(len(sessions_124)),        str(len(sessions_62)),         f"{len(sessions_124) - len(sessions_62):+d}"),
        ("Sesiones ganadas",                  str(wd_124),                   str(wd_62),                    f"{wd_124 - wd_62:+d}"),
        ("Señales consumidas",                str(total_sigs_124),           str(total_sigs_62),            f"{total_sigs_124 - total_sigs_62:+d}"),
        ("Win Rate real señales",             pct(wr_124),                   pct(wr_62),                    "—"),
        ("Entry 0L/0W (cap total aplicado)",  money(stake_ini_124),          money(stake_ini_62),           money(stake_ini_124 - stake_ini_62)),
        ("Entry max 3L/0W (cap total aplic.)",money(stake_max_124),          money(stake_max_62),           money(stake_max_124 - stake_max_62)),
        ("Exposición TOTAL máx (entry+G1+G2)",money(entry_max_124 * _total_mult), money(entry_max_62 * _total_mult), "—"),
        ("Cap 10% de la cuenta ($300)",       money(BASE_BALANCE * CAP_PCT), money(BASE_BALANCE * CAP_PCT), "—"),
    ]

    for r, (label, v124, v62, diff) in enumerate(rows_data, 5):
        even = (r % 2 == 0)
        ws.cell(row=r, column=1, value=label).border = thin_border()
        ws.cell(row=r, column=1).fill  = fill(GRIS_CLARO if even else BLANCO)
        ws.cell(row=r, column=1).font  = Font(bold=False, size=10)
        ws.cell(row=r, column=2, value=v124).fill  = fill(AZUL_CLARO)
        ws.cell(row=r, column=2).alignment = center()
        ws.cell(row=r, column=2).border = thin_border()
        ws.cell(row=r, column=2).font   = Font(bold=False, size=10, color=VERDE_OSCURO)
        ws.cell(row=r, column=3, value=v62).fill  = fill(GRIS_CLARO if even else BLANCO)
        ws.cell(row=r, column=3).alignment = center()
        ws.cell(row=r, column=3).border = thin_border()
        ws.cell(row=r, column=3).font   = Font(bold=False, size=10)
        ws.cell(row=r, column=4, value=diff).alignment = center()
        ws.cell(row=r, column=4).border = thin_border()
        is_num = isinstance(diff, str) and diff.startswith("$")
        try:
            val = float(diff.replace("$","").replace(",","").replace("+","")) if diff not in ("—","") else 0
            ws.cell(row=r, column=4).font = Font(
                bold=False, size=10,
                color=VERDE_OSCURO if val > 0 else (ROJO_OSCURO if val < 0 else "000000")
            )
        except Exception:
            pass
        ws.row_dimensions[r].height = 20

    # ── Nota al pie ────────────────────────────────────────────────────────
    last = 5 + len(rows_data)
    ws.row_dimensions[last].height = 10
    ws.merge_cells(f"A{last+1}:D{last+1}")
    note = ws[f"A{last+1}"]
    note.value = ("★  El cap 10% aplica sobre la EXPOSICIÓN TOTAL por señal (entry + G1 + G2 ≤ $30). "
                  "Entry máximo = $30 ÷ 7.44 ≈ $4.03 para que incluso si fallan los 3 intentos, "
                  "no se pierde más del 10% de la cuenta en una sola señal.")
    note.fill  = fill(AMARILLO)
    note.font  = Font(italic=True, size=9)
    note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[last+1].height = 32


# ═══════════════════════════════════════════════════════════════════════════
# 6. HOJA 2 — SESIONES DETALLADAS
# ═══════════════════════════════════════════════════════════════════════════

def sheet_sesiones(wb, sessions: list[SessionRow]):
    ws = wb.create_sheet("Sesiones Detalladas")
    ws.sheet_view.showGridLines = False

    cols = ["#","Fecha","Hora","Resultados","Stakes","Ops","Wins","Losses",
            "P&L sesión","Bal. Inicio","Bal. Fin","Meta día","P&L acum. día"]
    col_w = [5, 12, 9, 38, 34, 5, 5, 5, 12, 12, 12, 9, 14]
    for c, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    t = ws["A1"]
    t.value = "DETALLE DE SESIONES — MASANIELLO 12/4"
    t.fill  = fill(AZUL_HEADER)
    t.font  = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = center()
    ws.row_dimensions[1].height = 30

    write_header(ws, 2, cols)
    ws.row_dimensions[2].height = 28

    for r, s in enumerate(sessions, 3):
        even = (r % 2 == 0)
        style_data_row(ws, r, len(cols), even)
        resultados_str = "  ".join(s.resultados)
        stakes_str     = "  ".join(f"${x:.2f}" for x in s.stakes)

        data = [
            s.index, s.day, s.hora, resultados_str, stakes_str,
            s.ops_used, s.wins, s.losses, s.pnl,
            s.balance_pre, s.balance_post,
            "✔" if s.meta_hit else "—",
            s.acum_pnl_dia,
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = center() if c <= 3 or c in (6,7,8,12) else right()
            if c == 9:   # P&L sesion
                cell.number_format = '"$"#,##0.00'
                cell.font = Font(size=10, color=VERDE_OSCURO if s.pnl >= 0 else ROJO_OSCURO, bold=True)
            elif c in (10, 11, 13):
                cell.number_format = '"$"#,##0.00'
            if c == 12 and s.meta_hit:
                cell.font = Font(size=10, color=VERDE_OSCURO, bold=True)

        ws.row_dimensions[r].height = 18

    # Freeze header
    ws.freeze_panes = "A3"

    # Autofilter
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{2 + len(sessions)}"


# ═══════════════════════════════════════════════════════════════════════════
# 7. HOJA 3 — RESUMEN DIARIO
# ═══════════════════════════════════════════════════════════════════════════

def sheet_diario(wb, daily: dict, stats: dict):
    ws = wb.create_sheet("Resumen Diario")
    ws.sheet_view.showGridLines = False

    cols = ["Fecha","Sesiones","P&L día","Meta $60","Hora meta",
            "Bal. Inicio","Bal. Fin","Bal. Min","DD intradia","P&L/sesión avg"]
    col_w = [12, 9, 12, 9, 10, 13, 13, 13, 12, 14]
    for c, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    t = ws["A1"]
    t.value = "RESUMEN DIARIO — MASANIELLO 12/4"
    t.fill  = fill(AZUL_HEADER)
    t.font  = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = center()
    ws.row_dimensions[1].height = 30

    write_header(ws, 2, cols)
    ws.row_dimensions[2].height = 28

    dias = sorted(daily.keys(), key=lambda d: datetime.strptime(d, "%d/%m/%Y"))
    for r, day in enumerate(dias, 3):
        row  = daily[day]
        even = (r % 2 == 0)
        bg   = VERDE_CLARO if row["meta_hit"] else (ROJO_CLARO if row["pnl"] < 0 else (GRIS_CLARO if even else BLANCO))
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill   = fill(bg)
            cell.border = thin_border()
            cell.font   = Font(size=10)
            cell.alignment = center()

        dd_int = round((row["start_bal"] or 0) - (row["min_bal"] or row["start_bal"] or 0), 2)
        pnl_avg = round(row["pnl"] / row["sessions"], 2) if row["sessions"] else 0.0

        data = [
            day,
            row["sessions"],
            row["pnl"],
            "✔ META" if row["meta_hit"] else "✖",
            row["meta_time"] or "—",
            row["start_bal"],
            row["end_bal"],
            row["min_bal"],
            dd_int,
            pnl_avg,
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (3, 6, 7, 8, 9, 10):
                cell.number_format = '"$"#,##0.00'
            if c == 3:
                cell.font = Font(size=10, bold=True,
                                 color=VERDE_OSCURO if row["pnl"] >= 0 else ROJO_OSCURO)
            if c == 4:
                cell.font = Font(size=10, bold=True,
                                 color=VERDE_OSCURO if row["meta_hit"] else ROJO_OSCURO)
        ws.row_dimensions[r].height = 18

    # Totales
    last_r = 3 + len(dias)
    ws.row_dimensions[last_r].height = 6
    ws.row_dimensions[last_r + 1].height = 22
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=last_r + 1, column=c)
        cell.fill   = fill(AZUL_HEADER)
        cell.border = thin_border()
        cell.font   = Font(bold=True, size=10, color="FFFFFF")
        cell.alignment = center()

    meta_dias = sum(1 for d in daily.values() if d["meta_hit"])
    ws.cell(row=last_r+1, column=1, value="TOTALES / RESUMEN").alignment = center()
    ws.cell(row=last_r+1, column=2, value=sum(d["sessions"] for d in daily.values()))
    ws.cell(row=last_r+1, column=3, value=round(stats["final"] - CAPITAL, 2)).number_format = '"$"#,##0.00'
    ws.cell(row=last_r+1, column=4, value=f"{meta_dias}/{len(dias)} días")
    ws.cell(row=last_r+1, column=6, value=CAPITAL).number_format = '"$"#,##0.00'
    ws.cell(row=last_r+1, column=7, value=stats["final"]).number_format = '"$"#,##0.00'
    ws.cell(row=last_r+1, column=8, value=stats["min"]).number_format = '"$"#,##0.00'
    ws.cell(row=last_r+1, column=9, value=stats["max_dd"]).number_format = '"$"#,##0.00'

    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════════
# 8. HOJA 4 — EVOLUCIÓN DEL BALANCE (con gráfica)
# ═══════════════════════════════════════════════════════════════════════════

def sheet_balance(wb, daily_124: dict, daily_62: dict):
    ws = wb.create_sheet("Evolucion del Balance")
    ws.sheet_view.showGridLines = False

    cols = ["Fecha","Balance 12/4","Balance 6/2","DD dia 12/4","DD dia 6/2"]
    col_w = [12, 14, 14, 13, 13]
    for c, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    write_header(ws, 1, cols)
    ws.row_dimensions[1].height = 26

    dias = sorted(set(list(daily_124.keys()) + list(daily_62.keys())),
                  key=lambda d: datetime.strptime(d, "%d/%m/%Y"))

    bal_124 = CAPITAL
    bal_62  = CAPITAL

    for r, day in enumerate(dias, 2):
        even = (r % 2 == 0)
        d124 = daily_124.get(day)
        d62  = daily_62.get(day)

        if d124 and d124["end_bal"] is not None:
            bal_124 = d124["end_bal"]
        if d62  and d62["end_bal"]  is not None:
            bal_62  = d62["end_bal"]

        dd124 = round((d124["start_bal"] or bal_124) - (d124["min_bal"] or (d124["start_bal"] or bal_124)), 2) if d124 else 0.0
        dd62  = round((d62["start_bal"]  or bal_62)  - (d62["min_bal"]  or (d62["start_bal"] or bal_62)), 2)   if d62  else 0.0

        style_data_row(ws, r, len(cols), even)
        data = [day, bal_124, bal_62, dd124, dd62]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = center()
            if c in (2, 3, 4, 5):
                cell.number_format = '"$"#,##0.00'
        ws.row_dimensions[r].height = 16

    n_rows = len(dias)

    # ── Gráfica líneas: Balance 12/4 vs 6/2 ───────────────────────────────
    chart = LineChart()
    chart.title = "Evolución del Balance — 12/4 vs 6/2"
    chart.style = 10
    chart.y_axis.title = "Balance ($)"
    chart.x_axis.title = "Día"
    chart.height = 14
    chart.width  = 26

    ref124 = Reference(ws, min_col=2, min_row=1, max_row=1 + n_rows)
    ref62  = Reference(ws, min_col=3, min_row=1, max_row=1 + n_rows)
    chart.add_data(ref124, titles_from_data=True)
    chart.add_data(ref62,  titles_from_data=True)

    from openpyxl.chart.series import SeriesLabel
    chart.series[0].graphicalProperties.line.solidFill = "2E75B6"
    chart.series[0].graphicalProperties.line.width = 18000
    chart.series[1].graphicalProperties.line.solidFill = "FF8C00"
    chart.series[1].graphicalProperties.line.width = 15000

    dates_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + n_rows)
    from openpyxl.chart import Reference as Ref
    chart.set_categories(dates_ref)
    ws.add_chart(chart, f"G2")

    # ── Gráfica barras: DD diario ──────────────────────────────────────────
    chart2 = BarChart()
    chart2.type   = "col"
    chart2.title  = "Drawdown Intradia por Día"
    chart2.style  = 10
    chart2.y_axis.title = "Drawdown ($)"
    chart2.height = 14
    chart2.width  = 26

    refdd124 = Reference(ws, min_col=4, min_row=1, max_row=1 + n_rows)
    refdd62  = Reference(ws, min_col=5, min_row=1, max_row=1 + n_rows)
    chart2.add_data(refdd124, titles_from_data=True)
    chart2.add_data(refdd62,  titles_from_data=True)
    chart2.set_categories(dates_ref)
    chart2.series[0].graphicalProperties.solidFill = "2E75B6"
    chart2.series[1].graphicalProperties.solidFill = "FF8C00"
    ws.add_chart(chart2, f"G{2 + n_rows + 3}")


# ═══════════════════════════════════════════════════════════════════════════
# 9. HOJA 5 — ANÁLISIS DE RACHAS
# ═══════════════════════════════════════════════════════════════════════════

def sheet_rachas(wb, sessions: list[SessionRow]):
    ws = wb.create_sheet("Analisis de Rachas")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "ANÁLISIS DE RACHAS DE PÉRDIDAS — MASANIELLO 12/4"
    t.fill  = fill(AZUL_HEADER)
    t.font  = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = center()
    ws.row_dimensions[1].height = 30

    # ── Distribución de resultados por señal ─────────────────────────────
    total_sigs = sum(len(s.resultados) for s in sessions)
    cnt = {"WD": 0, "G1": 0, "G2": 0, "L": 0}
    for s in sessions:
        for r in s.resultados:
            if r in cnt:
                cnt[r] += 1

    write_header(ws, 3, ["Resultado", "Ocurrencias", "% del total", "Stake típico",
                          "Ganancia neta", "Pérdida neta"])

    tipo_datos = [
        ("WD  (victoria directa)",  cnt["WD"], masaniello_stake(BASE_BALANCE,0,0,N_OPS,W_NEEDED,PAYOUT_MULT),  True),
        ("G1  (victoria en gale 1)", cnt["G1"], masaniello_stake(BASE_BALANCE,0,0,N_OPS,W_NEEDED,PAYOUT_MULT) * PAYOUT_MULT/(PAYOUT_MULT-1), True),
        ("G2  (victoria en gale 2)", cnt["G2"], masaniello_stake(BASE_BALANCE,0,0,N_OPS,W_NEEDED,PAYOUT_MULT) * (PAYOUT_MULT/(PAYOUT_MULT-1))**2, True),
        ("L   (pérdida completa)",   cnt["L"],  0.0, False),
    ]

    stake0 = masaniello_stake(BASE_BALANCE, 0, 0, N_OPS, W_NEEDED, PAYOUT_MULT)

    for r, (label, occ, stake, is_w) in enumerate(tipo_datos, 4):
        even = (r % 2 == 0)
        bg   = VERDE_CLARO if is_w else ROJO_CLARO
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill   = fill(bg)
            ws.cell(row=r, column=c).border = thin_border()
            ws.cell(row=r, column=c).alignment = center()

        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=occ)
        ws.cell(row=r, column=3, value=round(occ/total_sigs*100, 1) if total_sigs else 0)
        ws.cell(row=r, column=3).number_format = '0.0"%"'
        cap = BASE_BALANCE * CAP_PCT
        s   = min(stake, cap)
        ws.cell(row=r, column=4, value=round(s, 2)).number_format = '"$"#,##0.00'
        if is_w:
            ws.cell(row=r, column=5, value=round(stake0 * (PAYOUT_MULT-1), 2)).number_format = '"$"#,##0.00'
            ws.cell(row=r, column=6, value=0).number_format = '"$"#,##0.00'
        else:
            ws.cell(row=r, column=5, value=0).number_format = '"$"#,##0.00'
            ws.cell(row=r, column=6, value=round(-min(masaniello_stake(BASE_BALANCE,0,0,N_OPS,W_NEEDED,PAYOUT_MULT)*(PAYOUT_MULT/(PAYOUT_MULT-1))**2, cap), 2)).number_format = '"$"#,##0.00'
        ws.row_dimensions[r].height = 20

    # ── Rachas de pérdidas consecutivas en sesiones ───────────────────────
    ws.row_dimensions[9].height = 10

    write_header(ws, 10, ["Racha L consecutivas", "Veces ocurrida", "Stake total comprometido",
                           "Balance necesario", "% del capital $300", "Nivel de riesgo"])
    ws.row_dimensions[10].height = 28

    rachas_sim = []
    racha_actual = 0
    rachas_detectadas = []
    for s in sessions:
        if s.wins == 0 and s.losses > 0:
            racha_actual += 1
        else:
            if racha_actual > 0:
                rachas_detectadas.append(racha_actual)
            racha_actual = 0
    if racha_actual > 0:
        rachas_detectadas.append(racha_actual)

    from collections import Counter
    racha_cnt = Counter(rachas_detectadas)
    max_racha = max(rachas_detectadas) if rachas_detectadas else 0

    stakes_acum = []
    for L in range(1, 6):
        total_stake = 0.0
        bal = BASE_BALANCE
        for li in range(L):
            s = masaniello_stake(BASE_BALANCE, li, 0, N_OPS, W_NEEDED, PAYOUT_MULT)
            cap = BASE_BALANCE * CAP_PCT
            s = min(s, cap)
            total_stake += s
        stakes_acum.append(total_stake)

    for r, L in enumerate(range(1, 6), 11):
        total = stakes_acum[L-1]
        veces = racha_cnt.get(L, 0)
        pct_cap = round(total / CAPITAL * 100, 1)
        nivel = "✔ Bajo" if pct_cap < 20 else ("⚠ Medio" if pct_cap < 50 else "🔴 Alto")
        color_n = VERDE_CLARO if pct_cap < 20 else (AMARILLO if pct_cap < 50 else ROJO_CLARO)
        even = (r % 2 == 0)
        bg   = GRIS_CLARO if even else BLANCO

        for c in range(1, 7):
            ws.cell(row=r, column=c).fill   = fill(bg)
            ws.cell(row=r, column=c).border = thin_border()
            ws.cell(row=r, column=c).alignment = center()

        ws.cell(row=r, column=1, value=f"{L} pérdida{'s' if L>1 else ''} seguida{'s' if L>1 else ''}").font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=veces)
        ws.cell(row=r, column=3, value=round(total,2)).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=4, value=round(total,2)).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=5, value=pct_cap).number_format = '0.0"%"'
        cell_n = ws.cell(row=r, column=6, value=nivel)
        cell_n.fill = fill(color_n)
        cell_n.font = Font(bold=True, size=10)
        ws.row_dimensions[r].height = 20

    # Nota peor racha
    ws.row_dimensions[17].height = 10
    ws.merge_cells("A18:F18")
    nota = ws["A18"]
    nota.value = (f"Peor racha observada en el historico real: {max_racha} pérdida(s) consecutiva(s).  "
                  f"Total rachas detectadas: {len(rachas_detectadas)}.  "
                  f"Máx DD por racha 12/4: {money(stakes_acum[min(max_racha,5)-1] if max_racha else 0)}")
    nota.fill  = fill(AMARILLO)
    nota.font  = Font(italic=True, size=9)
    nota.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[18].height = 32


# ═══════════════════════════════════════════════════════════════════════════
# 10. HOJA 6 — TABLA DE STAKES
# ═══════════════════════════════════════════════════════════════════════════

def sheet_stakes(wb):
    ws = wb.create_sheet("Tabla de Stakes 12-4")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "TABLA DE STAKES — MASANIELLO 12/4  (base $300, payout 92%)"
    t.fill  = fill(AZUL_HEADER)
    t.font  = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = center()
    ws.row_dimensions[1].height = 30

    write_header(ws, 2, ["Pérdidas (L)", "Victorias (W)", "Stake entrada",
                          "Stake G1", "Stake G2", "Total expuesto", "% del capital"])
    ws.row_dimensions[2].height = 28

    cap = BASE_BALANCE * CAP_PCT
    r = 3
    for l in range(0, 9):
        for w in range(0, W_NEEDED):
            if l + w >= N_OPS:
                continue
            s = masaniello_stake(BASE_BALANCE, l, w, N_OPS, W_NEEDED, PAYOUT_MULT)
            if s <= 0:
                continue
            pm = PAYOUT_MULT
            g1 = round(min(s * pm / (pm-1), cap), 2)
            g2 = round(min(s * (pm / (pm-1))**2, cap), 2)
            total = round(s + g1 + g2, 2)
            pct_cap = round(s / BASE_BALANCE * 100, 1)
            even = (r % 2 == 0)
            bg   = GRIS_CLARO if even else BLANCO
            if s >= cap * 0.9:
                bg = ROJO_CLARO
            elif s >= cap * 0.5:
                bg = AMARILLO

            for c in range(1, 8):
                ws.cell(row=r, column=c).fill   = fill(bg)
                ws.cell(row=r, column=c).border = thin_border()
                ws.cell(row=r, column=c).alignment = center()
                ws.cell(row=r, column=c).font   = Font(size=10)

            ws.cell(row=r, column=1, value=l)
            ws.cell(row=r, column=2, value=w)
            ws.cell(row=r, column=3, value=s).number_format = '"$"#,##0.00'
            ws.cell(row=r, column=4, value=g1).number_format = '"$"#,##0.00'
            ws.cell(row=r, column=5, value=g2).number_format = '"$"#,##0.00'
            ws.cell(row=r, column=6, value=total).number_format = '"$"#,##0.00'
            ws.cell(row=r, column=7, value=pct_cap).number_format = '0.0"%"'
            ws.row_dimensions[r].height = 18
            r += 1

    # Leyenda
    ws.row_dimensions[r].height = 10
    for label, bg_c in [("Verde = stake normal", BLANCO),
                         ("Amarillo = stake > 50% del cap", AMARILLO),
                         ("Rojo = stake en el tope del cap", ROJO_CLARO)]:
        r += 1
        ws.merge_cells(f"A{r}:G{r}")
        cell = ws[f"A{r}"]
        cell.value = label
        cell.fill  = fill(bg_c)
        cell.font  = Font(italic=True, size=9)
        cell.alignment = Alignment(horizontal="left")
        ws.row_dimensions[r].height = 18


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main() -> None:
    print("Leyendo historico Telegram...")
    outcomes = parse_outcomes(HISTORY)
    print(f"  {len(outcomes)} señales encontradas")

    print("Simulando Masaniello 12/4...")
    sessions_124, daily_124, stats_124 = simulate(outcomes, N_OPS, W_NEEDED)
    print(f"  {len(sessions_124)} sesiones | balance final ${stats_124['final']:.2f}")

    print("Simulando Masaniello 6/2 (referencia)...")
    sessions_62, daily_62, stats_62 = simulate(outcomes, N_OPS_62, W_NEEDED_62)
    print(f"  {len(sessions_62)} sesiones | balance final ${stats_62['final']:.2f}")

    print("Generando Excel...")
    wb = openpyxl.Workbook()

    sheet_resumen(wb, sessions_124, daily_124, stats_124,
                      sessions_62,  daily_62,  stats_62)
    sheet_sesiones(wb, sessions_124)
    sheet_diario(wb, daily_124, stats_124)
    sheet_balance(wb, daily_124, daily_62)
    sheet_rachas(wb, sessions_124)
    sheet_stakes(wb)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\n✔  Informe guardado en: {OUTPUT}")
    print()
    print("=" * 60)
    print("RESUMEN RAPIDO — MASANIELLO 12/4")
    print("=" * 60)
    meta_124 = sum(1 for d in daily_124.values() if d["meta_hit"])
    meta_62  = sum(1 for d in daily_62.values()  if d["meta_hit"])
    print(f"  Balance final 12/4 : ${stats_124['final']:.2f}")
    print(f"  Balance final  6/2  : ${stats_62['final']:.2f}")
    print(f"  MaxDD 12/4          : ${stats_124['max_dd']:.2f}")
    print(f"  MaxDD  6/2          : ${stats_62['max_dd']:.2f}")
    print(f"  Dias meta $60  12/4 : {meta_124}/{len(daily_124)}")
    print(f"  Dias meta $60   6/2 : {meta_62}/{len(daily_62)}")


if __name__ == "__main__":
    main()
