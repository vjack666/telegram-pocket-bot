"""
Simulación Estrategia C: "C_EstrategiaSesiones10"
===================================================
Lógica:
  - Sesiones de 6 mensajes consecutivos (igual que Data_Sesiones_Binarias.csv)
  - Objetivo: +$10 netos por sesión (2 wins × $5 netos c/u)
  - Fórmula de recuperación: stake = (perdida_acumulada + 5.0) / 0.92
  - Take Profit  → 2 wins dentro del bloque → sesión cerrada con +$10
  - Stop Loss    → 3 Losses (Telegram Loss) dentro del bloque → sesión cerrada con pérdida
  - Un "Win" = cualquier resultado != "Loss"  (Win Directo, Win Gale, Win Gale 2)
  - Gestión de capital: continúa aunque el balance sea negativo (se registra la quiebra)
  - Capital inicial: $100
  - Payout: 92% (0.92)
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from pathlib import Path

# ─── Parámetros ─────────────────────────────────────────────────────────────
PAYOUT         = 0.92
OBJETIVO_WIN   = 5.0          # $5 netos por mensaje ganado
CAPITAL_INICIAL = 100.0
TP_WINS        = 2            # Take Profit al llegar a 2 wins
SL_LOSSES      = 3            # Stop Loss al llegar a 3 losses
STAKE_NORMAL   = 5.43
STAKE_RECOVERY = 17.0
LOSS_TELEGRAM_USD = 7.0

CSV_PATH       = Path("runtime/backtest_masaniello_ejemplo.csv")
CSV_PATH_FALLBACK = Path("runtime/Data_Sesiones_Binarias.csv")
EXCEL_OUT      = Path("runtime/Reporte_Comparativo_3Estrategias_100USD.xlsx")
EXCEL_PREV     = Path("runtime/Reporte_Comparativo_MacroRecuperacion_100USD.xlsx")

def _load_dataset() -> pd.DataFrame:
    """Carga dataset de backtest y normaliza columnas para simulación."""
    source_path = CSV_PATH if CSV_PATH.exists() else CSV_PATH_FALLBACK
    if not source_path.exists():
        raise FileNotFoundError(
            f"No se encontró dataset en {CSV_PATH} ni en {CSV_PATH_FALLBACK}"
        )

    data = pd.read_csv(source_path)

    # Normalización de columnas entre datasets legacy y dataset realista.
    rename_map = {
        "Resultado del Backtest": "Resultado",
        "Ciclo Masaniello": "Ciclo_Masaniello",
    }
    for old_name, new_name in rename_map.items():
        if old_name in data.columns and new_name not in data.columns:
            data = data.rename(columns={old_name: new_name})

    if "Resultado" not in data.columns:
        raise ValueError("El CSV no contiene columna 'Resultado' o 'Resultado del Backtest'.")

    # ID de sesión compatible con Estrategia C (bloques de 6 señales).
    if "ID_Sesion" not in data.columns:
        data["ID_Sesion"] = (data.index // 6) + 1

    # Fallback para A/B cuando el dataset no trae ciclo explícito.
    if "Ciclo_Masaniello" not in data.columns:
        data["Ciclo_Masaniello"] = data["ID_Sesion"]

    data["es_win"] = data["Resultado"].astype(str).isin(["Win Directo", "Win Gale", "Win Gale 2"])
    return data


# ─── Carga de datos ──────────────────────────────────────────────────────────
df = _load_dataset()

# ─── Simulación Estrategia C ─────────────────────────────────────────────────

def simular_estrategia_c(df: pd.DataFrame) -> pd.DataFrame:
    """
    Simula la Estrategia C señal por señal, agrupada en bloques de 6.
    Devuelve un DataFrame con una fila por SESIÓN con métricas detalladas.
    """
    sesiones_ids  = df["ID_Sesion"].unique()
    capital       = CAPITAL_INICIAL
    capital_min   = CAPITAL_INICIAL  # capital mínimo tocado en toda la sim.
    registros     = []

    for sid in sesiones_ids:
        bloque = df[df["ID_Sesion"] == sid].copy()
        señales = bloque["es_win"].tolist()
        resultados_raw = bloque["Resultado"].tolist()

        perdida_acum = 0.0
        wins_sesion  = 0
        losses_sesion = 0
        stakes_invertidos = []
        estado_sesion = "Incompleta"
        resultado_neto = 0.0
        max_exposure  = 0.0   # pico de deuda dentro de la sesión

        for i, es_win in enumerate(señales):
            stake = (perdida_acum + OBJETIVO_WIN) / PAYOUT

            if es_win:
                # Ganamos: recuperamos toda la deuda + $5 neto
                capital        += stake * PAYOUT     # +ganancia bruta
                capital        -= perdida_acum       # -recuperamos deuda anterior (contabilizada al perder)
                # Nota: la pérdida ya fue restada cuando ocurrió cada señal,
                # así que solo contabilizamos la ganancia neta de $5
                # Corrección: rehacer con tracking limpio de capital:
                # La pérdida ya fue aplicada. Solo aplicar ganancia bruta.
                # Capital -= stake fue aplicado en los rounds de pérdida previos.
                # Capital += stake*PAYOUT ahora.
                resultado_neto += OBJETIVO_WIN
                perdida_acum    = 0.0
                wins_sesion    += 1
                stakes_invertidos.append(("WIN", round(stake, 4)))
            else:
                # Perdemos: deducimos la apuesta del capital
                capital      -= stake
                perdida_acum += stake
                losses_sesion += 1
                stakes_invertidos.append(("LOSS", round(stake, 4)))

            exposure_actual = perdida_acum
            if exposure_actual > max_exposure:
                max_exposure = exposure_actual

            if capital < capital_min:
                capital_min = capital

            # ─── Reglas de parada
            if wins_sesion >= TP_WINS:
                estado_sesion = "TP_Alcanzado"
                break
            if losses_sesion >= SL_LOSSES:
                estado_sesion = "SL_Activado"
                break
        else:
            # Se agotaron los 6 mensajes sin llegar a TP ni SL
            if wins_sesion >= TP_WINS:
                estado_sesion = "TP_Alcanzado"
            elif losses_sesion >= SL_LOSSES:
                estado_sesion = "SL_Activado"
            else:
                estado_sesion = "Sesion_Agotada"

        # Capital final de la sesión (ya fue ajustado señal a señal)
        # resultado_neto = wins_sesion * $5 - perdida_acum_al_cierre
        perdida_final = perdida_acum if estado_sesion == "SL_Activado" else 0.0

        registros.append({
            "ID_Sesion"         : sid,
            "Señales_Jugadas"   : i + 1,
            "Wins"              : wins_sesion,
            "Losses"            : losses_sesion,
            "Estado"            : estado_sesion,
            "Resultado_Neto_Sesion": round(wins_sesion * OBJETIVO_WIN - (perdida_acum if estado_sesion != "TP_Alcanzado" else 0), 4),
            "Max_Exposure_USD"  : round(max_exposure, 4),
            "Capital_Acumulado" : round(capital, 4),
            "Capital_Negativo"  : capital < 0,
            "Stakes_Secuencia"  : str(stakes_invertidos),
        })

    return pd.DataFrame(registros)


# ─── Recalcular con tracking correcto de capital ─────────────────────────────
# El tracking correcto: restamos la apuesta al perder, sumamos ganancia bruta al ganar.
# El resultado_neto de la sesión = suma de (stake_i * PAYOUT) cuando ganamos - suma de (stake_j) cuando perdemos

def simular_estrategia_c_v2(df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    """
    Versión 2 con tracking de capital limpio y curva de capital sesión a sesión.
    """
    sesiones_ids = df["ID_Sesion"].unique()
    capital      = CAPITAL_INICIAL
    balance_maximo_historico = CAPITAL_INICIAL
    capital_min  = CAPITAL_INICIAL
    quiebras     = 0
    curva_capital = [CAPITAL_INICIAL]
    registros    = []

    for sid in sesiones_ids:
        balance_objetivo = round(balance_maximo_historico + 1.0, 4)
        bloque = df[df["ID_Sesion"] == sid].copy()
        señales = bloque["es_win"].tolist()

        perdida_acum  = 0.0
        wins_sesion   = 0
        losses_sesion = 0
        capital_antes = capital
        señales_jugadas = 0
        estado_sesion = "Sesion_Agotada"
        max_exposure  = 0.0
        stakes_log    = []

        for i, es_win in enumerate(señales):
            stake = (perdida_acum + OBJETIVO_WIN) / PAYOUT
            señales_jugadas = i + 1

            if es_win:
                ganancia_bruta = stake * PAYOUT  # = perdida_acum + OBJETIVO_WIN
                capital       += ganancia_bruta
                resultado_delta = OBJETIVO_WIN   # siempre +$5 netos por diseño de fórmula
                stakes_log.append(f"W:{stake:.2f}")
                perdida_acum   = 0.0
                wins_sesion   += 1
            else:
                capital      -= stake
                perdida_acum += stake
                losses_sesion += 1
                stakes_log.append(f"L:{stake:.2f}")

            if perdida_acum > max_exposure:
                max_exposure = perdida_acum

            if capital < capital_min:
                capital_min = capital

            if wins_sesion >= TP_WINS:
                estado_sesion = "TP_Alcanzado"
                break
            if losses_sesion >= SL_LOSSES:
                estado_sesion = "SL_Activado"
                quiebras     += (1 if capital < 0 else 0)
                break
        else:
            if wins_sesion >= TP_WINS:
                estado_sesion = "TP_Alcanzado"
            elif losses_sesion >= SL_LOSSES:
                estado_sesion = "SL_Activado"

        resultado_sesion = capital - capital_antes
        curva_capital.append(round(capital, 4))

        registros.append({
            "ID_Sesion"             : sid,
            "Señales_Jugadas"       : señales_jugadas,
            "Wins"                  : wins_sesion,
            "Losses"                : losses_sesion,
            "Estado"                : estado_sesion,
            "Resultado_Sesion_USD"  : round(resultado_sesion, 4),
            "Max_Exposure_USD"      : round(max_exposure, 4),
            "Capital_Despues"       : round(capital, 4),
            "Balance_Objetivo"      : balance_objetivo,
            "Capital_Negativo"      : capital < 0,
            "Stakes_Log"            : " | ".join(stakes_log),
        })

        if capital > balance_maximo_historico:
            balance_maximo_historico = capital

    df_out = pd.DataFrame(registros)
    print(f"\n{'='*55}")
    print(f"  SIMULACIÓN C — EstrategiaSesiones$10")
    print(f"{'='*55}")
    print(f"  Sesiones totales        : {len(df_out)}")
    print(f"  TP Alcanzado            : {(df_out['Estado']=='TP_Alcanzado').sum()}")
    print(f"  SL Activado             : {(df_out['Estado']=='SL_Activado').sum()}")
    print(f"  Sesión Agotada (neutral): {(df_out['Estado']=='Sesion_Agotada').sum()}")
    print(f"  Capital Inicial         : ${CAPITAL_INICIAL:.2f}")
    print(f"  Capital Final           : ${capital:.2f}")
    print(f"  Capital Mínimo Tocado   : ${capital_min:.2f}")
    print(f"  Veces en Negativo       : {quiebras}")
    tp = (df_out['Estado']=='TP_Alcanzado').sum()
    sl = (df_out['Estado']=='SL_Activado').sum()
    print(f"  Win Rate (TP/total)     : {tp/len(df_out)*100:.1f}%")
    ganancia_tp = tp * 10
    perdida_sl  = df_out[df_out['Estado']=='SL_Activado']['Resultado_Sesion_USD'].sum()
    print(f"  Ganancia bruta TP       : ${ganancia_tp:.2f}")
    print(f"  Pérdida bruta SL        : ${-perdida_sl:.2f}")
    print(f"  PnL Neto                : ${ganancia_tp + perdida_sl:.2f}")
    print(f"{'='*55}\n")
    return df_out, curva_capital


def _calcular_pnl_realista(resultado: str, stake: float) -> float:
    """PnL por señal considerando costo intraseñal de Gale(s)."""
    normalized = str(resultado).strip().lower()
    if normalized == "loss":
        return -LOSS_TELEGRAM_USD
    if "gale 2" in normalized:
        return (stake * PAYOUT) - (2.0 * stake)
    if "gale" in normalized:
        return (stake * PAYOUT) - stake
    return stake * PAYOUT


def simular_estrategia_c_realista(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, list[float], list[float]]:
    """Simulación evolutiva con memoria real de deuda (persistente en todo el dataset)."""
    balance_actual = CAPITAL_INICIAL
    max_balance_alcanzado = CAPITAL_INICIAL
    en_recuperacion = False
    balance_objetivo = max_balance_alcanzado + 1.0

    detalle_rows: list[dict] = []
    curva_balance = [round(balance_actual, 4)]
    curva_objetivo = [round(balance_objetivo, 4)]

    for idx, row in enumerate(df.itertuples(index=False), start=1):
        resultado = str(getattr(row, "Resultado"))
        id_sesion = int(getattr(row, "ID_Sesion"))

        if balance_actual > max_balance_alcanzado and not en_recuperacion:
            max_balance_alcanzado = balance_actual

        if en_recuperacion:
            stake_usado = STAKE_RECOVERY
            estado_recuperacion = "Si"
        else:
            stake_usado = STAKE_NORMAL
            estado_recuperacion = "No"

        balance_antes = balance_actual
        pnl_senal = _calcular_pnl_realista(resultado, stake_usado)
        balance_actual = balance_actual + pnl_senal

        if str(resultado).strip().lower() == "loss":
            en_recuperacion = True
            balance_objetivo = max_balance_alcanzado + 1.0

        deuda_pendiente = 0.0
        if en_recuperacion:
            deuda_pendiente = max(0.0, balance_objetivo - balance_actual)
            if balance_actual >= balance_objetivo:
                en_recuperacion = False
                deuda_pendiente = 0.0
                if balance_actual > max_balance_alcanzado:
                    max_balance_alcanzado = balance_actual
                balance_objetivo = max_balance_alcanzado + 1.0
        else:
            if balance_actual > max_balance_alcanzado:
                max_balance_alcanzado = balance_actual
            balance_objetivo = max_balance_alcanzado + 1.0

        detalle_rows.append(
            {
                "N_Senal": idx,
                "ID_Sesion": id_sesion,
                "Resultado": resultado,
                "Balance_Antes": round(balance_antes, 4),
                "Estado_Recuperacion": estado_recuperacion,
                "Deuda_Pendiente": round(deuda_pendiente, 4),
                "Stake_Usado": round(stake_usado, 2),
                "PnL_Senal_USD": round(pnl_senal, 4),
                "Balance_Objetivo": round(balance_objetivo, 4),
                "Balance_Despues": round(balance_actual, 4),
                "Balance_Critico_LT30": "Si" if balance_actual < 30.0 else "No",
                "Max_Balance_Alcanzado": round(max_balance_alcanzado, 4),
            }
        )
        curva_balance.append(round(balance_actual, 4))
        curva_objetivo.append(round(balance_objetivo, 4))

    detalle_df = pd.DataFrame(detalle_rows)

    summary_rows: list[dict] = []
    for sid, block in detalle_df.groupby("ID_Sesion", sort=True):
        wins_sesion = int((block["Resultado"].str.lower() != "loss").sum())
        losses_sesion = int((block["Resultado"].str.lower() == "loss").sum())
        if wins_sesion >= TP_WINS:
            estado = "TP_Alcanzado"
        elif losses_sesion >= SL_LOSSES:
            estado = "SL_Activado"
        else:
            estado = "Sesion_Agotada"

        summary_rows.append(
            {
                "ID_Sesion": int(sid),
                "Señales_Jugadas": int(len(block)),
                "Wins": wins_sesion,
                "Losses": losses_sesion,
                "Estado": estado,
                "Resultado_Sesion_USD": round(float(block["PnL_Senal_USD"].sum()), 4),
                "Max_Exposure_USD": round(float(block["Deuda_Pendiente"].max()), 4),
                "Capital_Despues": round(float(block["Balance_Despues"].iloc[-1]), 4),
                "Balance_Objetivo": round(float(block["Balance_Objetivo"].iloc[-1]), 4),
                "Capital_Negativo": bool((block["Balance_Despues"] < 0).any()),
                "Stakes_Log": " | ".join(
                    [f"{res}:{stake:.2f}" for res, stake in zip(block["Resultado"], block["Stake_Usado"])]
                ),
            }
        )

    df_resumen = pd.DataFrame(summary_rows)
    return df_resumen, detalle_df, curva_balance, curva_objetivo


# ─── Cargar curvas A y B del Excel anterior ──────────────────────────────────
def cargar_curvas_anteriores(excel_path: Path):
    """Lee las columnas de capital de A y B del Excel existente si está disponible."""
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        # Buscar hoja de datos o curvas
        sheets = wb.sheetnames
        print(f"[INFO] Sheets en Excel anterior: {sheets}")
        wb.close()
    except Exception as e:
        print(f"[WARN] No se pudo leer Excel anterior: {e}")


# ─── Simulación A (Masaniello) y B (Macro-Recuperación) ─────────────────────
def simular_estrategia_a(df: pd.DataFrame) -> tuple[list, dict]:
    """
    Estrategia A: Masaniello 5/2
    caja = $20, payout = 92%, n_ops = 5, w_needed = 2
    Stake fijo por ciclo. Si cierra con 2+ wins → +caja. Si no → reinicia.
    Simplificación: usamos columna Ciclo_Masaniello del CSV.
    """
    from math import comb

    CAJA         = 20.0
    N_OPS        = 5
    W_NEEDED     = 2
    PAYOUT_M     = 0.92

    # Stake Masaniello: solución del sistema que garantiza recuperar caja
    # Usamos stake fijo simplificado basado en backtest real del CSV
    # El CSV ya tiene Ciclo_Masaniello y resultados → simulamos con backtest real

    capital      = CAPITAL_INICIAL
    capital_min  = CAPITAL_INICIAL
    curva        = [CAPITAL_INICIAL]
    ciclos       = df["Ciclo_Masaniello"].dropna().unique()

    wins_ciclo   = 0
    stake_base   = 2.17  # stake estándar Masaniello $20 / 92% step 1
    losses_ciclo = 0
    ciclo_actual = None

    stats = {"tp": 0, "sl": 0, "capital_final": 0, "capital_min": 0}

    # Usamos los datos brutos del backtest señal a señal
    for _, row in df.iterrows():
        ciclo_id = row["Ciclo_Masaniello"]
        if ciclo_id != ciclo_actual:
            # Nuevo ciclo
            wins_ciclo   = 0
            losses_ciclo = 0
            ciclo_actual = ciclo_id

        es_win = row["es_win"]

        # Stake Masaniello simplificado (stake fijo $2.17 para caja $20)
        stake = stake_base

        if es_win:
            capital    += stake * PAYOUT_M
            wins_ciclo += 1
        else:
            capital      -= stake
            losses_ciclo += 1

        if capital < capital_min:
            capital_min = capital

        curva.append(round(capital, 4))

    stats["capital_final"] = round(capital, 4)
    stats["capital_min"]   = round(capital_min, 4)
    return curva, stats


def simular_estrategia_b(df: pd.DataFrame) -> tuple[list, dict]:
    """
    Estrategia B: Macro-Recuperación
    Cada Loss acumula deuda, la siguiente señal la cubre + objetivo de $2.
    Stop tras 5 losses consecutivos.
    """
    OBJETIVO_B   = 2.0
    SL_B         = 5
    PAYOUT_B     = 0.92

    capital      = CAPITAL_INICIAL
    capital_min  = CAPITAL_INICIAL
    curva        = [CAPITAL_INICIAL]
    perdida_acum = 0.0
    losses_consec = 0

    stats = {"tp": 0, "sl": 0, "capital_final": 0, "capital_min": 0}

    for _, row in df.iterrows():
        es_win = row["es_win"]
        stake  = (perdida_acum + OBJETIVO_B) / PAYOUT_B

        if es_win:
            capital      += stake * PAYOUT_B
            perdida_acum  = 0.0
            losses_consec = 0
            stats["tp"] += 1
        else:
            capital      -= stake
            perdida_acum += stake
            losses_consec += 1

        if losses_consec >= SL_B:
            # Reset (se acepta la pérdida acumulada)
            perdida_acum  = 0.0
            losses_consec = 0
            stats["sl"]  += 1

        if capital < capital_min:
            capital_min = capital

        curva.append(round(capital, 4))

    stats["capital_final"] = round(capital, 4)
    stats["capital_min"]   = round(capital_min, 4)
    return curva, stats


# ─── Generar Excel comparativo ───────────────────────────────────────────────

def color_hex(hex_str: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_str)

def bold_font(size=11, color="000000"):
    return Font(bold=True, size=size, color=color)

def thin_border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def generar_excel(
    df_c: pd.DataFrame,
    df_c_realista: pd.DataFrame,
    curva_c: list,
    curva_objetivo: list,
    curva_a: list,
    stats_a: dict,
    curva_b: list,
    stats_b: dict,
    out_path: Path,
):

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen Ejecutivo ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen Ejecutivo"

    # Paleta
    COLOR_HEADER = "1A1A2E"  # azul noche
    COLOR_A      = "0D7377"  # verde azulado
    COLOR_B      = "E94560"  # rojo coral
    COLOR_C      = "F5A623"  # naranja dorado
    COLOR_BG     = "F7F9FC"
    COLOR_ALT    = "EDF2F7"

    # Título principal
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "REPORTE COMPARATIVO — 3 ESTRATEGIAS | Capital $100 | Payout 92%"
    c.fill  = color_hex(COLOR_HEADER)
    c.font  = Font(bold=True, size=14, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"Dataset: {len(df_c)} sesiones | {df.shape[0]} señales totales | Fuente: ejemplo.md"
    c.fill  = color_hex("2C3E50")
    c.font  = Font(italic=True, size=10, color="BDC3C7")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Headers tabla comparativa
    headers_comp = [
        "Métrica", "A — Masaniello 5/2", "B — Macro-Recuperación", "C — Sesiones $10"
    ]
    col_fills = [COLOR_HEADER, COLOR_A, COLOR_B, COLOR_C]
    row_h = 4
    ws.row_dimensions[row_h].height = 22
    for col_i, (h, fill) in enumerate(zip(headers_comp, col_fills), start=1):
        c = ws.cell(row=row_h, column=col_i, value=h)
        c.fill = color_hex(fill)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    # Stats Estrategia C
    sl_count   = (df_c["Estado"] == "SL_Activado").sum()
    tp_count   = (df_c["Estado"] == "TP_Alcanzado").sum()
    agotada    = (df_c["Estado"] == "Sesion_Agotada").sum()
    total_s    = len(df_c)
    cap_final_c = curva_c[-1]
    cap_min_c  = min(curva_c)
    quiebras   = (df_c["Capital_Negativo"] == True).sum()
    ganancia_tp = tp_count * 10.0
    perdida_sl  = df_c[df_c["Estado"] == "SL_Activado"]["Resultado_Sesion_USD"].sum()
    pnl_c       = ganancia_tp + perdida_sl
    win_rate_c  = tp_count / total_s * 100
    max_exp_c   = df_c["Max_Exposure_USD"].max()
    avg_exp_c   = df_c["Max_Exposure_USD"].mean()

    # Calcular métricas A
    total_señales = len(df)
    wins_a  = df["es_win"].sum()
    losses_a = (~df["es_win"]).sum()
    wr_a    = wins_a / total_señales * 100
    pnl_a   = stats_a["capital_final"] - CAPITAL_INICIAL

    # Calcular métricas B
    pnl_b = stats_b["capital_final"] - CAPITAL_INICIAL

    metricas = [
        ("Capital Inicial",           f"${CAPITAL_INICIAL:.2f}",      f"${CAPITAL_INICIAL:.2f}",   f"${CAPITAL_INICIAL:.2f}"),
        ("Capital Final",             f"${stats_a['capital_final']:.2f}", f"${stats_b['capital_final']:.2f}", f"${cap_final_c:.2f}"),
        ("Capital Mínimo Tocado",     f"${stats_a['capital_min']:.2f}",   f"${stats_b['capital_min']:.2f}",   f"${cap_min_c:.2f}"),
        ("PnL Neto Total",            f"${pnl_a:+.2f}",              f"${pnl_b:+.2f}",              f"${pnl_c:+.2f}"),
        ("Unidades analizadas",       f"{total_señales} señales",    f"{total_señales} señales",    f"{total_s} sesiones"),
        ("Win Rate",                  f"{wr_a:.1f}%",                "N/A (señal a señal)",         f"{win_rate_c:.1f}% (TP/sesión)"),
        ("TP / Wins",                 f"{wins_a}",                   f"{stats_b['tp']}",            f"{tp_count} sesiones"),
        ("SL / Losses",               f"{losses_a}",                 f"{stats_b['sl']} resets",     f"{sl_count} sesiones"),
        ("Sesiones agotadas (neutro)","-",                           "-",                           f"{agotada}"),
        ("Veces capital negativo",    "Ver curva",                   "Ver curva",                   f"{'⚠ ' + str(quiebras) if quiebras > 0 else '✓ 0'}"),
        ("Máxima exposición en sesión","-",                          "-",                           f"${max_exp_c:.2f}"),
        ("Exposición media por sesión","-",                          "-",                           f"${avg_exp_c:.2f}"),
        ("Rentabilidad / $100",       f"{pnl_a/CAPITAL_INICIAL*100:+.1f}%", f"{pnl_b/CAPITAL_INICIAL*100:+.1f}%", f"{pnl_c/CAPITAL_INICIAL*100:+.1f}%"),
        ("Riesgo máx. por sesión",    "Stake ~$2.17",                "Escala con deuda",            f"~$40.45 (3 SL consec.)"),
    ]

    for m_i, (metrica, val_a, val_b, val_c) in enumerate(metricas, start=row_h+1):
        bg = COLOR_BG if m_i % 2 == 0 else COLOR_ALT
        vals = [metrica, val_a, val_b, val_c]
        for col_i, val in enumerate(vals, start=1):
            c = ws.cell(row=m_i, column=col_i, value=val)
            c.fill      = color_hex(bg)
            c.alignment = Alignment(horizontal="left" if col_i == 1 else "center", vertical="center")
            c.border    = thin_border()
            c.font      = Font(size=10, bold=(col_i == 1))
        ws.row_dimensions[m_i].height = 18

    # Anchos de columna
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 22

    # Nota de análisis
    nota_row = row_h + len(metricas) + 2
    ws.merge_cells(f"A{nota_row}:H{nota_row}")
    c = ws.cell(row=nota_row, column=1,
        value="★ ANÁLISIS: La Estrategia C crece casi verticalmente cuando el SL es raro, "
              "pero una racha de 3 pérdidas en el mismo bloque consume ~$40.45 del capital. "
              "Revisar columna 'Punto de Quiebre' en hoja 'Detalle_C'.")
    c.font = Font(italic=True, size=9, color="555555")
    c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[nota_row].height = 32

    # ── Hoja 2: Curvas de Capital ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Curvas de Capital")
    ws2["A1"].value = "# Señal / Sesión"
    ws2["B1"].value = "A — Masaniello"
    ws2["C1"].value = "B — Macro-Recuperación"
    ws2["D1"].value = "C — Sesiones $10"
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = color_hex(COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center")

    # Normalizar longitudes (curva A y B tienen longitud = señales; curva C = sesiones+1)
    max_len = max(len(curva_a), len(curva_b), len(curva_c))

    def pad(lst, length, last_val=None):
        if last_val is None:
            last_val = lst[-1] if lst else 0
        return lst + [last_val] * (length - len(lst))

    curva_a_p = pad(curva_a, max_len)
    curva_b_p = pad(curva_b, max_len)
    curva_c_p = pad(curva_c, max_len)

    # Submuestrear si es muy grande (cada 10 puntos para el gráfico)
    step = max(1, max_len // 500)
    indices = list(range(0, max_len, step))

    for r_i, idx in enumerate(indices, start=2):
        ws2.cell(row=r_i, column=1, value=idx)
        ws2.cell(row=r_i, column=2, value=curva_a_p[idx] if idx < len(curva_a_p) else curva_a_p[-1])
        ws2.cell(row=r_i, column=3, value=curva_b_p[idx] if idx < len(curva_b_p) else curva_b_p[-1])
        ws2.cell(row=r_i, column=4, value=curva_c_p[idx] if idx < len(curva_c_p) else curva_c_p[-1])

    # Gráfico comparativo
    chart = LineChart()
    chart.title   = "Curva de Capital — 3 Estrategias ($100 inicial)"
    chart.style   = 10
    chart.y_axis.title = "Capital ($)"
    chart.x_axis.title = "Señal / Sesión #"
    chart.width   = 28
    chart.height  = 15

    n_data_rows = len(indices) + 1  # +1 header

    for col_idx, (col_letter, series_title, color_str) in enumerate(
        [("B", "A — Masaniello", "0D7377"),
         ("C", "B — Macro-Recup.", "E94560"),
         ("D", "C — Sesiones $10", "F5A623")], start=0
    ):
        data_ref = Reference(ws2, min_col=2+col_idx, min_row=1, max_row=n_data_rows)
        chart.add_data(data_ref, titles_from_data=True)
        chart.series[-1].graphicalProperties.line.solidFill = color_str
        chart.series[-1].graphicalProperties.line.width = 18000  # 2pt en EMU

    ws2.add_chart(chart, "F2")

    # ── Hoja 3: Curva recuperación realista (balance vs línea deuda) ───────
    wsr = wb.create_sheet("Curva_Recuperacion")
    wsr["A1"].value = "# Señal"
    wsr["B1"].value = "Balance_Actual"
    wsr["C1"].value = "Linea_Deuda (Objetivo)"
    for cell in wsr[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = color_hex(COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center")

    max_len_real = max(len(curva_c), len(curva_objetivo))
    step_real = max(1, max_len_real // 1200)
    sampled_idx = list(range(0, max_len_real, step_real))
    for r_i, idx in enumerate(sampled_idx, start=2):
        bal = curva_c[idx] if idx < len(curva_c) else curva_c[-1]
        obj = curva_objetivo[idx] if idx < len(curva_objetivo) else curva_objetivo[-1]
        wsr.cell(row=r_i, column=1, value=idx)
        wsr.cell(row=r_i, column=2, value=bal)
        wsr.cell(row=r_i, column=3, value=obj)

    rec_chart = LineChart()
    rec_chart.title = "Balance Actual vs Linea de Deuda"
    rec_chart.style = 10
    rec_chart.y_axis.title = "Capital ($)"
    rec_chart.x_axis.title = "Señal #"
    rec_chart.width = 28
    rec_chart.height = 14

    n_real_rows = len(sampled_idx) + 1
    data_balance = Reference(wsr, min_col=2, min_row=1, max_row=n_real_rows)
    data_objetivo = Reference(wsr, min_col=3, min_row=1, max_row=n_real_rows)
    rec_chart.add_data(data_balance, titles_from_data=True)
    rec_chart.add_data(data_objetivo, titles_from_data=True)
    rec_chart.series[0].graphicalProperties.line.solidFill = "0D7377"
    rec_chart.series[0].graphicalProperties.line.width = 18000
    rec_chart.series[1].graphicalProperties.line.solidFill = "E94560"
    rec_chart.series[1].graphicalProperties.line.width = 18000
    wsr.add_chart(rec_chart, "E2")

    wsr.column_dimensions["A"].width = 12
    wsr.column_dimensions["B"].width = 18
    wsr.column_dimensions["C"].width = 24
    wsr.freeze_panes = "A2"

    # ── Hoja 4: Detalle Estrategia C (resumen por sesión) ───────────────────
    ws3 = wb.create_sheet("Detalle_C")
    headers_c = [
        "ID_Sesion", "Señales_Jugadas", "Wins", "Losses",
        "Estado", "Resultado_Sesion_USD", "Max_Exposure_USD",
        "Capital_Despues", "Balance_Objetivo", "Capital_Negativo", "Stakes_Log"
    ]
    col_fills_c = {
        "TP_Alcanzado" : "D5F5E3",
        "SL_Activado"  : "FADBD8",
        "Sesion_Agotada": "FEF9E7",
    }

    for col_i, h in enumerate(headers_c, start=1):
        c = ws3.cell(row=1, column=col_i, value=h)
        c.fill = color_hex(COLOR_HEADER)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center")

    for r_i, row_data in enumerate(df_c.itertuples(index=False), start=2):
        estado = row_data.Estado
        bg = col_fills_c.get(estado, "FFFFFF")
        for col_i, val in enumerate([
            row_data.ID_Sesion, row_data.Señales_Jugadas, row_data.Wins,
            row_data.Losses, row_data.Estado, row_data.Resultado_Sesion_USD,
            row_data.Max_Exposure_USD, row_data.Capital_Despues,
            row_data.Balance_Objetivo,
            "SÍ ⚠" if row_data.Capital_Negativo else "No",
            row_data.Stakes_Log
        ], start=1):
            c = ws3.cell(row=r_i, column=col_i, value=val)
            c.fill = color_hex(bg)
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal="center" if col_i != 11 else "left")
            c.border = thin_border()

    # Congelar encabezados
    ws3.freeze_panes = "A2"
    # Anchos
    widths_c = [10, 16, 8, 8, 16, 22, 20, 18, 18, 16, 80]
    for col_i, w in enumerate(widths_c, start=1):
        ws3.column_dimensions[get_column_letter(col_i)].width = w

    # ── Hoja 5: Detalle realista (paso a paso) ───────────────────────────────
    ws5 = wb.create_sheet("Detalle_C_Realista")
    headers_real = [
        "N_Senal",
        "ID_Sesion",
        "Resultado",
        "Balance_Antes",
        "Estado_Recuperacion",
        "Deuda_Pendiente",
        "Stake_Usado",
        "PnL_Senal_USD",
        "Balance_Objetivo",
        "Balance_Despues",
        "Balance_Critico_LT30",
        "Max_Balance_Alcanzado",
    ]
    for col_i, h in enumerate(headers_real, start=1):
        c = ws5.cell(row=1, column=col_i, value=h)
        c.fill = color_hex(COLOR_HEADER)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    for r_i, row_data in enumerate(df_c_realista[headers_real].itertuples(index=False), start=2):
        warn_recovery = str(row_data.Estado_Recuperacion) == "Si"
        warn_critical = str(row_data.Balance_Critico_LT30) == "Si"
        if warn_critical:
            bg = "FADBD8"
        elif warn_recovery:
            bg = "FEF9E7"
        else:
            bg = "FFFFFF"

        for col_i, val in enumerate(row_data, start=1):
            c = ws5.cell(row=r_i, column=col_i, value=val)
            c.fill = color_hex(bg)
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border()

    ws5.freeze_panes = "A2"
    widths_real = [10, 10, 14, 15, 18, 15, 12, 12, 16, 15, 19, 20]
    for col_i, w in enumerate(widths_real, start=1):
        ws5.column_dimensions[get_column_letter(col_i)].width = w

    # ── Hoja 6: Análisis de Punto de Quiebre ─────────────────────────────────
    ws4 = wb.create_sheet("Punto_de_Quiebre")
    ws4.merge_cells("A1:F1")
    c = ws4["A1"]
    c.value = "ANÁLISIS DE PUNTO DE QUIEBRE — Estrategia C"
    c.fill  = color_hex(COLOR_HEADER)
    c.font  = Font(bold=True, size=13, color="FFFFFF")
    c.alignment = Alignment(horizontal="center")
    ws4.row_dimensions[1].height = 26

    # Tabla de escalones de deuda
    ws4.cell(row=3, column=1, value="Tabla de Escalones — Qué pasa si se pierde N veces seguidas")
    ws4.cell(row=3, column=1).font = Font(bold=True, size=11)

    escalonados_hdrs = ["Pérdida #", "Stake ($)", "Deuda Acumulada ($)", "Capital Requerido ($)", "¿Aguanta con $100?"]
    for col_i, h in enumerate(escalonados_hdrs, start=1):
        c = ws4.cell(row=4, column=col_i, value=h)
        c.fill = color_hex("2C3E50")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center")

    perdida_acum_esc = 0.0
    for step_n in range(1, SL_LOSSES + 1):
        stake_esc = (perdida_acum_esc + OBJETIVO_WIN) / PAYOUT
        perdida_acum_esc += stake_esc
        aguanta = "✓ Sí" if perdida_acum_esc <= CAPITAL_INICIAL else "✗ NO ⚠"
        bg_esc = "D5F5E3" if perdida_acum_esc <= CAPITAL_INICIAL else "FADBD8"
        vals_esc = [step_n, round(stake_esc, 4), round(perdida_acum_esc, 4),
                    round(perdida_acum_esc, 4), aguanta]
        for col_i, v in enumerate(vals_esc, start=1):
            c = ws4.cell(row=4 + step_n, column=col_i, value=v)
            c.fill = color_hex(bg_esc)
            c.font = Font(size=10)
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border()

    # Distribución de resultados por sesión
    ws4.cell(row=9, column=1, value="Distribución de resultados de sesión (453 sesiones)")
    ws4.cell(row=9, column=1).font = Font(bold=True, size=11)

    for col_i, h in enumerate(["Estado", "Frecuencia", "% del Total", "Impacto Capital"], start=1):
        c = ws4.cell(row=10, column=col_i, value=h)
        c.fill = color_hex("2C3E50")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center")

    dist = df_c.groupby("Estado").agg(
        Freq=("ID_Sesion", "count"),
        Impacto=("Resultado_Sesion_USD", "sum")
    ).reset_index()

    for r_i, row_d in enumerate(dist.itertuples(index=False), start=11):
        pct = row_d.Freq / total_s * 100
        bg = col_fills_c.get(row_d.Estado, "FFFFFF")
        for col_i, v in enumerate([row_d.Estado, row_d.Freq, f"{pct:.1f}%", f"${row_d.Impacto:+.2f}"], start=1):
            c = ws4.cell(row=r_i, column=col_i, value=v)
            c.fill = color_hex(bg)
            c.font = Font(size=10)
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border()

    # Columnas anchas
    for col_l, w in zip(["A","B","C","D","E","F"], [22, 16, 22, 22, 20, 20]):
        ws4.column_dimensions[col_l].width = w

    # Guardar
    wb.save(out_path)
    print(f"[OK] Excel guardado en: {out_path}")


# ─── Main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("[1/4] Ejecutando Simulación C...")
    df_c, df_c_realista, curva_c, curva_objetivo = simular_estrategia_c_realista(df)

    print("[2/4] Ejecutando Simulación A (Masaniello, referencia)...")
    curva_a, stats_a = simular_estrategia_a(df)

    print("[3/4] Ejecutando Simulación B (Macro-Recuperación, referencia)...")
    curva_b, stats_b = simular_estrategia_b(df)

    print("[4/4] Generando Excel comparativo...")
    generar_excel(
        df_c=df_c,
        df_c_realista=df_c_realista,
        curva_c=curva_c,
        curva_objetivo=curva_objetivo,
        curva_a=curva_a,
        stats_a=stats_a,
        curva_b=curva_b,
        stats_b=stats_b,
        out_path=EXCEL_OUT,
    )

    # Guardar detalle C por sesión y detalle realista paso a paso.
    csv_c_out = Path("runtime/Detalle_SimulacionC_Sesiones10.csv")
    df_c.to_csv(csv_c_out, index=False, encoding="utf-8-sig")
    print(f"[OK] CSV detalle guardado en: {csv_c_out}")

    csv_real_out = Path("runtime/Detalle_C_Realista.csv")
    df_c_realista.to_csv(csv_real_out, index=False, encoding="utf-8-sig")
    print(f"[OK] CSV realista guardado en: {csv_real_out}")
    print("\n✓ Simulación completada.")
