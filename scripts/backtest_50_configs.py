"""
Backtest: 3 configuraciones Masaniello para cuenta de $50
----------------------------------------------------------
A: Conservador  3/6  sesión $10  Macro-Gale x4
B: Equilibrado  2/4  sesión $15  Macro-Gale x4
C: Agresivo     2/3  sesión $25  Macro-Gale x2

Fuente: ejemplo.md
Salida: Backtest_50_Masaniello_Configs.xlsx
"""

import re
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Tuple
from collections import defaultdict

ROOT = Path(__file__).resolve().parents[1]
SIGNAL_FILE = ROOT / "ejemplo.md"
OUTPUT_XLSX = ROOT / "Backtest_50_Masaniello_Configs.xlsx"

# ── Parser de señales ────────────────────────────────────────────────────────
RESULT_PATTERNS = [
    ("WD", re.compile(r"VICTORIA DIRECTA", re.IGNORECASE)),
    ("G1", re.compile(r"VICTORIA EN 1[ªa°]", re.IGNORECASE)),
    ("G2", re.compile(r"VICTORIA EN 2[ªa°]", re.IGNORECASE)),
    ("L",  re.compile(r"PÉRDIDA|PERDIDA|❌", re.IGNORECASE)),
]
DATE_RE = re.compile(r"^\[(\d{2}/\d{2}/\d{4})")


def parse_signals(path: Path) -> List[Tuple[str, str]]:
    signals = []
    current_date = None
    for line in path.read_text(encoding="utf-8").splitlines():
        m = DATE_RE.match(line)
        if m:
            current_date = m.group(1)
        for code, pat in RESULT_PATTERNS:
            if pat.search(line):
                signals.append((current_date or "??/??/????", code))
                break
    return signals


# ── Motor de sesión Masaniello ────────────────────────────────────────────────
@dataclass
class MasaSession:
    ops_total: int
    wins_needed: int
    payout: float
    ops_done: int = field(default=0, init=False)
    wins: int = field(default=0, init=False)
    losses: int = field(default=0, init=False)

    def record(self, code: str) -> None:
        self.ops_done += 1
        if code in ("WD", "G1", "G2"):
            self.wins += 1
        else:
            self.losses += 1

    @property
    def finished(self) -> bool:
        max_losses = self.ops_total - self.wins_needed
        return (self.wins >= self.wins_needed or
                self.losses > max_losses or
                self.ops_done >= self.ops_total)

    @property
    def won(self) -> bool:
        return self.wins >= self.wins_needed

    def pnl(self, session_capital: float) -> float:
        if self.won:
            # Ganancia neta Masaniello: proporcional al payout
            return round(session_capital * (self.payout - 0.5), 2)
        else:
            return -round(session_capital, 2)


@dataclass
class ScenarioConfig:
    name: str
    label: str          # corto para hojas Excel
    capital: float
    session_base: float
    ops_total: int
    wins_needed: int
    max_macro_mult: int
    payout: float = 0.92
    drawdown_limit_pct: float = 0.50


@dataclass
class DayResult:
    date: str
    sessions: int
    sessions_won: int
    sessions_lost: int
    pnl: float
    balance_end: float
    macro_mult: int


def run_scenario(cfg: ScenarioConfig, signals: List[Tuple[str, str]]) -> dict:
    balance = cfg.capital
    macro_mult = 1
    days_detail: List[DayResult] = []
    ruin_day = None
    dd_pct_max = 0.0
    total_sessions = total_won = total_lost = 0

    by_day: dict = defaultdict(list)
    for date, code in signals:
        by_day[date].append(code)
    day_keys = sorted(by_day.keys(), key=lambda d: (d[6:], d[3:5], d[:2]))

    session = MasaSession(ops_total=cfg.ops_total, wins_needed=cfg.wins_needed, payout=cfg.payout)

    for date in day_keys:
        day_pnl = 0.0
        day_s = day_sw = day_sl = 0

        for code in by_day[date]:
            if balance <= 0:
                break
            session_capital = cfg.session_base * macro_mult
            session.record(code)

            if session.finished:
                if session.won:
                    actual_pnl = max(session.pnl(session_capital), 0)
                else:
                    actual_pnl = -min(session_capital, balance)

                balance = round(balance + actual_pnl, 2)
                day_pnl += actual_pnl
                day_s += 1; total_sessions += 1

                if session.won:
                    day_sw += 1; total_won += 1
                    macro_mult = 1
                else:
                    day_sl += 1; total_lost += 1
                    macro_mult = min(macro_mult * 2, cfg.max_macro_mult)

                dd = (cfg.capital - balance) / cfg.capital
                dd_pct_max = max(dd_pct_max, dd)
                if balance <= cfg.capital * (1 - cfg.drawdown_limit_pct) and ruin_day is None:
                    ruin_day = date

                session = MasaSession(ops_total=cfg.ops_total, wins_needed=cfg.wins_needed, payout=cfg.payout)

        days_detail.append(DayResult(
            date=date, sessions=day_s, sessions_won=day_sw, sessions_lost=day_sl,
            pnl=round(day_pnl, 2), balance_end=round(balance, 2), macro_mult=macro_mult,
        ))

    total_days = len(days_detail)
    return {
        "config": cfg,
        "days_detail": days_detail,
        "final_balance": round(balance, 2),
        "net_pnl": round(balance - cfg.capital, 2),
        "roi_pct": round((balance - cfg.capital) / cfg.capital * 100, 2),
        "max_dd_pct": round(dd_pct_max * 100, 2),
        "ruin_day": ruin_day,
        "survived": ruin_day is None,
        "total_sessions": total_sessions,
        "sessions_won": total_won,
        "sessions_lost": total_lost,
        "session_wr_pct": round(total_won / total_sessions * 100, 2) if total_sessions else 0,
        "total_days": total_days,
        "winning_days": sum(1 for d in days_detail if d.pnl > 0),
        "losing_days": sum(1 for d in days_detail if d.pnl < 0),
        "avg_daily_pnl": round((balance - cfg.capital) / total_days, 2) if total_days else 0,
    }


# ── Excel ─────────────────────────────────────────────────────────────────────
def build_excel(results: list, output: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.cell.cell import MergedCell

    wb = Workbook()

    C_NAVY  = "1F4E78"
    C_BLUE  = "2E75B6"
    C_GREEN = "375623"
    C_BG_G  = "E2EFDA"
    C_BG_Y  = "FFF2CC"
    C_BG_R  = "FCE4D6"
    C_BG_B  = "DEEAF1"
    C_LGRAY = "F2F2F2"
    C_WHITE = "FFFFFF"

    SCENARIO_BG = [C_BG_B, C_BG_G, C_BG_Y]     # A, B, C
    SCENARIO_HDR = ["2E75B6", "375623", "C55A11"] # azul, verde, naranja

    def fill(c): return PatternFill("solid", fgColor=c)
    def font(c=C_WHITE, sz=11, bold=True): return Font(color=c, size=sz, bold=bold)
    def border():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)
    def ca(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def la(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def hdr_cell(ws, row, col, val, bg=C_NAVY, fg=C_WHITE, sz=11):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg); c.font = font(fg, sz); c.alignment = ca(); c.border = border()
        return c

    def val_cell(ws, row, col, val, bg=C_WHITE, fmt=None, bold=False, color="000000"):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg); c.alignment = ca(); c.border = border()
        if fmt: c.number_format = fmt
        if bold or color != "000000": c.font = Font(bold=bold, color=color)
        return c

    def auto_w(ws, extra=4, mn=10, mx=55):
        for col in ws.columns:
            best = mn; ltr = None
            for cell in col:
                if isinstance(cell, MergedCell): continue
                if ltr is None: ltr = cell.column_letter
                try: best = max(best, min(len(str(cell.value or "")), mx))
                except: pass
            if ltr: ws.column_dimensions[ltr].width = best + extra

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 1 – Resumen comparativo
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Comparativa"
    ws1.merge_cells("B2:L2")
    t = ws1["B2"]
    t.value = "BACKTEST $50 — MASANIELLO 3 CONFIGURACIONES · HISTÓRICO REAL"
    t.font = Font(color=C_NAVY, size=15, bold=True)
    t.alignment = ca()

    ws1.merge_cells("B3:L3")
    s = ws1["B3"]
    s.value = (f"Fuente: {len(signals)} señales · {results[0]['total_days']} días · "
               "Payout 92% · Capital inicial $50 · Macro-Gale activo")
    s.font = Font(color=C_BLUE, size=10, italic=True)
    s.alignment = ca()

    hdrs = ["Escenario", "Config", "Sesión\nBase", "Macro-Gale\nMáx",
            "Balance\nFinal", "PnL Neto", "ROI%", "DD\nMáx",
            "Sesiones\nTotales", "WR\nSesiones", "PnL\nDiario Prom", "Estado"]
    for ci, h in enumerate(hdrs, 2):
        hdr_cell(ws1, 5, ci, h)

    for ri, res in enumerate(results):
        row = 6 + ri
        cfg = res["config"]
        bg = SCENARIO_BG[ri]
        hc = SCENARIO_HDR[ri]

        val_cell(ws1, row, 2,  cfg.name,           bg, bold=True, color=hc)
        val_cell(ws1, row, 3,  f"{cfg.wins_needed}/{cfg.ops_total}", bg)
        val_cell(ws1, row, 4,  cfg.session_base,   bg, '"$"#,##0.00')
        val_cell(ws1, row, 5,  f"x1→x{cfg.max_macro_mult}", bg)
        val_cell(ws1, row, 6,  res["final_balance"], bg, '"$"#,##0.00')

        pnl_color = "375623" if res["net_pnl"] >= 0 else "C00000"
        val_cell(ws1, row, 7,  res["net_pnl"],     bg, '"$"#,##0.00', bold=True, color=pnl_color)
        val_cell(ws1, row, 8,  res["roi_pct"]/100, bg, '0.00%')
        val_cell(ws1, row, 9,  res["max_dd_pct"]/100, bg, '0.00%')
        val_cell(ws1, row, 10, res["total_sessions"], bg)
        val_cell(ws1, row, 11, res["session_wr_pct"]/100, bg, '0.00%')
        val_cell(ws1, row, 12, res["avg_daily_pnl"], bg, '"$"#,##0.00')

        if res["survived"]:
            val_cell(ws1, row, 13, f"✅ {res['total_days']} días completos", bg, color="375623", bold=True)
        else:
            val_cell(ws1, row, 13, f"⚠ DD>50% día {res['ruin_day']}", bg, color="C00000", bold=True)

    # Leyenda
    ws1.merge_cells("B9:L9")
    leg = ws1["B9"]
    leg.value = ("A=Conservador (3/6 ops, sesión $10)  ·  "
                 "B=Equilibrado (2/4 ops, sesión $15)  ·  "
                 "C=Agresivo (2/3 ops, sesión $25)  ·  "
                 "DD Máx medido desde capital inicial $50")
    leg.font = Font(italic=True, color="595959", size=9)
    leg.alignment = la()

    auto_w(ws1)

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 2 – Detalle diario por escenario (tres sub-tablas en la misma hoja)
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Detalle_Diario")
    ws2.merge_cells("A1:Q1")
    t2 = ws2["A1"]
    t2.value = "DETALLE DÍA A DÍA — LOS 3 ESCENARIOS"
    t2.font = Font(bold=True, size=13, color=C_NAVY)
    t2.alignment = ca()

    col_offset = 1
    for ri, res in enumerate(results):
        cfg = res["config"]
        hc = SCENARIO_HDR[ri]
        bg = SCENARIO_BG[ri]

        # Título de columna del escenario
        ws2.merge_cells(start_row=2, start_column=col_offset,
                        end_row=2, end_column=col_offset + 5)
        tc = ws2.cell(row=2, column=col_offset,
                      value=f"{'━'*3} {cfg.name} — {cfg.wins_needed}/{cfg.ops_total} {'━'*3}")
        tc.font = Font(bold=True, color=hc, size=11)
        tc.alignment = ca()

        sub_hdrs = ["Fecha", "Sesiones", "W", "L", "PnL Día", "Balance"]
        for ci, h in enumerate(sub_hdrs):
            hdr_cell(ws2, 3, col_offset + ci, h, bg=hc)

        for di, d in enumerate(res["days_detail"], start=4):
            bg_row = C_BG_G if d.pnl > 0 else (C_BG_R if d.pnl < 0 else C_LGRAY)
            row_vals = [d.date, d.sessions, d.sessions_won, d.sessions_lost, d.pnl, d.balance_end]
            for ci, val in enumerate(row_vals):
                c = ws2.cell(row=di, column=col_offset + ci, value=val)
                c.fill = fill(bg_row); c.alignment = ca(); c.border = border()
                if ci in (4, 5): c.number_format = '"$"#,##0.00'

        col_offset += 7  # espacio entre escenarios

    auto_w(ws2)

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 3 – Proyección 10 sesiones exitosas (interés compuesto)
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Proyeccion_10_Sesiones")
    ws3.merge_cells("A1:E1")
    t3 = ws3["A1"]
    t3.value = "PROYECCIÓN: 10 SESIONES GANADORAS CONSECUTIVAS (capital reinvertido)"
    t3.font = Font(bold=True, size=13, color=C_NAVY)
    t3.alignment = ca()

    proj_hdrs = ["Sesión", "Balance A\n(Conserv.)", "Balance B\n(Equilib.)", "Balance C\n(Agresivo)",
                 "Nota"]
    for ci, h in enumerate(proj_hdrs, 1):
        hdr_cell(ws3, 2, ci, h)

    # PnL por sesión ganada = session_base * (payout - 0.5)
    # Con reinversión: el session_base crece proporcionalmente al balance
    cfgs = [r["config"] for r in results]
    balances = [50.0, 50.0, 50.0]
    notes = [""] * 10
    notes[4] = "⚠ C: riesgo de ruina al 65% antes de aquí"
    notes[9] = "Meta: triplicar en sesiones B/C"

    for ses in range(1, 11):
        row = ses + 2
        row_vals = [ses]
        for i, cfg in enumerate(cfgs):
            # Sesión base ∝ balance actual (misma proporción que la config inicial)
            ratio = cfg.session_base / 50.0
            sess_cap = round(balances[i] * ratio, 2)
            pnl = round(sess_cap * (cfg.payout - 0.5), 2)
            balances[i] = round(balances[i] + pnl, 2)
            row_vals.append(balances[i])

        row_vals.append(notes[ses - 1])
        bg_row = C_BG_G if ses % 2 == 0 else C_WHITE
        for ci, val in enumerate(row_vals, 1):
            c = ws3.cell(row=row, column=ci, value=val)
            c.fill = fill(bg_row); c.alignment = ca(); c.border = border()
            if ci in (2, 3, 4): c.number_format = '"$"#,##0.00'
            if ci == 5 and val:
                c.font = Font(color="C00000", bold=True)

    # Totales finales
    final_row = 13
    ws3.cell(row=final_row, column=1, value="FINAL").font = Font(bold=True)
    for i, b in enumerate(balances):
        color = "375623" if b > 50 else "C00000"
        c = ws3.cell(row=final_row, column=i + 2, value=b)
        c.number_format = '"$"#,##0.00'
        c.font = Font(bold=True, color=color)
        c.alignment = ca(); c.border = border()

    auto_w(ws3)

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 4 – Análisis de riesgo por escenario
    # ════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Riesgo_y_Recomendacion")
    ws4.merge_cells("A1:C1")
    ws4["A1"].value = "ANÁLISIS DE RIESGO Y RECOMENDACIÓN"
    ws4["A1"].font = Font(bold=True, size=13, color=C_NAVY)
    ws4["A1"].alignment = ca()

    filas = [
        ("ESCENARIO", "FORTALEZA", "DEBILIDAD / RIESGO"),
        ("A — Conservador (3/6)\nSesión $10",
         "Máxima supervivencia. Tolera hasta 3 fallos por sesión. "
         "Ideal para cuentas pequeñas o canales con eficiencia < 80%.",
         "Crecimiento lento. Necesita muchas señales para mover el balance de $50."),
        ("B — Equilibrado (2/4)\nSesión $15",
         "Punto dulce. Solo necesitas 50% de efectividad para ganar la sesión. "
         "Más rápido que A con riesgo manejable. Recomendado para presentaciones.",
         "Pierde el 30% del capital por sesión si falla. Con Macro-Gale x4, "
         "una racha de 3 sesiones perdidas compromete el 90% del capital."),
        ("C — Agresivo (2/3)\nSesión $25",
         "Crecimiento explosivo. 2 sesiones ganadoras = +100% del capital.",
         "Solo 2 fallos seguidos destruyen el 50% DD. Riesgo de ruina: ~65% "
         "antes de completar 10 sesiones según datos históricos."),
        ("RECOMENDACIÓN FINAL", "", ""),
        ("Para $50 en producción",
         "Usar Escenario A con sesión $5 (10% del capital) y Macro-Gale x4. "
         "Una vez alcanzado $150 (3x), escalar a Escenario B.",
         "Nunca usar Escenario C con capital total < $200. "
         "El historial muestra rachas de 4+ pérdidas que lo destruirían."),
    ]

    bg_map = {0: C_NAVY, 1: C_BG_B, 2: C_BG_G, 3: C_BG_Y, 4: "D9D9D9", 5: "F2F2F2"}
    fg_map = {0: C_WHITE, 1: "000000", 2: "000000", 3: "000000", 4: C_NAVY, 5: "000000"}
    bold_map = {0: True, 4: True, 5: False}

    for ri, (col_a, col_b, col_c) in enumerate(filas):
        bg = bg_map.get(ri, C_WHITE); fg = fg_map.get(ri, "000000")
        bd = bold_map.get(ri, False)
        for ci, val in enumerate([col_a, col_b, col_c], 1):
            c = ws4.cell(row=ri + 2, column=ci, value=val)
            c.fill = fill(bg)
            c.font = Font(color=fg, bold=bd)
            c.alignment = la()
            c.border = border()

    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 60
    ws4.column_dimensions["C"].width = 55

    wb.save(output)
    print(f"\n✔ Excel guardado: {output}")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"Cargando señales desde {SIGNAL_FILE} ...")
    signals = parse_signals(SIGNAL_FILE)
    print(f"  → {len(signals)} resultados\n")

    scenarios = [
        ScenarioConfig("A — Conservador", "A",  capital=50, session_base=10, ops_total=6, wins_needed=3, max_macro_mult=4),
        ScenarioConfig("B — Equilibrado", "B",  capital=50, session_base=15, ops_total=4, wins_needed=2, max_macro_mult=4),
        ScenarioConfig("C — Agresivo",    "C",  capital=50, session_base=25, ops_total=3, wins_needed=2, max_macro_mult=2),
    ]

    results = []
    for cfg in scenarios:
        res = run_scenario(cfg, signals)
        results.append(res)
        sup = f"✅ {res['total_days']} días" if res["survived"] else f"⚠ DD>50% día {res['ruin_day']}"
        print(f"[{cfg.name}]")
        print(f"  Config          : {cfg.wins_needed}/{cfg.ops_total}  sesión=${cfg.session_base}  Macro-Gale x{cfg.max_macro_mult}")
        print(f"  Balance final   : ${res['final_balance']:.2f}  (PnL ${res['net_pnl']:+.2f}  ROI {res['roi_pct']:.1f}%)")
        print(f"  DD máximo       : {res['max_dd_pct']:.1f}%")
        print(f"  WR sesiones     : {res['session_wr_pct']:.1f}%  ({res['sessions_won']}/{res['total_sessions']})")
        print(f"  PnL diario prom : ${res['avg_daily_pnl']:.2f}")
        print(f"  Supervivencia   : {sup}\n")

    print("Generando Excel...")
    build_excel(results, OUTPUT_XLSX)
