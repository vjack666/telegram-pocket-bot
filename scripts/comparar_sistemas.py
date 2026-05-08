"""
Comparativa completa — 3 sistemas sobre los mismos datos reales.
Datos: ejemplo.md (2,655 señales, 17/03/2026 - 07/05/2026)

Sistema A — Bot actual en producción
  Masaniello 12/4, payout 92%, base $300
  Cap por paso: entry ≤ $30, G1 ≤ $30, G2 ≤ $30
  (exposición máxima real si pierden los 3: ~$74.70)

Sistema B — Masaniello 12/4 con cap TOTAL corregido
  Mismos parámetros pero entry_max = $30 / 7.44 ≈ $4.03
  Exposición máxima garantizada: $30.00 (10% de $300)

Sistema C — Modo fijo $2/$4/$10 (fallback del bot)
  Sin sesiones, sin Masaniello
  Entrada fija $2, G1=$4, G2=$10 sin importar balance

Hojas generadas:
  1. Comparativa KPIs       — tabla side-by-side A vs B vs C
  2. Balance dia a dia      — gráfica 3 líneas
  3. Distribucion de riesgo — tabla de exposición real por señal
  4. Detalle A (bot actual) — sesiones y P&L
  5. Detalle B (cap total)  — sesiones y P&L
"""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

HISTORY = ROOT / "ejemplo.md"
OUTPUT  = ROOT / "runtime" / "comparativa_sistemas.xlsx"

# ── Parámetros comunes ──────────────────────────────────────────────────────
PAYOUT_MULT    = 1.92
NET_PAYOUT     = PAYOUT_MULT - 1  # 0.92
CAPITAL        = 300.0
N_OPS          = 12
W_NEEDED       = 4
CAP_PCT        = 0.10
CAP_ABS        = CAPITAL * CAP_PCT  # $30.00

# Ratio de gale para Masaniello: G1 = entry * r, G2 = entry * r^2
_R             = PAYOUT_MULT / NET_PAYOUT   # ≈ 2.0869
_TOTAL_MULT    = 1 + _R + _R ** 2           # ≈ 7.442

# Sistema A: cap por PASO (comportamiento actual del bot)
CAP_STEP_A     = CAP_ABS    # cada paso ≤ $30

# Sistema B: cap TOTAL (corregido)
ENTRY_MAX_B    = round(CAP_ABS / _TOTAL_MULT, 2)  # ≈ $4.03

# Sistema C: fijo
FIXED_ENTRY    = 2.00
FIXED_G1       = 4.00
FIXED_G2       = 10.00
META_DIARIA    = 60.0

# ── Colores ──────────────────────────────────────────────────────────────────
AZUL_HEADER  = "1F3864"
AZUL_MEDIO   = "2E75B6"
AZUL_CLARO   = "D6E4F7"
VERDE_OSC    = "1E6F4B"
VERDE_CLARO  = "D6F0E0"
ROJO_OSC     = "8B1A1A"
ROJO_CLARO   = "FDDEDE"
NARANJA      = "FF8C00"
NARANJA_CL   = "FFE5B4"
AMARILLO     = "FFF2CC"
GRIS_CLARO   = "F2F2F2"
BLANCO       = "FFFFFF"


# ═══════════════════════════════════════════════════════════════════════════
# 1. PARSER
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class Outcome:
    timestamp: datetime
    result: str    # WD, G1, G2, L
    is_win: bool


def parse_outcomes(path: Path) -> list[Outcome]:
    date_pat = re.compile(r"^\[(\d{2}/\d{2}/\d{4}) (\d{2}:\d{2}:\d{2})\]")
    res_pat  = re.compile(
        r"(VICTORIA DIRECTA|VICTORIA EN 1.*?MARTINGALA|VICTORIA EN 2.*?MARTINGALA|P[ÉE]RDIDA)",
        re.IGNORECASE,
    )
    label_map = {
        "victoria directa": "WD",
        "victoria en 1": "G1",
        "victoria en 2": "G2",
        "perdida": "L",
        "pérdida": "L",
    }
    outcomes: list[Outcome] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        ts_m = date_pat.match(line)
        res_m = res_pat.search(line)
        if not ts_m or not res_m:
            continue
        ts = datetime.strptime(f"{ts_m.group(1)} {ts_m.group(2)}", "%d/%m/%Y %H:%M:%S")
        raw = res_m.group(1).lower()
        label = next((v for k, v in label_map.items() if raw.startswith(k)), "?")
        outcomes.append(Outcome(ts, label, label != "L"))
    return outcomes


# ═══════════════════════════════════════════════════════════════════════════
# 2. FÓRMULA MASANIELLO
# ═══════════════════════════════════════════════════════════════════════════

def _fwd_prob(ops_left: int, wins_needed: int, pm: float) -> float:
    if wins_needed <= 0:         return 1.0
    if wins_needed > ops_left:   return 0.0
    if wins_needed == ops_left:  return pm ** ops_left
    pw = _fwd_prob(ops_left - 1, wins_needed - 1, pm)
    pl = _fwd_prob(ops_left - 1, wins_needed,     pm)
    d  = pw + (pm - 1) * pl
    return (pm * pw * pl / d) if d else 0.0


def masaniello_raw(losses: int, wins: int) -> float:
    """Stake Masaniello puro (sin cap) sobre BASE $300."""
    ops_left  = N_OPS - (losses + wins)
    wins_left = W_NEEDED - wins
    if ops_left <= 0 or wins_left <= 0 or wins_left > ops_left:
        return 0.0
    pw = _fwd_prob(ops_left - 1, wins_left - 1, PAYOUT_MULT)
    pl = _fwd_prob(ops_left - 1, wins_left,     PAYOUT_MULT)
    d  = pw + NET_PAYOUT * pl
    if not d:
        return CAPITAL
    s = CAPITAL * (1 - PAYOUT_MULT * pw / d)
    return round(max(0.01, min(s, CAPITAL)), 2)


# ═══════════════════════════════════════════════════════════════════════════
# 3. SIMULACIONES
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class DayStats:
    start_bal: float = 0.0
    end_bal:   float = 0.0
    min_bal:   float = 0.0
    pnl:       float = 0.0
    sessions:  int   = 0
    meta_hit:  bool  = False
    ops_used:  int   = 0


@dataclass
class SimResult:
    final:        float
    peak:         float
    min_bal:      float
    max_dd:       float
    sessions:     int
    sessions_won: int
    total_sigs:   int
    win_sigs:     int
    daily:        dict[str, DayStats]
    # riesgo real
    max_single_loss: float  # máxima pérdida en una señal perdida
    total_losses_count: int


def _simulate_masaniello(outcomes: list[Outcome], entry_fn) -> SimResult:
    """
    Simula Masaniello 12/4. entry_fn(losses, wins) devuelve el entry a usar.
    Pérdida real = entry + G1_real + G2_real (siempre proporcionales al entry).
    """
    balance       = CAPITAL
    peak          = CAPITAL
    min_bal       = CAPITAL
    max_dd        = 0.0
    sessions_won  = 0
    sessions_total = 0
    total_sigs    = 0
    win_sigs      = 0
    max_single_loss = 0.0
    total_losses_count = 0

    daily: dict[str, DayStats] = {}

    chunks = [outcomes[i:i + N_OPS] for i in range(0, len(outcomes), N_OPS)]

    for chunk in chunks:
        if not chunk:
            continue
        day  = chunk[0].timestamp.strftime("%d/%m/%Y")
        if day not in daily:
            daily[day] = DayStats(
                start_bal=round(balance, 2),
                end_bal=round(balance, 2),
                min_bal=round(balance, 2),
            )
        d = daily[day]
        if d.meta_hit:
            continue

        wins = losses = 0
        session_won = False

        for outcome in chunk:
            wins_needed = W_NEEDED - wins
            ops_left    = N_OPS - (wins + losses)
            if wins_needed <= 0 or wins_needed > ops_left:
                break

            entry = entry_fn(losses, wins)
            g1    = round(entry * _R,      2)
            g2    = round(entry * _R ** 2, 2)
            total_loss_signal = round(entry + g1 + g2, 2)

            total_sigs += 1
            d.ops_used += 1

            if outcome.is_win:
                wins     += 1
                win_sigs += 1
                gain      = round(entry * NET_PAYOUT, 2)
                balance   = round(balance + gain, 2)
                d.pnl     = round(d.pnl + gain, 2)
            else:
                losses            += 1
                total_losses_count += 1
                balance            = round(balance - total_loss_signal, 2)
                d.pnl              = round(d.pnl - total_loss_signal, 2)
                max_single_loss    = max(max_single_loss, total_loss_signal)

            min_bal    = min(min_bal, balance)
            d.min_bal  = min(d.min_bal, balance)
            peak       = max(peak, balance)
            dd         = round(peak - balance, 2)
            max_dd     = max(max_dd, dd)

            if wins >= W_NEEDED:
                session_won = True
                break

        sessions_total += 1
        if session_won:
            sessions_won += 1
        d.sessions += 1
        d.end_bal   = round(balance, 2)
        if d.pnl >= META_DIARIA:
            d.meta_hit = True

    return SimResult(
        final=round(balance, 2), peak=round(peak, 2), min_bal=round(min_bal, 2),
        max_dd=round(max_dd, 2), sessions=sessions_total, sessions_won=sessions_won,
        total_sigs=total_sigs, win_sigs=win_sigs, daily=daily,
        max_single_loss=round(max_single_loss, 2),
        total_losses_count=total_losses_count,
    )


def _simulate_fixed(outcomes: list[Outcome]) -> SimResult:
    """Sistema C: montos fijos $2/$4/$10, una señal = una operación sin sesiones."""
    balance            = CAPITAL
    peak               = CAPITAL
    min_bal            = CAPITAL
    max_dd             = 0.0
    wins               = 0
    losses_count       = 0
    max_single_loss    = 0.0
    daily: dict[str, DayStats] = {}

    for outcome in outcomes:
        day = outcome.timestamp.strftime("%d/%m/%Y")
        if day not in daily:
            daily[day] = DayStats(
                start_bal=round(balance, 2),
                end_bal=round(balance, 2),
                min_bal=round(balance, 2),
            )
        d = daily[day]

        d.ops_used += 1

        if outcome.result == "WD":
            gain     = round(FIXED_ENTRY * NET_PAYOUT, 2)
            balance  = round(balance + gain, 2)
            d.pnl    = round(d.pnl + gain, 2)
            wins    += 1
        elif outcome.result == "G1":
            pnl_g1   = round(-FIXED_ENTRY + FIXED_G1 * NET_PAYOUT, 2)
            balance  = round(balance + pnl_g1, 2)
            d.pnl    = round(d.pnl + pnl_g1, 2)
            wins    += 1
        elif outcome.result == "G2":
            pnl_g2   = round(-FIXED_ENTRY - FIXED_G1 + FIXED_G2 * NET_PAYOUT, 2)
            balance  = round(balance + pnl_g2, 2)
            d.pnl    = round(d.pnl + pnl_g2, 2)
            wins    += 1
        else:  # L
            total_l  = FIXED_ENTRY + FIXED_G1 + FIXED_G2
            balance  = round(balance - total_l, 2)
            d.pnl    = round(d.pnl - total_l, 2)
            losses_count       += 1
            max_single_loss     = max(max_single_loss, total_l)

        min_bal   = min(min_bal, balance)
        d.min_bal = min(d.min_bal, balance)
        peak      = max(peak, balance)
        dd        = round(peak - balance, 2)
        max_dd    = max(max_dd, dd)
        d.end_bal = round(balance, 2)
        if d.pnl >= META_DIARIA:
            d.meta_hit = True

    return SimResult(
        final=round(balance, 2), peak=round(peak, 2), min_bal=round(min_bal, 2),
        max_dd=round(max_dd, 2), sessions=len(outcomes), sessions_won=wins,
        total_sigs=len(outcomes), win_sigs=wins, daily=daily,
        max_single_loss=round(max_single_loss, 2),
        total_losses_count=losses_count,
    )


# ═══════════════════════════════════════════════════════════════════════════
# 4. HELPERS EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def fill(h: str) -> PatternFill:
    return PatternFill("solid", fgColor=h)

def thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def write_header(ws, row: int, cols: list[str], bg=AZUL_HEADER):
    for c, t in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=t)
        cell.fill      = fill(bg)
        cell.font      = Font(bold=True, size=10, color="FFFFFF")
        cell.alignment = center()
        cell.border    = thin_border()
    ws.row_dimensions[row].height = 28

def money(v: float) -> str:
    return f"${v:,.2f}"

def pct(v: float) -> str:
    return f"{v:.1f}%"

def _color_diff(v: float) -> str:
    return VERDE_OSC if v > 0 else (ROJO_OSC if v < 0 else "000000")


# ═══════════════════════════════════════════════════════════════════════════
# 5. HOJA 1 — COMPARATIVA KPIs
# ═══════════════════════════════════════════════════════════════════════════

def sheet_kpis(wb, ra: SimResult, rb: SimResult, rc: SimResult):
    ws = wb.active
    ws.title = "Comparativa KPIs"
    ws.sheet_view.showGridLines = False

    col_widths = [36, 20, 20, 20, 14, 14]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Título
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "COMPARATIVA DE SISTEMAS — DATOS REALES TELEGRAM (2,655 señales)"
    t.fill  = fill(AZUL_HEADER)
    t.font  = Font(bold=True, size=14, color="FFFFFF")
    t.alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = "Periodo: 17/03/2026 – 07/05/2026  |  52 días  |  Capital inicial: $300  |  Payout: 92%"
    sub.fill  = fill(AZUL_MEDIO)
    sub.font  = Font(size=10, color="FFFFFF")
    sub.alignment = center()
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    # Cabeceras de sistemas
    write_header(ws, 4, [
        "Métrica",
        "A — Bot actual\n(12/4, cap $30/paso)",
        "B — Cap total corregido\n(12/4, total ≤ $30)",
        "C — Fijo $2/$4/$10",
        "B vs A",
        "B vs C",
    ])

    # Sub-header de riesgo
    ws.merge_cells("A5:A6")
    for col, (label, bg) in enumerate([
        ("A — Bot actual\ncap por paso: entry≤$30  G1≤$30  G2≤$30\nPérdida máx/señal: ~$74.70", AZUL_CLARO),
        ("B — Cap total\nentry_max=$4.03  G1=$8.41  G2=$17.57\nPérdida máx/señal: $30.00 exactos", VERDE_CLARO),
        ("C — Fijo\nentry=$2  G1=$4  G2=$10\nPérdida máx/señal: $16.00", NARANJA_CL),
    ], start=2):
        ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
        cell = ws.cell(row=5, column=col, value=label)
        cell.fill      = fill(bg)
        cell.alignment = center()
        cell.border    = thin_border()
        cell.font      = Font(size=9, italic=True)
    for col in [5, 6]:
        ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
    ws.row_dimensions[5].height = 40
    ws.row_dimensions[6].height = 20

    # Datos
    dias_a = len(ra.daily)
    dias_b = len(rb.daily)
    dias_c = len(rc.daily)
    meta_a = sum(1 for d in ra.daily.values() if d.meta_hit)
    meta_b = sum(1 for d in rb.daily.values() if d.meta_hit)
    meta_c = sum(1 for d in rc.daily.values() if d.meta_hit)
    wr_a   = ra.win_sigs / ra.total_sigs * 100 if ra.total_sigs else 0
    wr_b   = rb.win_sigs / rb.total_sigs * 100 if rb.total_sigs else 0
    wr_c   = rc.win_sigs / rc.total_sigs * 100 if rc.total_sigs else 0
    roi_a  = (ra.final - CAPITAL) / CAPITAL * 100
    roi_b  = (rb.final - CAPITAL) / CAPITAL * 100
    roi_c  = (rc.final - CAPITAL) / CAPITAL * 100

    rows_data = [
        # (label, val_a, val_b, val_c, num_a, num_b)  — num_* para diff
        ("Balance final",                  money(ra.final),  money(rb.final),  money(rc.final),  ra.final,   rb.final),
        ("ROI total (%)",                  pct(roi_a),       pct(roi_b),       pct(roi_c),       roi_a,      roi_b),
        ("Balance máximo alcanzado",       money(ra.peak),   money(rb.peak),   money(rc.peak),   ra.peak,    rb.peak),
        ("Balance mínimo observado",       money(ra.min_bal),money(rb.min_bal),money(rc.min_bal),ra.min_bal, rb.min_bal),
        ("DrawDown máximo ($)",            money(ra.max_dd), money(rb.max_dd), money(rc.max_dd), -ra.max_dd, -rb.max_dd),
        ("DrawDown máximo (%)",            pct(ra.max_dd/CAPITAL*100), pct(rb.max_dd/CAPITAL*100), pct(rc.max_dd/CAPITAL*100), None, None),
        ("Pérdida máxima por señal ($)",   money(ra.max_single_loss), money(rb.max_single_loss), money(rc.max_single_loss), -ra.max_single_loss, -rb.max_single_loss),
        ("Pérdida máx como % del capital", pct(ra.max_single_loss/CAPITAL*100), pct(rb.max_single_loss/CAPITAL*100), pct(rc.max_single_loss/CAPITAL*100), None, None),
        ("Días analizados",                str(dias_a),      str(dias_b),      str(dias_c),      None, None),
        ("Días con meta $60",              f"{meta_a}/{dias_a}", f"{meta_b}/{dias_b}", f"{meta_c}/{dias_c}", meta_a, meta_b),
        ("Señales totales",                str(ra.total_sigs), str(rb.total_sigs), str(rc.total_sigs), None, None),
        ("Señales ganadas",                str(ra.win_sigs), str(rb.win_sigs),  str(rc.win_sigs), ra.win_sigs, rb.win_sigs),
        ("Win Rate señales",               pct(wr_a),        pct(wr_b),        pct(wr_c),        None, None),
        ("Total señales perdidas (L)",     str(ra.total_losses_count), str(rb.total_losses_count), str(rc.total_losses_count), None, None),
    ]

    for r, (label, va, vb, vc, na, nb) in enumerate(rows_data, 7):
        even = (r % 2 == 0)
        bg   = GRIS_CLARO if even else BLANCO

        ws.cell(row=r, column=1, value=label).fill   = fill(bg)
        ws.cell(row=r, column=1).border  = thin_border()
        ws.cell(row=r, column=1).font    = Font(bold=True, size=10)

        for col, val, col_bg in [(2, va, AZUL_CLARO), (3, vb, VERDE_CLARO), (4, vc, NARANJA_CL)]:
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill      = fill(col_bg)
            cell.alignment = center()
            cell.border    = thin_border()
            cell.font      = Font(size=10)

        # Diff B vs A
        if na is not None and nb is not None:
            diff_val = nb - na
            txt = f"{diff_val:+,.2f}" if isinstance(diff_val, float) else f"{diff_val:+d}"
            cell = ws.cell(row=r, column=5, value=txt)
            cell.font = Font(size=10, color=_color_diff(diff_val))
        else:
            ws.cell(row=r, column=5, value="—")
        ws.cell(row=r, column=5).alignment = center()
        ws.cell(row=r, column=5).border    = thin_border()
        ws.cell(row=r, column=5).fill      = fill(bg)
        # Col 6 vacía por ahora
        ws.cell(row=r, column=6, value="").alignment = center()
        ws.cell(row=r, column=6).border = thin_border()
        ws.cell(row=r, column=6).fill   = fill(bg)

        ws.row_dimensions[r].height = 20

    # Nota al pie
    last = 7 + len(rows_data)
    ws.row_dimensions[last].height = 8
    ws.merge_cells(f"A{last+1}:F{last+1}")
    n = ws[f"A{last+1}"]
    n.value = (
        "★  Sistema A = bot en producción hoy (Masaniello 12/4, APP_MAX_TRADE_PCT=0.10 aplica a CADA paso). "
        "Sistema B = corrección propuesta: total exposure ≤ 10% por señal (entry_max=$4.03). "
        "Sistema C = modo fijo de respaldo ($2/$4/$10). "
        "Los 3 sistemas usan exactamente las mismas 2,655 señales históricas."
    )
    n.fill  = fill(AMARILLO)
    n.font  = Font(italic=True, size=9)
    n.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[last+1].height = 40


# ═══════════════════════════════════════════════════════════════════════════
# 6. HOJA 2 — BALANCE DÍA A DÍA (gráfica)
# ═══════════════════════════════════════════════════════════════════════════

def sheet_balance(wb, ra: SimResult, rb: SimResult, rc: SimResult):
    ws = wb.create_sheet("Balance dia a dia")
    ws.sheet_view.showGridLines = False

    col_widths = [12, 14, 14, 14]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    write_header(ws, 1, ["Fecha", "A — Bot actual", "B — Cap total", "C — Fijo 2/4/10"])

    all_days = sorted(
        set(list(ra.daily.keys()) + list(rb.daily.keys()) + list(rc.daily.keys())),
        key=lambda d: datetime.strptime(d, "%d/%m/%Y"),
    )

    bal_a = bal_b = bal_c = CAPITAL

    for r, day in enumerate(all_days, 2):
        da = ra.daily.get(day)
        db = rb.daily.get(day)
        dc = rc.daily.get(day)

        if da and da.end_bal: bal_a = da.end_bal
        if db and db.end_bal: bal_b = db.end_bal
        if dc and dc.end_bal: bal_c = dc.end_bal

        even = (r % 2 == 0)
        bg = GRIS_CLARO if even else BLANCO
        for col in range(1, 5):
            ws.cell(row=r, column=col).fill   = fill(bg)
            ws.cell(row=r, column=col).border = thin_border()
            ws.cell(row=r, column=col).alignment = center()

        ws.cell(row=r, column=1, value=day)
        for col, val in [(2, bal_a), (3, bal_b), (4, bal_c)]:
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = '"$"#,##0.00'
        ws.row_dimensions[r].height = 16

    n_rows = len(all_days)

    chart = LineChart()
    chart.title  = "Balance acumulado — 3 sistemas"
    chart.style  = 10
    chart.y_axis.title = "Balance ($)"
    chart.x_axis.title = "Día"
    chart.height = 16
    chart.width  = 30

    for col, color, width in [(2, "2E75B6", 20000), (3, "1E6F4B", 20000), (4, "FF8C00", 15000)]:
        ref = Reference(ws, min_col=col, min_row=1, max_row=1 + n_rows)
        chart.add_data(ref, titles_from_data=True)

    chart.series[0].graphicalProperties.line.solidFill = "2E75B6"
    chart.series[0].graphicalProperties.line.width = 20000
    chart.series[1].graphicalProperties.line.solidFill = "1E6F4B"
    chart.series[1].graphicalProperties.line.width = 20000
    chart.series[2].graphicalProperties.line.solidFill = "FF8C00"
    chart.series[2].graphicalProperties.line.width = 15000

    dates_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + n_rows)
    chart.set_categories(dates_ref)
    ws.add_chart(chart, "F2")


# ═══════════════════════════════════════════════════════════════════════════
# 7. HOJA 3 — DISTRIBUCIÓN DE RIESGO
# ═══════════════════════════════════════════════════════════════════════════

def sheet_riesgo(wb):
    ws = wb.create_sheet("Distribucion de Riesgo")
    ws.sheet_view.showGridLines = False

    col_widths = [22, 14, 14, 14, 16, 16, 16]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "TABLA DE RIESGO REAL POR SEÑAL — Los 3 sistemas"
    t.fill  = fill(AZUL_HEADER)
    t.font  = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = center()
    ws.row_dimensions[1].height = 30

    write_header(ws, 2, ["Estado L/W", "Entry", "G1", "G2",
                          "A pérdida total", "B pérdida total", "C pérdida total"])

    r = 3
    for l in range(0, 7):
        for w in range(0, W_NEEDED):
            if l + w >= N_OPS:
                continue
            raw_entry = masaniello_raw(l, w)
            if raw_entry <= 0:
                continue

            # Sistema A: cap por paso
            entry_a = min(raw_entry, CAP_STEP_A)
            g1_a    = min(entry_a * _R,      CAP_STEP_A)
            g2_a    = min(entry_a * _R ** 2, CAP_STEP_A)
            loss_a  = round(entry_a + g1_a + g2_a, 2)
            pct_a   = round(loss_a / CAPITAL * 100, 1)

            # Sistema B: cap total
            entry_b = min(raw_entry, ENTRY_MAX_B)
            g1_b    = round(entry_b * _R,      2)
            g2_b    = round(entry_b * _R ** 2, 2)
            loss_b  = round(entry_b + g1_b + g2_b, 2)
            pct_b   = round(loss_b / CAPITAL * 100, 1)

            # Sistema C: fijo
            loss_c = FIXED_ENTRY + FIXED_G1 + FIXED_G2

            even = (r % 2 == 0)
            bg_a = ROJO_CLARO if pct_a > 15 else (AMARILLO if pct_a > 10 else VERDE_CLARO)
            bg_b = VERDE_CLARO
            bg_c = NARANJA_CL

            for col in range(1, 8):
                ws.cell(row=r, column=col).border    = thin_border()
                ws.cell(row=r, column=col).alignment = center()
                ws.cell(row=r, column=col).font      = Font(size=10)
                ws.cell(row=r, column=col).fill      = fill(GRIS_CLARO if even else BLANCO)

            ws.cell(row=r, column=1, value=f"{l}L / {w}W")
            ws.cell(row=r, column=2, value=round(entry_a, 2)).number_format = '"$"#,##0.00'
            ws.cell(row=r, column=3, value=round(g1_a, 2)).number_format = '"$"#,##0.00'
            ws.cell(row=r, column=4, value=round(g2_a, 2)).number_format = '"$"#,##0.00'

            cell_a = ws.cell(row=r, column=5, value=f"${loss_a:.2f} ({pct_a}%)")
            cell_a.fill = fill(bg_a)
            cell_a.font = Font(size=10, bold=True, color=ROJO_OSC if pct_a > 10 else VERDE_OSC)

            cell_b = ws.cell(row=r, column=6, value=f"${loss_b:.2f} ({pct_b}%)")
            cell_b.fill = fill(bg_b)
            cell_b.font = Font(size=10, bold=True, color=VERDE_OSC)

            cell_c = ws.cell(row=r, column=7, value=f"${loss_c:.2f} ({loss_c/CAPITAL*100:.1f}%)")
            cell_c.fill = fill(bg_c)
            cell_c.font = Font(size=10, bold=True)

            ws.row_dimensions[r].height = 18
            r += 1

    # Leyenda
    ws.row_dimensions[r].height = 8
    for label, bg_c in [
        ("🔵 A (bot actual): cap $30 por paso → pérdida puede ser 24.9% del capital", AZUL_CLARO),
        ("🟢 B (corregido): entry_max=$4.03 → pérdida SIEMPRE ≤ $30 (10% exacto)", VERDE_CLARO),
        ("🟠 C (fijo 2/4/10): pérdida máxima $16.00 (5.3% del capital)", NARANJA_CL),
    ]:
        r += 1
        ws.merge_cells(f"A{r}:G{r}")
        cell = ws[f"A{r}"]
        cell.value = label
        cell.fill  = fill(bg_c)
        cell.font  = Font(italic=True, size=10)
        cell.alignment = Alignment(horizontal="left")
        ws.row_dimensions[r].height = 20


# ═══════════════════════════════════════════════════════════════════════════
# 8. HOJAS 4 y 5 — DETALLE DIARIO A y B
# ═══════════════════════════════════════════════════════════════════════════

def sheet_diario(wb, result: SimResult, title: str, tab_color: str):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color

    col_widths = [12, 9, 12, 10, 12, 12, 12, 12]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(col_widths))}1")
    t = ws["A1"]
    t.value = title
    t.fill  = fill(AZUL_HEADER)
    t.font  = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    write_header(ws, 2, ["Fecha", "Sesiones", "P&L día", "Meta $60",
                          "Bal. Inicio", "Bal. Fin", "Bal. Min", "DD intradia"])

    dias = sorted(result.daily.keys(), key=lambda d: datetime.strptime(d, "%d/%m/%Y"))
    for r, day in enumerate(dias, 3):
        d    = result.daily[day]
        even = (r % 2 == 0)
        bg   = VERDE_CLARO if d.meta_hit else (ROJO_CLARO if d.pnl < 0 else (GRIS_CLARO if even else BLANCO))
        for c in range(1, len(col_widths) + 1):
            ws.cell(row=r, column=c).fill   = fill(bg)
            ws.cell(row=r, column=c).border = thin_border()
            ws.cell(row=r, column=c).font   = Font(size=10)
            ws.cell(row=r, column=c).alignment = center()

        dd_int = round((d.start_bal or 0) - (d.min_bal or d.start_bal or 0), 2)
        data = [day, d.sessions, d.pnl, "✔ META" if d.meta_hit else "✖",
                d.start_bal, d.end_bal, d.min_bal, dd_int]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (3, 5, 6, 7, 8):
                cell.number_format = '"$"#,##0.00'
            if c == 3:
                cell.font = Font(size=10, bold=True,
                                 color=VERDE_OSC if d.pnl >= 0 else ROJO_OSC)
            if c == 4:
                cell.font = Font(size=10, bold=True,
                                 color=VERDE_OSC if d.meta_hit else ROJO_OSC)
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main() -> None:
    print("Leyendo histórico Telegram...")
    outcomes = parse_outcomes(HISTORY)
    print(f"  {len(outcomes)} señales encontradas")

    # Sistema A: bot actual — cap por paso
    print("\nSimulando Sistema A (bot actual, cap $30/paso)...")
    def entry_a(l, w):
        raw = masaniello_raw(l, w)
        return min(raw, CAP_STEP_A)

    ra = _simulate_masaniello(outcomes, entry_a)
    print(f"  Balance final: ${ra.final:.2f}  MaxDD: ${ra.max_dd:.2f}")
    print(f"  Pérdida máx/señal: ${ra.max_single_loss:.2f} ({ra.max_single_loss/CAPITAL*100:.1f}% del capital)")

    # Sistema B: cap total corregido
    print("\nSimulando Sistema B (cap total ≤ $30)...")
    def entry_b(l, w):
        raw = masaniello_raw(l, w)
        return min(raw, ENTRY_MAX_B)

    rb = _simulate_masaniello(outcomes, entry_b)
    print(f"  Balance final: ${rb.final:.2f}  MaxDD: ${rb.max_dd:.2f}")
    print(f"  Pérdida máx/señal: ${rb.max_single_loss:.2f} ({rb.max_single_loss/CAPITAL*100:.1f}% del capital)")

    # Sistema C: fijo
    print("\nSimulando Sistema C (fijo $2/$4/$10)...")
    rc = _simulate_fixed(outcomes)
    print(f"  Balance final: ${rc.final:.2f}  MaxDD: ${rc.max_dd:.2f}")
    print(f"  Pérdida máx/señal: ${rc.max_single_loss:.2f} ({rc.max_single_loss/CAPITAL*100:.1f}% del capital)")

    print("\nGenerando Excel...")
    wb = openpyxl.Workbook()

    sheet_kpis(wb, ra, rb, rc)
    sheet_balance(wb, ra, rb, rc)
    sheet_riesgo(wb)
    sheet_diario(wb, ra, "Detalle A — Bot actual", "2E75B6")
    sheet_diario(wb, rb, "Detalle B — Cap total", "1E6F4B")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\n✔  Comparativa guardada en: {OUTPUT}")

    print()
    print("=" * 65)
    print("RESUMEN FINAL — COMPARATIVA 3 SISTEMAS")
    print("=" * 65)
    dias_a = len(ra.daily); meta_a = sum(1 for d in ra.daily.values() if d.meta_hit)
    dias_b = len(rb.daily); meta_b = sum(1 for d in rb.daily.values() if d.meta_hit)
    dias_c = len(rc.daily); meta_c = sum(1 for d in rc.daily.values() if d.meta_hit)

    for name, r, dias, meta in [
        ("A — Bot actual (12/4, cap/paso)", ra, dias_a, meta_a),
        ("B — Cap total ≤$30 (12/4, corr.)", rb, dias_b, meta_b),
        ("C — Fijo $2/$4/$10",             rc, dias_c, meta_c),
    ]:
        roi = (r.final - CAPITAL) / CAPITAL * 100
        print(f"\n  {name}")
        print(f"    Balance:   ${r.final:.2f}  (ROI: {roi:+.1f}%)")
        print(f"    MaxDD:     ${r.max_dd:.2f}  ({r.max_dd/CAPITAL*100:.1f}%)")
        print(f"    Pérd/señal:${r.max_single_loss:.2f}  ({r.max_single_loss/CAPITAL*100:.1f}%)")
        print(f"    Meta $60:  {meta}/{dias} días")


if __name__ == "__main__":
    main()
