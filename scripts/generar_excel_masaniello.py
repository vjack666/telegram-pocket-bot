"""
Genera un Excel con los 64 patrones posibles de una sesion Masaniello 6/2.
Para cada patron muestra operacion a operacion: stake, resultado, saldo,
drawdown, y metricas de estabilidad/riesgo/recuperacion.

Salida: runtime/masaniello_escenarios.xlsx
"""
from __future__ import annotations

import sys
from itertools import product
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.sim_objetivo_60_por_dia import masaniello_stake

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── Parametros ──────────────────────────────────────────────────────────────
CAPITAL    = 300.0
N_OPS      = 6
W_NEEDED   = 2
PAYOUT     = 1.92       # payout multiplicador (92%)
META_PNL   = 60.0       # objetivo de ganancia de la sesion tipo

# ── Colores ──────────────────────────────────────────────────────────────────
C_HEADER   = "1565C0"   # azul oscuro
C_WIN      = "C8E6C9"   # verde claro
C_LOSS     = "FFCDD2"   # rojo claro
C_NEUTRAL  = "E3F2FD"   # azul muy claro
C_META_OK  = "A5D6A7"   # verde medio
C_META_NO  = "EF9A9A"   # rojo medio
C_STABLE   = "B2EBF2"   # cian
C_AGRESIVE = "FFE082"   # amarillo
C_DANGER   = "FF8A65"   # naranja
C_DEAD     = "E57373"   # rojo fuerte
C_SKIP     = "F5F5F5"   # gris muy claro

def fmt(v):
    return PatternFill("solid", fgColor=v)

thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(ws, row, col, value, fill=None, bold=False, number_fmt=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fmt(fill)
    if bold:
        c.font = Font(bold=True, color="FFFFFF" if fill == C_HEADER else "000000")
    c.border = border
    c.alignment = Alignment(horizontal=align, vertical="center")
    if number_fmt:
        c.number_format = number_fmt
    return c

# ── Generar los 64 patrones ───────────────────────────────────────────────────
def simulate_pattern(pattern: tuple[str, ...]) -> dict:
    """
    Simula una sesion Masaniello 6/2 con un patron de W/L dado.
    El bot para al llegar a W_NEEDED victorias o cuando no puede ganar mas
    con las operaciones restantes.
    Devuelve dict con datos por operacion y resumen.
    """
    balance   = CAPITAL
    wins = losses = 0
    ops: list[dict] = []
    peak = CAPITAL
    max_dd = 0.0

    for i, result in enumerate(pattern):
        wins_needed = W_NEEDED - wins
        ops_left    = N_OPS - (wins + losses)
        # Condicion de parada
        if wins_needed <= 0 or wins_needed > ops_left:
            break

        stake  = masaniello_stake(CAPITAL, losses, wins, N_OPS, W_NEEDED, PAYOUT)
        is_win = result == "W"
        pnl_op = round(stake * (PAYOUT - 1), 2) if is_win else round(-stake, 2)
        balance_before = balance
        balance        = round(balance + pnl_op, 2)
        peak           = max(peak, balance)
        dd             = round(peak - balance, 2)
        max_dd         = max(max_dd, dd)

        wins   += is_win
        losses += not is_win

        ops.append({
            "op_num":         i + 1,
            "result":         result,
            "stake":          stake,
            "pnl_op":         pnl_op,
            "balance_before": balance_before,
            "balance_after":  balance,
            "drawdown":       dd,
            "wins_acc":       wins,
            "losses_acc":     losses,
        })

    total_pnl    = round(balance - CAPITAL, 2)
    meta_hit     = wins >= W_NEEDED
    survival     = balance > 0
    win_rate     = wins / len(ops) * 100 if ops else 0
    ops_executed = len(ops)

    # Nivel de riesgo: bajo/medio/alto/critico
    if max_dd == 0:
        risk = "Bajo"
        risk_color = C_STABLE
    elif max_dd <= CAPITAL * 0.10:
        risk = "Moderado"
        risk_color = C_STABLE
    elif max_dd <= CAPITAL * 0.20:
        risk = "Medio"
        risk_color = C_AGRESIVE
    elif max_dd <= CAPITAL * 0.30:
        risk = "Alto"
        risk_color = C_DANGER
    else:
        risk = "Critico"
        risk_color = C_DEAD

    # Clasificacion patron
    if meta_hit and max_dd == 0:
        tipo = "Estable"
    elif meta_hit and max_dd <= CAPITAL * 0.10:
        tipo = "Rentable"
    elif meta_hit:
        tipo = "Agresivo"
    elif not meta_hit and survival:
        tipo = "Incompleto"
    else:
        tipo = "Peligroso"

    return {
        "ops":          ops,
        "total_pnl":    total_pnl,
        "meta_hit":     meta_hit,
        "max_dd":       max_dd,
        "win_rate":     win_rate,
        "ops_executed": ops_executed,
        "wins":         wins,
        "losses":       losses,
        "final_bal":    balance,
        "risk":         risk,
        "risk_color":   risk_color,
        "tipo":         tipo,
        "survival":     survival,
    }

all_patterns = list(product("WL", repeat=N_OPS))

# ── Crear workbook ─────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1: DETALLE OPERACION POR OPERACION
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Detalle Operaciones"
ws1.freeze_panes = "A3"

# Cabecera titulo
ws1.merge_cells("A1:V1")
t = ws1.cell(row=1, column=1,
    value=f"MASANIELLO 6/2 — 64 ESCENARIOS POSIBLES — Capital ${CAPITAL:.0f} | Payout {(PAYOUT-1)*100:.0f}% | Meta ${META_PNL:.0f}")
t.font = Font(bold=True, size=13, color="FFFFFF")
t.fill = fmt(C_HEADER)
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 24

# Sub-cabeceras
headers = [
    "#", "Patron", "Tipo",
    "Op1 Stake","Op1","Saldo1",
    "Op2 Stake","Op2","Saldo2",
    "Op3 Stake","Op3","Saldo3",
    "Op4 Stake","Op4","Saldo4",
    "Op5 Stake","Op5","Saldo5",
    "Op6 Stake","Op6","Saldo6",
    "Ops Ejecutadas", "W", "L",
    "PnL Sesion", "Saldo Final",
    "Drawdown Max", "Riesgo",
    "Meta $60?", "Supervivencia",
]

for col, h in enumerate(headers, start=1):
    c = ws1.cell(row=2, column=col, value=h)
    c.fill = fmt(C_HEADER)
    c.font = Font(bold=True, color="FFFFFF", size=9)
    c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 30

# Datos
for row_num, pattern in enumerate(all_patterns, start=1):
    sim   = simulate_pattern(pattern)
    r     = row_num + 2
    pat_str = "".join(pattern)

    # Colores de fila segun tipo
    row_fill = {
        "Estable":     C_STABLE,
        "Rentable":    C_META_OK,
        "Agresivo":    C_AGRESIVE,
        "Incompleto":  C_NEUTRAL,
        "Peligroso":   C_DANGER,
    }.get(sim["tipo"], "FFFFFF")

    col = 1
    cell_style(ws1, r, col, row_num,        fill=row_fill); col += 1
    cell_style(ws1, r, col, pat_str,        fill=row_fill, bold=True); col += 1
    cell_style(ws1, r, col, sim["tipo"],    fill=row_fill, bold=True); col += 1

    # Operaciones (hasta 6, celdas vacias si no se ejecutaron)
    for op_idx in range(N_OPS):
        if op_idx < len(sim["ops"]):
            op = sim["ops"][op_idx]
            fill_stake = C_WIN if op["result"] == "W" else C_LOSS
            cell_style(ws1, r, col,     f"${op['stake']:.2f}", fill=fill_stake,  number_fmt='"$"#,##0.00'); col += 1
            cell_style(ws1, r, col,     op["result"],          fill=fill_stake); col += 1
            cell_style(ws1, r, col,     op["balance_after"],   fill=fill_stake,  number_fmt='"$"#,##0.00'); col += 1
        else:
            for _ in range(3):
                cell_style(ws1, r, col, "—", fill=C_SKIP); col += 1

    cell_style(ws1, r, col, sim["ops_executed"],                      fill=row_fill); col += 1
    cell_style(ws1, r, col, sim["wins"],                              fill=C_WIN); col += 1
    cell_style(ws1, r, col, sim["losses"],                            fill=C_LOSS); col += 1
    cell_style(ws1, r, col, sim["total_pnl"],   number_fmt='"$"#,##0.00', fill=C_META_OK if sim["total_pnl"] >= 0 else C_META_NO); col += 1
    cell_style(ws1, r, col, sim["final_bal"],   number_fmt='"$"#,##0.00', fill=row_fill); col += 1
    cell_style(ws1, r, col, sim["max_dd"],      number_fmt='"$"#,##0.00', fill=sim["risk_color"]); col += 1
    cell_style(ws1, r, col, sim["risk"],        fill=sim["risk_color"]); col += 1
    cell_style(ws1, r, col, "SI" if sim["meta_hit"]  else "NO",  fill=C_META_OK if sim["meta_hit"]  else C_META_NO); col += 1
    cell_style(ws1, r, col, "SI" if sim["survival"]  else "NO",  fill=C_META_OK if sim["survival"]  else C_DEAD); col += 1

# Anchos de columna
ws1.column_dimensions["A"].width = 5
ws1.column_dimensions["B"].width = 10
ws1.column_dimensions["C"].width = 12
for i in range(4, 22):
    ws1.column_dimensions[get_column_letter(i)].width = 10
for i in range(22, len(headers) + 1):
    ws1.column_dimensions[get_column_letter(i)].width = 13

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2: RESUMEN Y CLASIFICACION
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resumen y Clasificacion")
ws2.freeze_panes = "A3"

ws2.merge_cells("A1:L1")
t2 = ws2.cell(row=1, column=1,
    value="RESUMEN — Clasificacion de los 64 patrones Masaniello 6/2")
t2.font = Font(bold=True, size=13, color="FFFFFF")
t2.fill = fmt(C_HEADER)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 24

headers2 = [
    "#", "Patron", "Tipo", "Ops Ejecutadas",
    "W", "L", "PnL ($)", "Saldo Final ($)",
    "Drawdown Max ($)", "DD como % Capital",
    "Riesgo", "Meta $60?",
]
for col, h in enumerate(headers2, start=1):
    c = ws2.cell(row=2, column=col, value=h)
    c.fill = fmt(C_HEADER)
    c.font = Font(bold=True, color="FFFFFF", size=9)
    c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[2].height = 30

for row_num, pattern in enumerate(all_patterns, start=1):
    sim = simulate_pattern(pattern)
    r   = row_num + 2
    pat_str = "".join(pattern)

    row_fill = {
        "Estable":     C_STABLE,
        "Rentable":    C_META_OK,
        "Agresivo":    C_AGRESIVE,
        "Incompleto":  C_NEUTRAL,
        "Peligroso":   C_DANGER,
    }.get(sim["tipo"], "FFFFFF")

    vals = [
        row_num, pat_str, sim["tipo"], sim["ops_executed"],
        sim["wins"], sim["losses"], sim["total_pnl"], sim["final_bal"],
        sim["max_dd"], round(sim["max_dd"] / CAPITAL * 100, 1),
        sim["risk"], "SI" if sim["meta_hit"] else "NO",
    ]
    fills = [
        row_fill, row_fill, row_fill, row_fill,
        C_WIN, C_LOSS,
        C_META_OK if sim["total_pnl"] >= 0 else C_META_NO,
        row_fill, sim["risk_color"], sim["risk_color"],
        sim["risk_color"],
        C_META_OK if sim["meta_hit"] else C_META_NO,
    ]
    fmts = [
        None, None, None, None,
        None, None, '"$"#,##0.00', '"$"#,##0.00',
        '"$"#,##0.00', '0.0"%"',
        None, None,
    ]
    for col, (v, f, nf) in enumerate(zip(vals, fills, fmts), start=1):
        cell_style(ws2, r, col, v, fill=f, number_fmt=nf)

for i, w in enumerate([5, 10, 12, 14, 6, 6, 12, 14, 14, 14, 10, 10], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════════
# HOJA 3: ESTADISTICAS GLOBALES
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Estadisticas Globales")

all_sims = [simulate_pattern(p) for p in all_patterns]

ws3.merge_cells("A1:D1")
t3 = ws3.cell(row=1, column=1, value="ESTADISTICAS GLOBALES — 64 ESCENARIOS")
t3.font = Font(bold=True, size=13, color="FFFFFF")
t3.fill = fmt(C_HEADER)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 24

def stat_row(ws, r, label, value, fmt_str=None, fill="FFFFFF"):
    cell_style(ws, r, 1, label, fill=fill, bold=True, align="left")
    c = ws.cell(row=r, column=2, value=value)
    c.fill = fmt(fill)
    c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center")
    if fmt_str:
        c.number_format = fmt_str

metas_ok   = sum(1 for s in all_sims if s["meta_hit"])
tipos      = {}
for s in all_sims:
    tipos[s["tipo"]] = tipos.get(s["tipo"], 0) + 1

riesgos = {}
for s in all_sims:
    riesgos[s["risk"]] = riesgos.get(s["risk"], 0) + 1

pnls   = [s["total_pnl"] for s in all_sims]
dds    = [s["max_dd"]    for s in all_sims]

r = 3
ws3.cell(row=r, column=1, value="PARAMETROS DE LA SESION").font = Font(bold=True, size=11)
r += 1
for lbl, val, nf, fl in [
    ("Capital inicial",         CAPITAL,     '"$"#,##0.00', C_NEUTRAL),
    ("N operaciones (max)",     N_OPS,       None,          C_NEUTRAL),
    ("Victorias necesarias",    W_NEEDED,    None,          C_NEUTRAL),
    ("Payout multiplicador",    PAYOUT,      '0.00"x"',     C_NEUTRAL),
    ("Payout %",                (PAYOUT-1)*100, '0"%"',     C_NEUTRAL),
    ("Meta ganancia sesion",    META_PNL,    '"$"#,##0.00', C_NEUTRAL),
    ("Stake inicial (0L/0W)",   masaniello_stake(CAPITAL,0,0,N_OPS,W_NEEDED,PAYOUT), '"$"#,##0.00', C_NEUTRAL),
    ("Stake tras 1 perdida",    masaniello_stake(CAPITAL,1,0,N_OPS,W_NEEDED,PAYOUT), '"$"#,##0.00', C_NEUTRAL),
    ("Stake tras 2 perdidas",   masaniello_stake(CAPITAL,2,0,N_OPS,W_NEEDED,PAYOUT), '"$"#,##0.00', C_NEUTRAL),
]:
    stat_row(ws3, r, lbl, val, nf, fl); r += 1

r += 1
ws3.cell(row=r, column=1, value="RESULTADOS GLOBALES").font = Font(bold=True, size=11)
r += 1
for lbl, val, nf, fl in [
    ("Total patrones analizados",      64,               None,            C_NEUTRAL),
    ("Patrones que cumplen meta",      metas_ok,         None,            C_META_OK),
    ("Patrones que NO cumplen meta",   64 - metas_ok,    None,            C_META_NO),
    ("% patrones exitosos",            metas_ok/64*100,  '0.0"%"',        C_META_OK),
    ("PnL promedio por patron",        sum(pnls)/64,     '"$"#,##0.00',   C_NEUTRAL),
    ("PnL maximo (mejor patron)",      max(pnls),        '"$"#,##0.00',   C_META_OK),
    ("PnL minimo (peor patron)",       min(pnls),        '"$"#,##0.00',   C_META_NO),
    ("Drawdown promedio",              sum(dds)/64,      '"$"#,##0.00',   C_AGRESIVE),
    ("Drawdown maximo",                max(dds),         '"$"#,##0.00',   C_DANGER),
]:
    stat_row(ws3, r, lbl, val, nf, fl); r += 1

r += 1
ws3.cell(row=r, column=1, value="CLASIFICACION POR TIPO").font = Font(bold=True, size=11)
r += 1
tipo_colors = {"Estable": C_STABLE, "Rentable": C_META_OK, "Agresivo": C_AGRESIVE,
               "Incompleto": C_NEUTRAL, "Peligroso": C_DANGER}
for tipo, count in sorted(tipos.items(), key=lambda x: -x[1]):
    stat_row(ws3, r, tipo, count, fill=tipo_colors.get(tipo, "FFFFFF")); r += 1

r += 1
ws3.cell(row=r, column=1, value="CLASIFICACION POR NIVEL DE RIESGO").font = Font(bold=True, size=11)
r += 1
riesgo_colors = {"Bajo": C_STABLE, "Moderado": C_STABLE, "Medio": C_AGRESIVE,
                 "Alto": C_DANGER, "Critico": C_DEAD}
for riesgo, count in sorted(riesgos.items(), key=lambda x: -x[1]):
    stat_row(ws3, r, riesgo, count, fill=riesgo_colors.get(riesgo, "FFFFFF")); r += 1

ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 16

# ── Guardar ────────────────────────────────────────────────────────────────
out = ROOT / "runtime" / "masaniello_escenarios.xlsx"
wb.save(out)
print(f"Excel guardado: {out}")
print(f"  Hoja 1: Detalle operacion por operacion (64 filas x {len(headers)} cols)")
print(f"  Hoja 2: Resumen y clasificacion")
print(f"  Hoja 3: Estadisticas globales")
print()
print(f"  Patrones que cumplen meta $60: {metas_ok}/64 ({metas_ok/64*100:.1f}%)")
print(f"  Tipos:")
for tipo, count in sorted(tipos.items(), key=lambda x: -x[1]):
    print(f"    {tipo:<12}: {count}")
