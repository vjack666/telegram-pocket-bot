from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
HISTORY = ROOT / "ejemplo.md"
OUTPUT = ROOT / "runtime" / "informe_maestro.xlsx"

# Trading constants
PAYOUT_MULT = 1.92
NET_PAYOUT = 0.92
R = PAYOUT_MULT / NET_PAYOUT
TOTAL_MULT = 1 + R + R**2

# Session architecture requested by user
CAPITAL = 300.0
CAP_PCT = 0.10
SIGNALS_PER_SESSION = 6
TARGET_WINS_PER_SESSION = 2
MAX_LOSSES_PER_SESSION = 3
SESSIONS_PER_DAY_MAX = 6
TARGET_SESSION_PROFIT = 9.0  # target range 8-10
RECOVERY_FACTORS = (1.0, 1.6, 2.2)

# Styles
BLUE_DARK = "1F3864"
BLUE = "2E75B6"
GREEN = "1E6F4B"
RED = "8B1A1A"
GRAY = "F2F2F2"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"
GREEN_BG = "D6F0E0"
RED_BG = "FDDEDE"


@dataclass
class Outcome:
    timestamp: datetime
    result: str
    is_win: bool


@dataclass
class SessionResult:
    date: str
    session_index: int
    signals_used: int
    wins: int
    losses: int
    session_pnl: float
    start_balance: float
    end_balance: float
    stop_reason: str


@dataclass
class SimSummary:
    initial_balance: float
    final_balance: float
    max_drawdown: float
    total_days: int
    total_sessions: int
    won_sessions: int
    protected_sessions: int
    total_signals_used: int
    avg_signals_per_day: float
    avg_pnl_per_day: float
    days_target_40: int
    days_target_60: int
    wr_history: float


def parse_outcomes(path: Path) -> list[Outcome]:
    date_pat = re.compile(r"^\[(\d{2}/\d{2}/\d{4}) (\d{2}:\d{2}:\d{2})\]")
    res_pat = re.compile(
        r"(VICTORIA DIRECTA|VICTORIA EN 1.*?MARTINGALA|VICTORIA EN 2.*?MARTINGALA|P[EE]RDIDA|P[ÉE]RDIDA)",
        re.IGNORECASE,
    )
    label_map = {
        "victoria directa": "WD",
        "victoria en 1": "G1",
        "victoria en 2": "G2",
        "perdida": "L",
        "perdida": "L",
        "pérdida": "L",
    }

    outcomes: list[Outcome] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        d = date_pat.match(line)
        if not d:
            continue
        r = res_pat.search(line)
        if not r:
            continue
        ts = datetime.strptime(" ".join(d.groups()), "%d/%m/%Y %H:%M:%S")
        raw = r.group(1).lower()
        label = "?"
        for key, val in label_map.items():
            if raw.startswith(key):
                label = val
                break
        if label == "?":
            continue
        outcomes.append(Outcome(timestamp=ts, result=label, is_win=(label != "L")))
    return outcomes


def scan_project(root: Path) -> list[dict]:
    rows = []
    valid_ext = {".py", ".md", ".env", ".txt", ".session", ".json", ".jsonl"}
    for p in root.rglob("*"):
        if not p.is_file():
            continue
        if "build" in p.parts:
            continue
        if p.suffix.lower() not in valid_ext and p.name != ".env":
            continue
        try:
            content = p.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        low = content.lower()
        rows.append(
            {
                "file": str(p.relative_to(root)).replace("\\", "/"),
                "lines": len(content.splitlines()),
                "size_kb": round(len(content.encode("utf-8")) / 1024.0, 2),
                "masaniello": int("masaniello" in low),
                "equity": int("equity" in low),
                "martingale": int(("martingale" in low) or ("gale" in low)),
                "signal": int(("signal" in low) or ("señal" in low) or ("senal" in low)),
            }
        )
    rows.sort(key=lambda x: x["file"])
    return rows


def chunk_by_day_and_session(outcomes: list[Outcome], signals_per_session: int) -> dict[str, list[list[Outcome]]]:
    by_day: dict[str, list[Outcome]] = defaultdict(list)
    for o in outcomes:
        by_day[o.timestamp.strftime("%d/%m/%Y")].append(o)

    sessions_by_day: dict[str, list[list[Outcome]]] = {}
    for day, day_outcomes in by_day.items():
        chunks: list[list[Outcome]] = []
        for i in range(0, len(day_outcomes), signals_per_session):
            chunks.append(day_outcomes[i : i + signals_per_session])
        sessions_by_day[day] = chunks
    return sessions_by_day


def simulate_2wins_architecture(outcomes: list[Outcome]) -> tuple[list[SessionResult], dict[str, float], SimSummary]:
    sessions_by_day = chunk_by_day_and_session(outcomes, SIGNALS_PER_SESSION)

    entry_cap_max = CAPITAL * CAP_PCT / TOTAL_MULT
    max_factor = max(RECOVERY_FACTORS)
    # Base stake chosen to allow recovery steps while keeping hard total cap.
    base_entry = min(TARGET_SESSION_PROFIT / TARGET_WINS_PER_SESSION / NET_PAYOUT, entry_cap_max / max_factor)

    balance = CAPITAL
    peak = CAPITAL
    max_dd = 0.0

    sessions_results: list[SessionResult] = []
    daily_pnl: dict[str, float] = defaultdict(float)
    total_signals_used = 0

    for day in sorted(sessions_by_day.keys(), key=lambda d: datetime.strptime(d, "%d/%m/%Y")):
        day_sessions = sessions_by_day[day][:SESSIONS_PER_DAY_MAX]
        for idx, chunk in enumerate(day_sessions, start=1):
            start_balance = balance
            wins = 0
            losses = 0
            used = 0
            session_pnl = 0.0
            loss_streak = 0
            stop_reason = "session_exhausted"

            for outcome in chunk:
                if wins >= TARGET_WINS_PER_SESSION:
                    stop_reason = "target_wins_reached"
                    break
                if losses >= MAX_LOSSES_PER_SESSION:
                    stop_reason = "loss_guard"
                    break

                factor = RECOVERY_FACTORS[min(loss_streak, len(RECOVERY_FACTORS) - 1)]
                entry = round(min(base_entry * factor, entry_cap_max), 2)
                if entry <= 0:
                    stop_reason = "entry_zero"
                    break

                used += 1
                total_signals_used += 1

                if outcome.is_win:
                    gain = round(entry * NET_PAYOUT, 2)
                    balance = round(balance + gain, 2)
                    session_pnl = round(session_pnl + gain, 2)
                    wins += 1
                    loss_streak = 0
                else:
                    total_loss = round(entry * TOTAL_MULT, 2)
                    balance = round(balance - total_loss, 2)
                    session_pnl = round(session_pnl - total_loss, 2)
                    losses += 1
                    loss_streak += 1

                peak = max(peak, balance)
                max_dd = max(max_dd, round(peak - balance, 2))

            if wins >= TARGET_WINS_PER_SESSION:
                stop_reason = "target_wins_reached"
            elif losses >= MAX_LOSSES_PER_SESSION:
                stop_reason = "loss_guard"

            daily_pnl[day] = round(daily_pnl[day] + session_pnl, 2)
            sessions_results.append(
                SessionResult(
                    date=day,
                    session_index=idx,
                    signals_used=used,
                    wins=wins,
                    losses=losses,
                    session_pnl=session_pnl,
                    start_balance=round(start_balance, 2),
                    end_balance=round(balance, 2),
                    stop_reason=stop_reason,
                )
            )

    total_days = len(daily_pnl)
    total_sessions = len(sessions_results)
    won_sessions = sum(1 for s in sessions_results if s.wins >= TARGET_WINS_PER_SESSION)
    protected_sessions = sum(1 for s in sessions_results if s.stop_reason == "loss_guard")
    days_target_40 = sum(1 for p in daily_pnl.values() if p >= 40.0)
    days_target_60 = sum(1 for p in daily_pnl.values() if p >= 60.0)

    wr_history = (sum(1 for o in outcomes if o.is_win) / len(outcomes) * 100.0) if outcomes else 0.0

    summary = SimSummary(
        initial_balance=CAPITAL,
        final_balance=round(balance, 2),
        max_drawdown=round(max_dd, 2),
        total_days=total_days,
        total_sessions=total_sessions,
        won_sessions=won_sessions,
        protected_sessions=protected_sessions,
        total_signals_used=total_signals_used,
        avg_signals_per_day=round(total_signals_used / total_days, 2) if total_days else 0.0,
        avg_pnl_per_day=round((balance - CAPITAL) / total_days, 2) if total_days else 0.0,
        days_target_40=days_target_40,
        days_target_60=days_target_60,
        wr_history=round(wr_history, 2),
    )

    return sessions_results, dict(sorted(daily_pnl.items(), key=lambda x: datetime.strptime(x[0], "%d/%m/%Y"))), summary


def border_thin() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def header_cell(cell, bg=BLUE_DARK) -> None:
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.alignment = center()
    cell.border = border_thin()


def body_cell(cell, even: bool) -> None:
    cell.fill = PatternFill("solid", fgColor=(GRAY if even else WHITE))
    cell.font = Font(size=10)
    cell.alignment = center()
    cell.border = border_thin()


def build_excel(
    scan_rows: list[dict],
    sessions: list[SessionResult],
    daily_pnl: dict[str, float],
    summary: SimSummary,
    outcomes: list[Outcome],
) -> None:
    wb = openpyxl.Workbook()

    # Sheet 1: executive summary
    ws = wb.active
    ws.title = "Resumen_Ejecutivo"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 44

    ws.merge_cells("A1:C1")
    ws["A1"] = "INFORME MAESTRO - SISTEMA SESIONES 2 WINS"
    header_cell(ws["A1"])  # type: ignore[arg-type]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    ws["A2"] = "Base: datos reales de ejemplo.md + escaneo del codigo del proyecto"
    ws["A2"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A2"].font = Font(color=WHITE, size=10)
    ws["A2"].alignment = center()

    rows = [
        ("Capital base", f"${summary.initial_balance:,.2f}", "Modo seguro con cap total por senal"),
        ("WR historico", f"{summary.wr_history:.2f}%", "Calculado sobre todas las senales parseadas"),
        ("Balance final simulado", f"${summary.final_balance:,.2f}", "Arquitectura: 2 wins por sesion"),
        ("PnL promedio diario", f"${summary.avg_pnl_per_day:+.2f}", "Objetivo recomendado: 40-60 por dia"),
        ("Max drawdown", f"${summary.max_drawdown:,.2f}", "Controlado por cap% + guard de perdidas"),
        ("Dias >= $40", f"{summary.days_target_40}/{summary.total_days}", "Meta conservadora"),
        ("Dias >= $60", f"{summary.days_target_60}/{summary.total_days}", "Meta objetivo"),
        ("Senales promedio por dia", f"{summary.avg_signals_per_day}", "Deberia caer mucho vs operar todo"),
        (
            "Sesiones ganadas",
            f"{summary.won_sessions}/{summary.total_sessions}",
            "Sesion ganada = alcanza 2 wins antes del corte",
        ),
        (
            "Sesiones protegidas",
            str(summary.protected_sessions),
            "Corte por maximo de perdidas en sesion",
        ),
    ]

    ws["A4"] = "KPI"
    ws["B4"] = "Valor"
    ws["C4"] = "Lectura"
    for c in ("A4", "B4", "C4"):
        header_cell(ws[c])

    for i, (k, v, obs) in enumerate(rows, start=5):
        even = (i % 2 == 0)
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
        ws[f"C{i}"] = obs
        for c in (f"A{i}", f"B{i}", f"C{i}"):
            body_cell(ws[c], even)

    verdict = "NEGATIVO"
    verdict_bg = RED_BG
    if summary.avg_pnl_per_day > 0 and summary.days_target_40 >= int(summary.total_days * 0.5):
        verdict = "OPERABLE"
        verdict_bg = GREEN_BG

    ws.merge_cells("A17:C18")
    ws["A17"] = (
        f"VEREDICTO: {verdict}.\n"
        "Esta arquitectura es mejor para control de riesgo porque deja de operar al cumplir objetivo de sesion."
    )
    ws["A17"].fill = PatternFill("solid", fgColor=verdict_bg)
    ws["A17"].font = Font(bold=True, size=11, color=(GREEN if verdict == "OPERABLE" else RED))
    ws["A17"].alignment = Alignment(wrap_text=True, vertical="center")

    # Sheet 2: config
    cfg = wb.create_sheet("Parametros")
    cfg.sheet_view.showGridLines = False
    cfg.column_dimensions["A"].width = 28
    cfg.column_dimensions["B"].width = 22
    cfg.column_dimensions["C"].width = 38

    cfg["A1"] = "Parametro"
    cfg["B1"] = "Valor"
    cfg["C1"] = "Comentario"
    for c in ("A1", "B1", "C1"):
        header_cell(cfg[c])

    param_rows = [
        ("Capital", CAPITAL, "Base operativa"),
        ("Cap total por senal", CAP_PCT, "Riesgo maximo total por senal"),
        ("Senales por sesion", SIGNALS_PER_SESSION, "Bloque esperado del grupo"),
        ("Objetivo wins/sesion", TARGET_WINS_PER_SESSION, "Stop al llegar a este valor"),
        ("Max perdidas/sesion", MAX_LOSSES_PER_SESSION, "Proteccion de sesion"),
        ("Sesiones max/dia", SESSIONS_PER_DAY_MAX, "Cap de actividad diaria"),
        ("Target profit sesion", TARGET_SESSION_PROFIT, "Rango objetivo 8-10"),
        ("Recovery factors", ", ".join(str(x) for x in RECOVERY_FACTORS), "Recuperacion suave"),
        ("Payout mult", PAYOUT_MULT, "1 + net payout"),
        ("Net payout", NET_PAYOUT, "Ganancia neta por unidad arriesgada"),
        ("Exposure multiplier", round(TOTAL_MULT, 4), "entry + g1 + g2"),
        (
            "Entry cap max",
            round(CAPITAL * CAP_PCT / TOTAL_MULT, 2),
            "Entry maxima para no romper cap total",
        ),
    ]

    for i, row in enumerate(param_rows, start=2):
        even = (i % 2 == 0)
        for j, val in enumerate(row, start=1):
            cell = cfg.cell(row=i, column=j, value=val)
            body_cell(cell, even)

    # Sheet 3: simulation detail
    sim = wb.create_sheet("Simulacion_Sesiones")
    sim.sheet_view.showGridLines = False
    sim_cols = [
        "Fecha",
        "Sesion",
        "Senales usadas",
        "Wins",
        "Losses",
        "PnL sesion",
        "Balance inicio",
        "Balance fin",
        "Stop reason",
    ]
    for i, name in enumerate(sim_cols, start=1):
        cell = sim.cell(row=1, column=i, value=name)
        header_cell(cell)
        sim.column_dimensions[get_column_letter(i)].width = 16

    for r, s in enumerate(sessions, start=2):
        even = (r % 2 == 0)
        values = [
            s.date,
            s.session_index,
            s.signals_used,
            s.wins,
            s.losses,
            s.session_pnl,
            s.start_balance,
            s.end_balance,
            s.stop_reason,
        ]
        for c, v in enumerate(values, start=1):
            cell = sim.cell(row=r, column=c, value=v)
            body_cell(cell, even)
            if c in (6, 7, 8):
                cell.number_format = '"$"#,##0.00'
            if c == 6:
                cell.font = Font(size=10, bold=True, color=(GREEN if float(v) >= 0 else RED))

    sim.freeze_panes = "A2"

    # Sheet 4: code scan
    code = wb.create_sheet("Codigo_Sistema")
    code.sheet_view.showGridLines = False
    scan_cols = ["file", "lines", "size_kb", "masaniello", "equity", "martingale", "signal"]
    widths = [58, 10, 10, 12, 10, 12, 10]
    for i, name in enumerate(scan_cols, start=1):
        cell = code.cell(row=1, column=i, value=name)
        header_cell(cell)
        code.column_dimensions[get_column_letter(i)].width = widths[i - 1]

    for r, row in enumerate(scan_rows, start=2):
        even = (r % 2 == 0)
        for c, key in enumerate(scan_cols, start=1):
            cell = code.cell(row=r, column=c, value=row[key])
            body_cell(cell, even)

    # Sheet 5: charts and daily summary
    ch = wb.create_sheet("Graficas")
    ch.sheet_view.showGridLines = False
    ch.column_dimensions["A"].width = 14
    ch.column_dimensions["B"].width = 18
    ch.column_dimensions["C"].width = 18
    ch.column_dimensions["D"].width = 18

    ch["A1"] = "Fecha"
    ch["B1"] = "PnL diario"
    ch["C1"] = "Balance acumulado"
    ch["D1"] = "Senales usadas"
    for c in ("A1", "B1", "C1", "D1"):
        header_cell(ch[c])

    balance = CAPITAL
    signals_by_day = Counter()
    for s in sessions:
        signals_by_day[s.date] += s.signals_used

    dates = list(daily_pnl.keys())
    for i, day in enumerate(dates, start=2):
        pnl = daily_pnl[day]
        balance = round(balance + pnl, 2)
        ch.cell(row=i, column=1, value=day)
        ch.cell(row=i, column=2, value=pnl)
        ch.cell(row=i, column=3, value=balance)
        ch.cell(row=i, column=4, value=signals_by_day.get(day, 0))
        for c in range(1, 5):
            body_cell(ch.cell(row=i, column=c), i % 2 == 0)
        ch.cell(row=i, column=2).number_format = '"$"#,##0.00'
        ch.cell(row=i, column=3).number_format = '"$"#,##0.00'

    if dates:
        line = LineChart()
        line.title = "Equity Curve (balance diario acumulado)"
        line.style = 10
        line.height = 10
        line.width = 20
        line.y_axis.title = "Balance ($)"
        line.x_axis.title = "Dia"
        data = Reference(ch, min_col=3, min_row=1, max_row=1 + len(dates))
        cats = Reference(ch, min_col=1, min_row=2, max_row=1 + len(dates))
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        ch.add_chart(line, "F2")

        bar = BarChart()
        bar.title = "PnL diario"
        bar.style = 10
        bar.height = 10
        bar.width = 20
        bar.y_axis.title = "PnL ($)"
        bar.x_axis.title = "Dia"
        d2 = Reference(ch, min_col=2, min_row=1, max_row=1 + len(dates))
        bar.add_data(d2, titles_from_data=True)
        bar.set_categories(cats)
        ch.add_chart(bar, "F22")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)


def main() -> None:
    print("Escaneando proyecto...")
    scan_rows = scan_project(ROOT)
    print(f"  Archivos analizados: {len(scan_rows)}")

    print("Leyendo historico real de senales...")
    outcomes = parse_outcomes(HISTORY)
    if not outcomes:
        raise RuntimeError("No se encontraron senales validas en ejemplo.md")
    print(f"  Senales parseadas: {len(outcomes)}")

    print("Simulando arquitectura de sesiones 2-wins...")
    sessions, daily_pnl, summary = simulate_2wins_architecture(outcomes)
    print(f"  Sesiones simuladas: {summary.total_sessions}")
    print(f"  Dias con meta >= $40: {summary.days_target_40}/{summary.total_days}")
    print(f"  Dias con meta >= $60: {summary.days_target_60}/{summary.total_days}")
    print(f"  PnL diario promedio: ${summary.avg_pnl_per_day:+.2f}")

    print("Generando Excel maestro...")
    build_excel(scan_rows, sessions, daily_pnl, summary, outcomes)
    print(f"\nOK -> Informe generado en: {OUTPUT}")


if __name__ == "__main__":
    main()
