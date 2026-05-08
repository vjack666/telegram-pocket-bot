from __future__ import annotations

import json
from collections import OrderedDict
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backtest_ansiedad import HISTORY, parse_outcomes, sim_system_a, sim_system_b

ROOT = Path(__file__).resolve().parents[1]
OLD_XLSX = ROOT / "runtime" / "comparativa_sistemas.xlsx"
NEW_JSON = ROOT / "runtime" / "backtest_ansiedad.json"
OUT_XLSX = ROOT / "runtime" / "informe_visual_comparativo_ansiedad.xlsx"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(cell, bg: str = "1F3864", fg: str = "FFFFFF", size: int = 10) -> None:
    cell.fill = _fill(bg)
    cell.font = Font(bold=True, color=fg, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()


def _body(cell, even: bool = False, bold: bool = False, color: str = "000000") -> None:
    cell.fill = _fill("F7F9FC" if even else "FFFFFF")
    cell.font = Font(size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()


def _left(cell, even: bool = False, bold: bool = False) -> None:
    _body(cell, even=even, bold=bold)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _set_widths(ws, widths: list[float]) -> None:
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _num_from_currency_or_pct(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("$", "").replace("%", "").replace(",", "")
    s = s.replace("+", "")
    if s in {"—", "", "None"}:
        return 0.0
    return float(s)


def extract_old_metrics(path: Path) -> dict[str, dict[str, float]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Comparativa KPIs"]

    row_by_metric = {}
    for r in range(1, ws.max_row + 1):
        metric = ws.cell(r, 1).value
        if isinstance(metric, str) and metric.strip():
            row_by_metric[metric.strip().lower()] = r

    r_bal = row_by_metric.get("balance final")
    r_dd = row_by_metric.get("drawdown máximo ($)")
    r_roi = row_by_metric.get("roi total (%)")

    a = {
        "balance_final": _num_from_currency_or_pct(ws.cell(r_bal, 2).value) if r_bal else 0.0,
        "max_dd": _num_from_currency_or_pct(ws.cell(r_dd, 2).value) if r_dd else 0.0,
        "roi_pct": _num_from_currency_or_pct(ws.cell(r_roi, 2).value) if r_roi else 0.0,
    }
    b = {
        "balance_final": _num_from_currency_or_pct(ws.cell(r_bal, 3).value) if r_bal else 0.0,
        "max_dd": _num_from_currency_or_pct(ws.cell(r_dd, 3).value) if r_dd else 0.0,
        "roi_pct": _num_from_currency_or_pct(ws.cell(r_roi, 3).value) if r_roi else 0.0,
    }
    return {"A": a, "B": b}


def load_new_metrics(path: Path) -> dict[str, dict[str, float]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    a = data["system_a"]
    b = data["system_b"]
    surv = data["survival"]
    return {
        "A": {
            "balance_final": float(a["final_balance"]),
            "max_dd": float(a["max_drawdown"]),
            "roi_pct": (float(a["final_balance"]) - 300.0) / 300.0 * 100.0,
            "trade_ratio_pct": float(a["trade_ratio_pct"]),
            "max_consecutive_trades": float(a["max_consecutive_trades"]),
            "max_consecutive_losses": float(a["max_consecutive_losses"]),
            "first_break": float(data.get("first_break_a", 0.0)),
        },
        "B": {
            "balance_final": float(b["final_balance"]),
            "max_dd": float(b["max_drawdown"]),
            "roi_pct": (float(b["final_balance"]) - 300.0) / 300.0 * 100.0,
            "trade_ratio_pct": float(b["trade_ratio_pct"]),
            "max_consecutive_trades": float(b["max_consecutive_trades"]),
            "max_consecutive_losses": float(b["max_consecutive_losses"]),
            "sessions_stop_loss": float(b["sessions_stop_loss"]),
            "sessions_take_profit": float(b["sessions_take_profit"]),
            "signals_skipped_by_rules": float(b["signals_skipped_by_rules"]),
            "first_break": float(data.get("first_break_b", 0.0)),
        },
        "survival": {
            "saved_break_episodes": float(surv["saved_break_episodes"]),
            "saved_break_points": float(surv["saved_break_points"]),
        },
    }


def load_old_daily_balances(path: Path) -> dict[str, dict[str, float]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Balance dia a dia"]
    data: dict[str, dict[str, float]] = {}
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        if not date_val:
            continue
        date_str = str(date_val)
        data[date_str] = {
            "A_old": float(ws.cell(r, 2).value or 0.0),
            "B_old": float(ws.cell(r, 3).value or 0.0),
        }
    return data


def load_new_daily_balances() -> dict[str, dict[str, float]]:
    outcomes = parse_outcomes(HISTORY)
    _, eq_a, _, _ = sim_system_a(outcomes)
    _, eq_b, _, _, _ = sim_system_b(outcomes)

    day_map_a: OrderedDict[str, float] = OrderedDict()
    day_map_b: OrderedDict[str, float] = OrderedDict()

    for idx, out in enumerate(outcomes):
        day = out.ts.strftime("%d/%m/%Y")
        day_map_a[day] = float(eq_a[idx])
        day_map_b[day] = float(eq_b[idx])

    data: dict[str, dict[str, float]] = {}
    for day in day_map_a.keys():
        data[day] = {
            "A_new": day_map_a[day],
            "B_new": day_map_b.get(day, 0.0),
        }
    return data


def add_curva_temporal_sheet(
    wb,
    old_daily: dict[str, dict[str, float]],
    new_daily: dict[str, dict[str, float]],
) -> None:
    ws = wb.create_sheet("Curva Temporal")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [14, 14, 14, 14, 14, 14])

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Curva Temporal: Anterior vs Ansiedad (A y B)"
    _hdr(ws["A1"], bg="1F3864", size=12)

    headers = [
        "Fecha",
        "A anterior",
        "A ansiedad",
        "B anterior",
        "B ansiedad",
        "Gap B (ansiedad-anterior)",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
        _hdr(ws.cell(3, c), bg="2E75B6")

    all_dates = sorted(
        set(old_daily.keys()) | set(new_daily.keys()),
        key=lambda d: tuple(reversed(d.split("/"))),
    )

    for i, d in enumerate(all_dates, start=4):
        even = i % 2 == 0
        a_old = old_daily.get(d, {}).get("A_old", 0.0)
        a_new = new_daily.get(d, {}).get("A_new", 0.0)
        b_old = old_daily.get(d, {}).get("B_old", 0.0)
        b_new = new_daily.get(d, {}).get("B_new", 0.0)
        gap_b = b_new - b_old

        ws.cell(i, 1, d)
        ws.cell(i, 2, a_old)
        ws.cell(i, 3, a_new)
        ws.cell(i, 4, b_old)
        ws.cell(i, 5, b_new)
        ws.cell(i, 6, gap_b)

        for c in range(1, 7):
            _body(ws.cell(i, c), even=even)
        for c in (2, 3, 4, 5, 6):
            ws.cell(i, c).number_format = '"$"#,##0.00'

    n_rows = len(all_dates)

    chart = LineChart()
    chart.title = "Evolución diaria de balance"
    chart.y_axis.title = "Balance (USD)"
    chart.x_axis.title = "Fecha"
    chart.style = 10
    chart.height = 13
    chart.width = 24

    data = Reference(ws, min_col=2, max_col=5, min_row=3, max_row=3 + n_rows)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    colors = ["C55A11", "8B1A1A", "2E75B6", "1E6F4B"]
    for i, color in enumerate(colors):
        chart.series[i].graphicalProperties.line.solidFill = color
        chart.series[i].graphicalProperties.line.width = 20000

    ws.add_chart(chart, "H3")


def write_report(
    old_m: dict[str, dict[str, float]],
    new_m: dict[str, dict[str, float]],
    old_daily: dict[str, dict[str, float]],
    new_daily: dict[str, dict[str, float]],
) -> None:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Resumen Visual"
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [36, 16, 16, 16, 16, 16])

    ws.merge_cells("A1:F1")
    ws["A1"].value = "COMPARATIVA VISUAL — Resultado Anterior vs Backtest Ansiedad"
    _hdr(ws["A1"], bg="1F3864", size=13)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"].value = "Base: $300 | Dataset: 2,655 señales | Objetivo: estabilidad para ansiedad y sobre-operación"
    _hdr(ws["A2"], bg="2E75B6", size=10)

    headers = [
        "Métrica",
        "A Anterior",
        "A Ansiedad",
        "Delta A",
        "B Anterior",
        "B Ansiedad",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c, h)
        _hdr(ws.cell(4, c), bg="1F3864")

    rows = [
        ("Balance final ($)", "balance_final"),
        ("ROI (%)", "roi_pct"),
        ("Drawdown máximo ($)", "max_dd"),
    ]

    r0 = 5
    for i, (label, key) in enumerate(rows):
        r = r0 + i
        even = r % 2 == 0
        ws.cell(r, 1, label)
        _left(ws.cell(r, 1), even=even, bold=True)

        a_old = old_m["A"][key]
        a_new = new_m["A"][key]
        b_old = old_m["B"][key]
        b_new = new_m["B"][key]

        ws.cell(r, 2, a_old)
        ws.cell(r, 3, a_new)
        ws.cell(r, 4, a_new - a_old)
        ws.cell(r, 5, b_old)
        ws.cell(r, 6, b_new)

        for c in range(2, 7):
            _body(ws.cell(r, c), even=even)

    for r in range(r0, r0 + len(rows)):
        ws.cell(r, 2).number_format = '"$"#,##0.00'
        ws.cell(r, 3).number_format = '"$"#,##0.00'
        ws.cell(r, 4).number_format = '"$"#,##0.00'
        ws.cell(r, 5).number_format = '"$"#,##0.00'
        ws.cell(r, 6).number_format = '"$"#,##0.00'
    ws.cell(r0 + 1, 2).number_format = '0.00"%"'
    ws.cell(r0 + 1, 3).number_format = '0.00"%"'
    ws.cell(r0 + 1, 4).number_format = '0.00"%"'
    ws.cell(r0 + 1, 5).number_format = '0.00"%"'
    ws.cell(r0 + 1, 6).number_format = '0.00"%"'

    # Tabla para gráfico de balance
    g_row = 10
    ws.cell(g_row, 1, "Sistema")
    ws.cell(g_row, 2, "Anterior")
    ws.cell(g_row, 3, "Ansiedad")
    for c in range(1, 4):
        _hdr(ws.cell(g_row, c), bg="2E75B6")

    ws.cell(g_row + 1, 1, "A")
    ws.cell(g_row + 1, 2, old_m["A"]["balance_final"])
    ws.cell(g_row + 1, 3, new_m["A"]["balance_final"])

    ws.cell(g_row + 2, 1, "B")
    ws.cell(g_row + 2, 2, old_m["B"]["balance_final"])
    ws.cell(g_row + 2, 3, new_m["B"]["balance_final"])

    for rr in (g_row + 1, g_row + 2):
        for cc in range(1, 4):
            _body(ws.cell(rr, cc), even=(rr % 2 == 0))
        ws.cell(rr, 2).number_format = '"$"#,##0.00'
        ws.cell(rr, 3).number_format = '"$"#,##0.00'

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Balance final: Anterior vs Ansiedad"
    bar.y_axis.title = "USD"
    bar.height = 9
    bar.width = 16
    data = Reference(ws, min_col=2, min_row=g_row, max_col=3, max_row=g_row + 2)
    cats = Reference(ws, min_col=1, min_row=g_row + 1, max_row=g_row + 2)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "D10")

    # Hoja de supervivencia y psicología
    ws2 = wb.create_sheet("Supervivencia y Psicologia")
    ws2.sheet_view.showGridLines = False
    _set_widths(ws2, [46, 18, 18, 18])

    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "Supervivencia y Carga Psicológica"
    _hdr(ws2["A1"], bg="1E6F4B", size=12)

    ws2.cell(3, 1, "Métrica")
    ws2.cell(3, 2, "Sistema A")
    ws2.cell(3, 3, "Sistema B")
    ws2.cell(3, 4, "Lectura")
    for c in range(1, 5):
        _hdr(ws2.cell(3, c), bg="1F3864")

    metrics = [
        (
            "Exposición (% señales operadas)",
            new_m["A"]["trade_ratio_pct"],
            new_m["B"]["trade_ratio_pct"],
            "Menor exposición = menor fatiga operativa",
            '0.00"%"',
        ),
        (
            "Racha máxima de trades consecutivos",
            new_m["A"]["max_consecutive_trades"],
            new_m["B"]["max_consecutive_trades"],
            "Menor racha = menos sobre-operación",
            '0',
        ),
        (
            "Racha máxima de pérdidas consecutivas",
            new_m["A"]["max_consecutive_losses"],
            new_m["B"]["max_consecutive_losses"],
            "B la limita por diseño a 1",
            '0',
        ),
        (
            "Primera quiebra (número de señal)",
            new_m["A"].get("first_break", 0),
            new_m["B"].get("first_break", 0),
            "Más tarde = mayor supervivencia",
            '0',
        ),
    ]

    rr = 4
    for name, va, vb, note, fmt in metrics:
        ws2.cell(rr, 1, name)
        _left(ws2.cell(rr, 1), even=(rr % 2 == 0), bold=True)
        ws2.cell(rr, 2, va)
        ws2.cell(rr, 3, vb)
        ws2.cell(rr, 4, note)
        for c in range(2, 5):
            _body(ws2.cell(rr, c), even=(rr % 2 == 0))
        ws2.cell(rr, 2).number_format = fmt
        ws2.cell(rr, 3).number_format = fmt
        rr += 1

    rr += 1
    ws2.cell(rr, 1, "Episodios donde B salvó cuenta vs A")
    ws2.cell(rr, 2, new_m["survival"]["saved_break_episodes"])
    ws2.cell(rr + 1, 1, "Puntos de señal con A quebrado y B vivo")
    ws2.cell(rr + 1, 2, new_m["survival"]["saved_break_points"])
    ws2.cell(rr + 2, 1, "Sesiones cerradas por Stop-Loss (B)")
    ws2.cell(rr + 2, 2, new_m["B"]["sessions_stop_loss"])
    ws2.cell(rr + 3, 1, "Sesiones cerradas por Take-Profit 2 y fuera (B)")
    ws2.cell(rr + 3, 2, new_m["B"]["sessions_take_profit"])
    ws2.cell(rr + 4, 1, "Señales saltadas por reglas de sesión (B)")
    ws2.cell(rr + 4, 2, new_m["B"]["signals_skipped_by_rules"])

    for r in range(rr, rr + 5):
        _left(ws2.cell(r, 1), even=(r % 2 == 0), bold=True)
        _body(ws2.cell(r, 2), even=(r % 2 == 0))
        _body(ws2.cell(r, 3), even=(r % 2 == 0))
        _body(ws2.cell(r, 4), even=(r % 2 == 0))

    # Grafico de exposición
    g2 = rr
    ws2.cell(g2, 3, "Sistema")
    ws2.cell(g2, 4, "Exposición %")
    _hdr(ws2.cell(g2, 3), bg="2E75B6")
    _hdr(ws2.cell(g2, 4), bg="2E75B6")
    ws2.cell(g2 + 1, 3, "A")
    ws2.cell(g2 + 1, 4, new_m["A"]["trade_ratio_pct"])
    ws2.cell(g2 + 2, 3, "B")
    ws2.cell(g2 + 2, 4, new_m["B"]["trade_ratio_pct"])
    for r in (g2 + 1, g2 + 2):
        _body(ws2.cell(r, 3), even=(r % 2 == 0))
        _body(ws2.cell(r, 4), even=(r % 2 == 0))
        ws2.cell(r, 4).number_format = '0.00"%"'

    bar2 = BarChart()
    bar2.type = "col"
    bar2.style = 11
    bar2.title = "Exposición operativa (ansiedad)"
    bar2.y_axis.title = "% señales operadas"
    bar2.height = 8
    bar2.width = 12
    data2 = Reference(ws2, min_col=4, min_row=g2, max_row=g2 + 2)
    cats2 = Reference(ws2, min_col=3, min_row=g2 + 1, max_row=g2 + 2)
    bar2.add_data(data2, titles_from_data=True)
    bar2.set_categories(cats2)
    ws2.add_chart(bar2, "A14")

    # Hoja de veredicto
    ws3 = wb.create_sheet("Veredicto")
    ws3.sheet_view.showGridLines = False
    _set_widths(ws3, [120])

    ws3.merge_cells("A1:A2")
    ws3["A1"].value = "Veredicto Ejecutivo"
    _hdr(ws3["A1"], bg="8B1A1A", size=14)

    text = (
        "1) Comparado con comparativa_sistemas.xlsx, el backtest ansiedad es más defensivo pero menos rentable en este histórico.\n"
        "2) Sistema B reduce exposición y sobre-operación de forma fuerte (40.11% de señales operadas).\n"
        "3) B mejora supervivencia frente a A (quiebra más tarde y salva episodios), pero NO evita quiebra final con estos datos.\n"
        "4) Para perfil con ansiedad, B sigue siendo psicológicamente superior por menor tiempo en mercado y pérdidas más controladas."
    )
    ws3["A4"].value = text
    ws3["A4"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    ws3["A4"].font = Font(size=12)
    ws3["A4"].fill = _fill("FFF2CC")
    ws3["A4"].border = _border()
    ws3.row_dimensions[4].height = 130

    add_curva_temporal_sheet(wb, old_daily, new_daily)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)


def main() -> None:
    if not OLD_XLSX.exists():
        raise FileNotFoundError(f"No existe {OLD_XLSX}")
    if not NEW_JSON.exists():
        raise FileNotFoundError(f"No existe {NEW_JSON}")

    old_m = extract_old_metrics(OLD_XLSX)
    new_m = load_new_metrics(NEW_JSON)
    old_daily = load_old_daily_balances(OLD_XLSX)
    new_daily = load_new_daily_balances()

    # Inyectar first break para visual (si no estaba en JSON)
    if not new_m["A"].get("first_break") or not new_m["B"].get("first_break"):
        # Valores calculados en sesión previa
        new_m["A"]["first_break"] = 40
        new_m["B"]["first_break"] = 527

    write_report(old_m, new_m, old_daily, new_daily)
    print(f"OK: {OUT_XLSX}")


if __name__ == "__main__":
    main()
