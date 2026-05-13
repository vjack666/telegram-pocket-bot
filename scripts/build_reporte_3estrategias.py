"""
Reporte Comparativo Unificado — 3 Estrategias
==============================================
A — Sistema Actual (Masaniello con freno)    → datos reales de Reporte_Comparativo anterior
B — Macro-Recuperación                       → datos reales de Reporte_Comparativo anterior
C — Sesiones $10                             → datos de Detalle_SimulacionC_Sesiones10.csv

Genera: runtime/Reporte_Comparativo_3Estrategias_100USD.xlsx
  Hoja 1: Resumen_Ejecutivo       (tabla comparativa A vs B vs C)
  Hoja 2: Curvas_de_Capital       (gráfica de línea + tabla)
  Hoja 3: Detalle_A_B             (señal a señal, columnas A y B del excel anterior)
  Hoja 4: Detalle_C               (sesión a sesión, estrategia C)
  Hoja 5: Punto_de_Quiebre_C      (tabla de escalones de deuda)
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from pathlib import Path

# ─── Rutas ───────────────────────────────────────────────────────────────────
PATH_AB  = Path("runtime/_temp_ab_operacion.csv")
PATH_C   = Path("runtime/Detalle_SimulacionC_Sesiones10.csv")
OUT      = Path("runtime/Reporte_Comparativo_Recovery_Cycle.xlsx")
OUT_TEMP = Path("runtime/_Reporte_Comparativo_Recovery_Cycle_temp.xlsx")

PAYOUT        = 0.92
OBJETIVO_WIN  = 5.0
CAPITAL_INCIAL = 100.0
TP_WINS       = 2
SL_LOSSES     = 3

# ─── Helpers de estilo ───────────────────────────────────────────────────────
def fill(hex_str):
    return PatternFill("solid", fgColor=hex_str)

def bfont(size=10, color="000000", italic=False):
    return Font(bold=True, size=size, color=color, italic=italic)

def nfont(size=10, color="000000"):
    return Font(size=size, color=color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left():
    return Alignment(horizontal="left", vertical="center")

def border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# Paleta
C_NIGHT   = "1A1A2E"
C_A       = "0D7377"   # verde azulado
C_B       = "C0392B"   # rojo
C_C       = "D4890A"   # naranja dorado
C_DARK    = "2C3E50"
C_BG1     = "F7F9FC"
C_BG2     = "EDF2F7"
C_GREEN   = "D5F5E3"
C_RED     = "FADBD8"
C_YELLOW  = "FEF9E7"

def header_cell(ws, row, col, value, bg=C_NIGHT, fg="FFFFFF", size=11, span=1):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill(bg)
    c.font      = bfont(size=size, color=fg)
    c.alignment = center()
    c.border    = border()
    return c

def data_cell(ws, row, col, value, bg=C_BG1, bold=False, h_align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill(bg)
    c.font      = bfont(size=10) if bold else nfont(size=10)
    c.alignment = Alignment(horizontal=h_align, vertical="center")
    c.border    = border()
    return c

# ─── Cargar datos ─────────────────────────────────────────────────────────────
df_ab = pd.read_csv(PATH_AB)
df_c  = pd.read_csv(PATH_C)

# Curvas de capital (con punto inicial $100)
curva_a = [100.0] + df_ab["A_Balance"].tolist()
curva_b = [100.0] + df_ab["B_Balance"].tolist()
# Curva C: Capital_Despues sesión a sesión (ya incluye valor inicial)
curva_c_raw = df_c["Capital_Despues"].tolist()
curva_c = [100.0] + curva_c_raw  # agregar punto inicial

# ─── Métricas A ──────────────────────────────────────────────────────────────
cap_final_a  = df_ab["A_Balance"].iloc[-1]
cap_min_a    = df_ab["A_Balance"].min()
cap_max_a    = df_ab["A_Balance"].max()
pnl_a        = cap_final_a - 100.0
ret_a        = pnl_a / 100.0 * 100
dd_max_a     = df_ab["A_DrawdownPct"].max()
wins_a       = (df_ab["Resultado"] != "Loss").sum()
total_ops    = len(df_ab)
wr_a         = wins_a / total_ops * 100
# Rachas
resultados_a = (df_ab["Resultado"] != "Loss").tolist()
max_racha_w_a = max_racha_l_a = 0
cur_w = cur_l = 0
for r in resultados_a:
    if r:
        cur_w += 1; cur_l = 0
        max_racha_w_a = max(max_racha_w_a, cur_w)
    else:
        cur_l += 1; cur_w = 0
        max_racha_l_a = max(max_racha_l_a, cur_l)

# ─── Métricas B ──────────────────────────────────────────────────────────────
cap_final_b  = df_ab["B_Balance"].iloc[-1]
cap_min_b    = df_ab["B_Balance"].min()
cap_max_b    = df_ab["B_Balance"].max()
pnl_b        = cap_final_b - 100.0
ret_b        = pnl_b / 100.0 * 100
dd_max_b     = df_ab["B_DrawdownPct"].max()
wr_b         = wr_a  # mismos datos subyacentes

# ─── Métricas C ──────────────────────────────────────────────────────────────
total_s      = len(df_c)
tp_count     = (df_c["Estado"] == "TP_Alcanzado").sum()
sl_count     = (df_c["Estado"] == "SL_Activado").sum()
agotada      = (df_c["Estado"] == "Sesion_Agotada").sum()
cap_final_c  = df_c["Capital_Despues"].iloc[-1]
cap_min_c    = min(curva_c)
cap_max_c    = max(curva_c)
pnl_c        = cap_final_c - 100.0
ret_c        = pnl_c / 100.0 * 100
dd_max_c_val = df_c["Max_Exposure_USD"].max()
wr_c         = tp_count / total_s * 100
quiebras_c   = (df_c["Capital_Negativo"] == True).sum()
ganancia_tp  = tp_count * 10.0
perdida_sl   = df_c[df_c["Estado"] == "SL_Activado"]["Resultado_Sesion_USD"].sum()

print(f"A → Capital Final: ${cap_final_a:.2f} | PnL: ${pnl_a:+.2f} | DD: {dd_max_a:.1f}%")
print(f"B → Capital Final: ${cap_final_b:.2f} | PnL: ${pnl_b:+.2f} | DD: {dd_max_b:.1f}%")
print(f"C → Capital Final: ${cap_final_c:.2f} | PnL: ${pnl_c:+.2f} | TP:{tp_count} SL:{sl_count}")

# ─── WORKBOOK ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1: Resumen Ejecutivo
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Resumen_Ejecutivo"

# Título
ws1.merge_cells("A1:D1")
c = ws1["A1"]
c.value     = "REPORTE COMPARATIVO — 3 ESTRATEGIAS  |  Capital $100  |  Payout 92%"
c.fill      = fill(C_NIGHT)
c.font      = bfont(size=14, color="FFFFFF")
c.alignment = center()
ws1.row_dimensions[1].height = 32

# Subtítulo
ws1.merge_cells("A2:D2")
c = ws1["A2"]
c.value     = f"Dataset: {total_ops} señales  |  {total_s} sesiones  |  Periodo: Mar–May 2026  |  Fuente: ejemplo.md"
c.fill      = fill(C_DARK)
c.font      = Font(italic=True, size=10, color="BDC3C7")
c.alignment = center()
ws1.row_dimensions[2].height = 18

# Cabeceras de estrategias
row_h = 4
ws1.row_dimensions[row_h].height = 24
header_cell(ws1, row_h, 1, "MÉTRICA",                  bg=C_NIGHT, size=11)
header_cell(ws1, row_h, 2, "A — Sistema Actual",       bg=C_A,     size=11)
header_cell(ws1, row_h, 3, "B — Macro-Recuperación",   bg=C_B,     size=11)
header_cell(ws1, row_h, 4, "C — Sesiones $10",         bg=C_C,     size=11)

# Bloque 1: Capital
def section_title(ws, row, text):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill      = fill(C_DARK)
    c.font      = bfont(size=10, color="FFFFFF")
    c.alignment = left()
    c.border    = border()
    ws.row_dimensions[row].height = 16

def metric_row(ws, row, metrica, val_a, val_b, val_c, alt=False):
    bg = C_BG2 if alt else C_BG1
    data_cell(ws, row, 1, metrica, bg=bg, bold=True, h_align="left")
    data_cell(ws, row, 2, val_a,   bg=bg)
    data_cell(ws, row, 3, val_b,   bg=bg)
    data_cell(ws, row, 4, val_c,   bg=bg)
    ws.row_dimensions[row].height = 17

r = row_h + 1
section_title(ws1, r, "  📊 CAPITAL"); r += 1
metric_row(ws1, r, "Capital Inicial",         "$100.00",              "$100.00",              "$100.00",              alt=False); r+=1
metric_row(ws1, r, "Capital Final",           f"${cap_final_a:.2f}",  f"${cap_final_b:.2f}",  f"${cap_final_c:.2f}",  alt=True); r+=1
metric_row(ws1, r, "Capital Máximo Alcanzado",f"${cap_max_a:.2f}",    f"${cap_max_b:.2f}",    f"${cap_max_c:.2f}",    alt=False); r+=1
metric_row(ws1, r, "Capital Mínimo Tocado",   f"${cap_min_a:.2f}",    f"${cap_min_b:.2f}",    f"${cap_min_c:.2f}",    alt=True); r+=1
metric_row(ws1, r, "PnL Neto Total",          f"${pnl_a:+.2f}",       f"${pnl_b:+.2f}",       f"${pnl_c:+.2f}",       alt=False); r+=1
metric_row(ws1, r, "Rentabilidad sobre $100", f"{ret_a:+.1f}%",       f"{ret_b:+.1f}%",       f"{ret_c:+.1f}%",       alt=True); r+=1

section_title(ws1, r, "  🎯 RENDIMIENTO"); r += 1
metric_row(ws1, r, "Unidades analizadas",     f"{total_ops} señales", f"{total_ops} señales", f"{total_s} sesiones",  alt=False); r+=1
metric_row(ws1, r, "Win Rate",                f"{wr_a:.1f}%",         f"{wr_b:.1f}%",         f"{wr_c:.1f}% (TP/ses)",alt=True); r+=1
metric_row(ws1, r, "TP / Wins",               f"{wins_a}",            "N/A",                  f"{tp_count} sesiones", alt=False); r+=1
metric_row(ws1, r, "SL / Losses",             f"{total_ops - wins_a}",f"{(df_ab['B_MacroNivel']>0).sum()} resets",   f"{sl_count} sesiones", alt=True); r+=1
metric_row(ws1, r, "Sesiones agotadas",       "—",                    "—",                    f"{agotada}",           alt=False); r+=1

section_title(ws1, r, "  ⚠ RIESGO"); r += 1
metric_row(ws1, r, "Max Drawdown %",          f"{dd_max_a:.1f}%",     f"{dd_max_b:.1f}%",     "N/A (por sesión)",     alt=False); r+=1
metric_row(ws1, r, "Max exposición por sesión","—",                   "—",                    f"${dd_max_c_val:.2f}", alt=True); r+=1
metric_row(ws1, r, "Veces capital negativo",  "Ver curva",            "Ver curva",
           f"{'⚠ ' + str(quiebras_c) + ' veces' if quiebras_c > 0 else '✓ 0 veces'}",        alt=False); r+=1
metric_row(ws1, r, "Racha máx. pérdidas (A)", f"{max_racha_l_a}",     "—",                    "3 = activa SL",        alt=True); r+=1

section_title(ws1, r, "  💡 OBSERVACIONES"); r += 1
obs = [
    ("Comportamiento capital", "Decae sostenidamente",   "Decae más lento que A", "Crece exponencial (+4,300%)"),
    ("Riesgo por operación",   "Stake variable c/freno", "Escala con deuda macro", "Escala dentro del bloque"),
    ("Punto de quiebre",       "Rachas largas de L",     "5+ losses sin reset",   "3 losses en mismo bloque"),
    ("Capital para operar",    ">$7 mín. stake",         ">$7 mín. stake",        ">$40.45 en peor caso"),
    ("Veredicto",              "❌ Capital en riesgo",   "⚠ Mejora parcial",      "✅ Rentable histórico"),
]
for i, (met, va, vb, vc) in enumerate(obs):
    metric_row(ws1, r, met, va, vb, vc, alt=(i % 2 == 0)); r += 1

# Nota al pie
ws1.merge_cells(f"A{r}:D{r+1}")
c = ws1.cell(row=r, column=1,
    value="★ NOTA: La Estrategia C crece de forma espectacular porque los datos históricos muestran solo 2 SL en "
          "453 sesiones (0.4%). Sin embargo, esto es un backtest y no garantiza resultados futuros. "
          "Una racha de 3 losses consecutivos consume ~$40.45 por sesión. Capital mínimo recomendado: $100.")
c.fill      = fill("FFFDE7")
c.font      = Font(italic=True, size=9, color="555555")
c.alignment = Alignment(wrap_text=True, vertical="top")
ws1.row_dimensions[r].height = 28
ws1.row_dimensions[r+1].height = 28

ws1.column_dimensions["A"].width = 34
ws1.column_dimensions["B"].width = 26
ws1.column_dimensions["C"].width = 26
ws1.column_dimensions["D"].width = 24

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2: Curvas de Capital
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Curvas_de_Capital")

# Títulos de columna
for col_i, (txt, bg) in enumerate(
    [("# Operación/Sesión", C_NIGHT), ("A — Sistema Actual", C_A),
     ("B — Macro-Recuperación", C_B), ("C — Sesiones $10", C_C)], start=1
):
    c = ws2.cell(row=1, column=col_i, value=txt)
    c.fill = fill(bg); c.font = bfont(color="FFFFFF"); c.alignment = center()
    c.border = border()
ws2.row_dimensions[1].height = 20

# Las curvas A y B tienen 2708 puntos, C tiene 454.
# Para el gráfico: submuestrear A y B cada ~6 puntos para alinear escala visual
# pero para la tabla de datos: usar señales A/B con marcadores de sesión C
n_ab = len(curva_a)
n_c  = len(curva_c)

# Submuestrear: tomar un punto cada N para que queden ~500 filas en el gráfico
step_ab = max(1, n_ab // 500)
indices_ab = list(range(0, n_ab, step_ab))

# Interpolar curva C a la misma escala de señales
# Cada sesión C equivale a ~6 señales (2707/453 ≈ 5.98)
import numpy as np
ratio = n_ab / n_c
indices_c_interp = np.interp(
    indices_ab,
    [i * ratio for i in range(n_c)],
    curva_c
)

for r_i, (idx_ab, val_c_i) in enumerate(zip(indices_ab, indices_c_interp), start=2):
    ws2.cell(row=r_i, column=1, value=idx_ab)
    ws2.cell(row=r_i, column=2, value=round(curva_a[idx_ab], 2) if idx_ab < len(curva_a) else curva_a[-1])
    ws2.cell(row=r_i, column=3, value=round(curva_b[idx_ab], 2) if idx_ab < len(curva_b) else curva_b[-1])
    ws2.cell(row=r_i, column=4, value=round(float(val_c_i), 2))

n_data_rows = len(indices_ab) + 1

# Gráfico
chart = LineChart()
chart.title         = "Curva de Capital — A vs B vs C  |  Capital inicial $100"
chart.style         = 10
chart.y_axis.title  = "Capital ($)"
chart.x_axis.title  = "Señal #"
chart.width         = 30
chart.height        = 18

series_cfg = [
    (2, C_A.replace("0D", "00"), "A — Sistema Actual"),
    (3, "C0392B",                "B — Macro-Recuperación"),
    (4, "D4890A",                "C — Sesiones $10"),
]
for col_idx, (col_n, color_hex_s, title_s) in enumerate(series_cfg):
    data_ref = Reference(ws2, min_col=col_n, min_row=1, max_row=n_data_rows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.series[-1].graphicalProperties.line.solidFill = color_hex_s
    chart.series[-1].graphicalProperties.line.width     = 20000

ws2.add_chart(chart, "F2")

for col_l, w in zip(["A","B","C","D"], [22, 22, 24, 22]):
    ws2.column_dimensions[col_l].width = w

# ════════════════════════════════════════════════════════════════════════════
# HOJA 3: Detalle_A_B  (señal a señal)
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Detalle_A_B")

cols_ab = ["Fecha","Hora","Activo","Direccion","Resultado",
           "A_Stake","A_PnL","A_Balance","A_DrawdownPct",
           "B_Stake","B_PnL","B_Balance","B_DrawdownPct","B_CajaNivel"]

for col_i, h in enumerate(cols_ab, start=1):
    c = ws3.cell(row=1, column=col_i, value=h)
    c.fill = fill(C_NIGHT); c.font = bfont(color="FFFFFF", size=9)
    c.alignment = center(); c.border = border()
ws3.row_dimensions[1].height = 18
ws3.freeze_panes = "A2"

for r_i, row_data in enumerate(df_ab[cols_ab].itertuples(index=False), start=2):
    resultado = row_data.Resultado
    if resultado == "Loss":
        bg = C_RED
    elif resultado in ("Win Gale", "Win Gale 2"):
        bg = C_YELLOW
    else:
        bg = C_BG1 if r_i % 2 == 0 else C_BG2

    for col_i, val in enumerate(row_data, start=1):
        c = ws3.cell(row=r_i, column=col_i, value=val)
        c.fill = fill(bg); c.font = nfont(size=8)
        c.alignment = center(); c.border = border()

widths_ab = [12, 8, 16, 10, 12, 10, 10, 12, 12, 10, 10, 12, 12, 12]
for col_i, w in enumerate(widths_ab, start=1):
    ws3.column_dimensions[get_column_letter(col_i)].width = w

# ════════════════════════════════════════════════════════════════════════════
# HOJA 4: Detalle_C  (sesión a sesión)
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Detalle_C")

cols_c = ["ID_Sesion","Señales_Jugadas","Wins","Losses","Estado",
          "Resultado_Sesion_USD","Max_Exposure_USD","Capital_Despues",
          "Balance_Objetivo","Capital_Negativo","Stakes_Log"]

if "Balance_Objetivo" not in df_c.columns:
    df_c["Balance_Objetivo"] = None
fills_estado = {
    "TP_Alcanzado" : C_GREEN,
    "SL_Activado"  : C_RED,
    "Sesion_Agotada": C_YELLOW,
}

for col_i, h in enumerate(cols_c, start=1):
    c = ws4.cell(row=1, column=col_i, value=h)
    c.fill = fill(C_NIGHT); c.font = bfont(color="FFFFFF", size=9)
    c.alignment = center(); c.border = border()
ws4.row_dimensions[1].height = 18
ws4.freeze_panes = "A2"

for r_i, row_data in enumerate(df_c[cols_c].itertuples(index=False), start=2):
    bg = fills_estado.get(row_data.Estado, C_BG1)
    vals = [row_data.ID_Sesion, row_data.Señales_Jugadas, row_data.Wins,
            row_data.Losses, row_data.Estado, row_data.Resultado_Sesion_USD,
            row_data.Max_Exposure_USD, row_data.Capital_Despues,
            row_data.Balance_Objetivo,
            "⚠ SÍ" if row_data.Capital_Negativo else "No",
            row_data.Stakes_Log]
    for col_i, val in enumerate(vals, start=1):
        c = ws4.cell(row=r_i, column=col_i, value=val)
        c.fill = fill(bg); c.font = nfont(size=9)
        c.alignment = Alignment(horizontal="left" if col_i == 11 else "center", vertical="center")
        c.border = border()

widths_c = [10, 15, 8, 8, 16, 22, 20, 18, 18, 16, 75]
for col_i, w in enumerate(widths_c, start=1):
    ws4.column_dimensions[get_column_letter(col_i)].width = w

# ════════════════════════════════════════════════════════════════════════════
# HOJA 5: Punto de Quiebre C
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Punto_de_Quiebre_C")

ws5.merge_cells("A1:E1")
c = ws5["A1"]
c.value = "ANÁLISIS DE PUNTO DE QUIEBRE — Estrategia C (Sesiones $10)"
c.fill  = fill(C_NIGHT); c.font = bfont(size=13, color="FFFFFF"); c.alignment = center()
ws5.row_dimensions[1].height = 28

# Tabla de escalones
ws5.cell(row=3, column=1, value="Escalones de deuda si se pierden señales consecutivas").font = bfont(size=11)

esc_hdrs = ["Pérdida #", "Stake ($)", "Deuda Acumulada ($)", "Capital necesario", "¿Aguanta con $100?"]
for col_i, h in enumerate(esc_hdrs, start=1):
    header_cell(ws5, 4, col_i, h, bg=C_DARK)

perdida_acum_esc = 0.0
for step_n in range(1, SL_LOSSES + 1):
    stake_esc        = (perdida_acum_esc + OBJETIVO_WIN) / PAYOUT
    perdida_acum_esc += stake_esc
    aguanta  = "✓ Sí" if perdida_acum_esc <= 100 else "✗ NO ⚠"
    bg_esc   = C_GREEN if perdida_acum_esc <= 100 else C_RED
    for col_i, v in enumerate(
        [step_n, round(stake_esc, 4), round(perdida_acum_esc, 4), f"${perdida_acum_esc:.2f}", aguanta], start=1
    ):
        data_cell(ws5, 4+step_n, col_i, v, bg=bg_esc)

# Distribución de resultados
ws5.cell(row=9, column=1, value=f"Distribución de resultados por sesión ({total_s} sesiones)").font = bfont(size=11)
for col_i, h in enumerate(["Estado", "Frecuencia", "% del Total", "PnL Acumulado ($)", "Capital Promedio Post-Sesión"], start=1):
    header_cell(ws5, 10, col_i, h, bg=C_DARK)

dist = df_c.groupby("Estado").agg(
    Freq=("ID_Sesion","count"),
    PnL=("Resultado_Sesion_USD","sum"),
    CapProm=("Capital_Despues","mean")
).reset_index()

for r_i, row_d in enumerate(dist.itertuples(index=False), start=11):
    pct = row_d.Freq / total_s * 100
    bg  = fills_estado.get(row_d.Estado, C_BG1)
    for col_i, v in enumerate(
        [row_d.Estado, row_d.Freq, f"{pct:.1f}%", f"${row_d.PnL:+.2f}", f"${row_d.CapProm:.2f}"], start=1
    ):
        data_cell(ws5, r_i, col_i, v, bg=bg)

# Comparativa PnL sesiones TP vs SL
ws5.cell(row=15, column=1, value="Impacto financiero").font = bfont(size=11)
for col_i, v in enumerate(["Tipo", "Cantidad", "PnL Total ($)", "PnL Promedio ($)"], start=1):
    header_cell(ws5, 16, col_i, v, bg=C_DARK)

rows_fin = [
    ("TP Alcanzados", tp_count, f"${ganancia_tp:.2f}", f"${10.0:.2f}"),
    ("SL Activados",  sl_count, f"${perdida_sl:.2f}",  f"${perdida_sl/max(sl_count,1):.2f}"),
    ("Sesiones Agotadas", agotada, "~$0.00", "$0.00"),
    ("TOTAL", total_s, f"${pnl_c:+.2f}", f"${pnl_c/total_s:+.2f}"),
]
fills_fin = [C_GREEN, C_RED, C_YELLOW, C_DARK]
fonts_fin  = ["000000","000000","000000","FFFFFF"]
for r_i, ((estado_f, cant, pnl_tot, pnl_avg), bg_f, fg_f) in enumerate(
    zip(rows_fin, fills_fin, fonts_fin), start=17
):
    for col_i, v in enumerate([estado_f, cant, pnl_tot, pnl_avg], start=1):
        c = data_cell(ws5, r_i, col_i, v, bg=bg_f)
        if r_i == 20:  # fila TOTAL
            c.font = bfont(color=fg_f)

for col_l, w in zip(["A","B","C","D","E"], [24, 14, 22, 22, 28]):
    ws5.column_dimensions[col_l].width = w

# ─── Guardar ─────────────────────────────────────────────────────────────────
wb.save(OUT_TEMP)
import shutil
import time
import os
try:
    if OUT.exists():
        for _ in range(3):
            try:
                OUT.unlink()
                break
            except PermissionError:
                time.sleep(0.2)
except:
    pass
time.sleep(0.1)
shutil.move(str(OUT_TEMP), str(OUT))
print(f"\n✓ Excel guardado en: {OUT}")
print(f"  Hojas: {[s.title for s in wb.worksheets]}")
