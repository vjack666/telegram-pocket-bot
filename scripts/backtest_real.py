"""
BACKTEST REAL — Auditoría del bot antes y después de los fixes de riesgo
========================================================================
Fuente: ejemplo.md — 2,655 señales reales (17/03 – 07/05/2026)

Compara 3 sistemas corriendo sobre los MISMOS datos:

  Sistema A — "Bot actual ANTES del fix"
      Masaniello 12/4, cap POR PASO (entry≤$30, G1≤$30, G2≤$30)
      → exposición real puede ser $90+ por señal perdida

  Sistema B — "Bot actual DESPUÉS del fix (12/4)"
      Masaniello 12/4, cap TOTAL (entry+G1+G2 ≤ $30)
      → fiel a los cambios aplicados en engine.py

  Sistema C — "Arquitectura óptima (6/2 + max_losses=3)"
      Masaniello 6/2, cap TOTAL, máx 3 pérdidas por sesión
      → arquitectura 2-wins propuesta

Hojas Excel:
  1. Comparativa        — KPIs de los 3 sistemas lado a lado + gráfica
  2. Detalle Sistema A  — día a día + sesiones
  3. Detalle Sistema B  — día a día + sesiones
  4. Detalle Sistema C  — día a día + sesiones
  5. Equity Curves      — curva de capital acumulada diaria (los 3 sistemas)
  6. Auditoria Riesgo   — exposición máxima, rachas, drawdown, distribución
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT    = Path(__file__).resolve().parents[1]
HISTORY = ROOT / "ejemplo.md"
OUTPUT  = ROOT / "runtime" / "backtest_real.xlsx"

# ── Constantes globales ───────────────────────────────────────────────────
PAYOUT_MULT = 1.92
NET_PAYOUT  = 0.92
G1_MULT     = round(PAYOUT_MULT / NET_PAYOUT, 6)       # ≈ 2.08696
G2_MULT     = round(G1_MULT ** 2, 6)                   # ≈ 4.35538
TOTAL_MULT  = round(1.0 + G1_MULT + G2_MULT, 6)        # ≈ 7.44234
BASE        = 300.0
CAP_PCT     = 0.10

# ── Colores ───────────────────────────────────────────────────────────────
AH  = "1F3864"; AM  = "2E75B6"; ACL = "D6E4F7"
GH  = "1E6F4B"; GB  = "D6F0E0"
RH  = "8B1A1A"; RB  = "FDDEDE"
YB  = "FFF2CC"; OB  = "FFE5B4"
GR  = "F2F2F2"; WH  = "FFFFFF"
OG  = "E26B0A"   # naranja oscuro


# ═══════════════════════════════════════════════════════════════════════════
# 1. PARSER
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class Outcome:
    timestamp: datetime
    result: str    # WD / G1 / G2 / L
    is_win: bool


def parse_outcomes(path: Path) -> list[Outcome]:
    date_pat = re.compile(r"^\[(\d{2}/\d{2}/\d{4}) (\d{2}:\d{2}:\d{2})\]")
    res_pat  = re.compile(
        r"(VICTORIA DIRECTA|VICTORIA EN 1.*?MARTINGALA|VICTORIA EN 2.*?MARTINGALA|P[EÉ]RDIDA)",
        re.IGNORECASE,
    )
    label_map = {
        "victoria directa": "WD", "victoria en 1": "G1",
        "victoria en 2": "G2", "perdida": "L", "pérdida": "L",
    }
    outcomes: list[Outcome] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        d = date_pat.match(line)
        if not d:
            continue
        r = res_pat.search(line)
        if not r:
            continue
        ts  = datetime.strptime(" ".join(d.groups()), "%d/%m/%Y %H:%M:%S")
        raw = r.group(1).lower()
        lbl = next((v for k, v in label_map.items() if raw.startswith(k)), "?")
        if lbl == "?":
            continue
        outcomes.append(Outcome(ts, lbl, lbl != "L"))
    return outcomes


# ═══════════════════════════════════════════════════════════════════════════
# 2. FÓRMULA MASANIELLO
# ═══════════════════════════════════════════════════════════════════════════

def _fwd(ops: int, wins: int, pm: float) -> float:
    if wins <= 0:       return 1.0
    if wins > ops:      return 0.0
    if wins == ops:     return pm ** ops
    pw = _fwd(ops - 1, wins - 1, pm)
    pl = _fwd(ops - 1, wins, pm)
    d  = pw + (pm - 1) * pl
    return (pm * pw * pl / d) if d else 0.0


def masaniello_raw(losses: int, wins: int, n: int, w: int, pm: float) -> float:
    """Stake normalizado (sobre base=1)."""
    ops_left  = n - (losses + wins)
    wins_left = w - wins
    if ops_left <= 0 or wins_left <= 0 or wins_left > ops_left:
        return 0.0
    pw = _fwd(ops_left - 1, wins_left - 1, pm)
    pl = _fwd(ops_left - 1, wins_left, pm)
    d  = pw + (pm - 1) * pl
    if not d:
        return 1.0
    return max(0.001, min(1.0 - pm * pw / d, 1.0))


# Tablas precalculadas por (n, w)
_TABLES: dict[tuple[int, int], dict[tuple[int, int], float]] = {}

def get_table(n: int, w: int) -> dict[tuple[int, int], float]:
    key = (n, w)
    if key not in _TABLES:
        _TABLES[key] = {
            (l, wins): masaniello_raw(l, wins, n, w, PAYOUT_MULT)
            for l in range(n)
            for wins in range(n)
            if l + wins < n
        }
    return _TABLES[key]


# ═══════════════════════════════════════════════════════════════════════════
# 3. TRES CONFIGURACIONES DE SISTEMA
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class SystemConfig:
    name:       str
    color:      str    # hex para Excel
    n_ops:      int
    w_needed:   int
    cap_pct:    float
    max_losses: int    # 0 = sin límite
    cap_mode:   str    # "per_step" | "total"
    base:       float  = BASE


SYSTEMS: list[SystemConfig] = [
    SystemConfig(
        name="A — Bot anterior (12/4, cap x paso)",
        color="C55A11",
        n_ops=12, w_needed=4, cap_pct=0.10, max_losses=0,
        cap_mode="per_step",
    ),
    SystemConfig(
        name="B — Bot actual (12/4, cap total)",
        color="2E75B6",
        n_ops=12, w_needed=4, cap_pct=0.10, max_losses=0,
        cap_mode="total",
    ),
    SystemConfig(
        name="C — Arquitectura óptima (6/2, cap total, guard=3)",
        color="1E6F4B",
        n_ops=6, w_needed=2, cap_pct=0.10, max_losses=3,
        cap_mode="total",
    ),
]


# ═══════════════════════════════════════════════════════════════════════════
# 4. SIMULACIÓN FIEL AL ENGINE
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class SessionRecord:
    date:         str
    idx:          int
    wins:         int
    losses:       int
    signals_used: int
    pnl:          float
    start_bal:    float
    end_bal:      float
    stop_reason:  str    # target_wins / max_losses / session_exhausted


@dataclass
class DayRecord:
    date:      str
    pnl:       float       = 0.0
    start_bal: float       = 0.0
    end_bal:   float       = 0.0
    min_bal:   float       = 0.0
    sessions:  int         = 0
    meta_40:   bool        = False
    meta_60:   bool        = False


@dataclass
class SimResult:
    cfg:              SystemConfig
    balance:          float
    peak:             float
    min_bal:          float
    max_dd:           float
    max_dd_pct:       float
    sessions:         list[SessionRecord] = field(default_factory=list)
    days:             list[DayRecord]     = field(default_factory=list)
    days_meta_40:     int = 0
    days_meta_60:     int = 0
    days_total:       int = 0
    sessions_total:   int = 0
    sessions_won:     int = 0
    sessions_guarded: int = 0
    sigs_used:        int = 0
    max_loss_per_sig: float = 0.0
    worst_streak:     int = 0    # racha de sesiones perdidas
    best_streak:      int = 0    # racha de sesiones ganadas


def _calc_amounts(cfg: SystemConfig, entry_raw: float) -> tuple[float, float, float]:
    """Devuelve (entry, g1, g2) respetando el modo de cap."""
    cap_money = cfg.base * cfg.cap_pct
    if cfg.cap_mode == "per_step":
        # BUG original: cap aplicado por paso individualmente
        cap = round(cap_money, 2)
        entry = round(min(entry_raw, cap), 2)
        g1    = round(min(entry_raw * G1_MULT, cap), 2)
        g2    = round(min(entry_raw * G2_MULT, cap), 2)
    else:
        # FIX correcto: cap sobre la exposición TOTAL
        entry_max = round(cap_money / TOTAL_MULT, 4)
        entry = round(min(entry_raw, entry_max), 2)
        g1    = round(entry * G1_MULT, 2)
        g2    = round(entry * G2_MULT, 2)
    return max(0.01, entry), g1, g2


def simulate(outcomes: list[Outcome], cfg: SystemConfig) -> SimResult:
    table = get_table(cfg.n_ops, cfg.w_needed)

    balance = cfg.base
    peak    = cfg.base
    min_bal = cfg.base
    max_dd  = 0.0

    # Agrupar señales por día
    by_day: dict[str, list[Outcome]] = defaultdict(list)
    for o in outcomes:
        by_day[o.timestamp.strftime("%d/%m/%Y")].append(o)

    all_sessions: list[SessionRecord] = []
    all_days:     list[DayRecord]     = []

    sigs_used        = 0
    max_loss_per_sig = 0.0
    w_streak = l_streak = 0
    best_streak = worst_streak = 0

    for day_str in sorted(by_day, key=lambda d: datetime.strptime(d, "%d/%m/%Y")):
        day_outs = by_day[day_str]
        dr = DayRecord(date=day_str, start_bal=round(balance, 2), min_bal=round(balance, 2))

        # Dividir en bloques de N_OPS señales → "sesiones"
        chunks = [day_outs[i:i + cfg.n_ops] for i in range(0, len(day_outs), cfg.n_ops)]

        for s_idx, chunk in enumerate(chunks, start=1):
            s_wins = s_losses = s_used = 0
            s_pnl  = 0.0
            start_bal = round(balance, 2)
            blocked   = False
            stop_reason = "session_exhausted"

            for outcome in chunk:
                if s_wins >= cfg.w_needed:
                    stop_reason = "target_wins"
                    break
                if cfg.max_losses > 0 and s_losses >= cfg.max_losses:
                    stop_reason = "max_losses"
                    blocked = True
                    break

                raw_stake = table.get((s_losses, s_wins), 0.0)
                entry_raw = raw_stake * cfg.base
                entry, g1, g2 = _calc_amounts(cfg, entry_raw)
                if entry <= 0:
                    break

                s_used    += 1
                sigs_used += 1

                if outcome.result == "WD":
                    gain = round(entry * NET_PAYOUT, 2)
                    balance = round(balance + gain, 2)
                    s_pnl   = round(s_pnl + gain, 2)
                    s_wins += 1
                elif outcome.result == "G1":
                    # Pierde entry, gana G1
                    net = round(g1 * NET_PAYOUT - entry, 2)
                    balance = round(balance + net, 2)
                    s_pnl   = round(s_pnl + net, 2)
                    s_wins += 1
                elif outcome.result == "G2":
                    # Pierde entry+G1, gana G2
                    net = round(g2 * NET_PAYOUT - entry - g1, 2)
                    balance = round(balance + net, 2)
                    s_pnl   = round(s_pnl + net, 2)
                    s_wins += 1
                else:  # L
                    total_loss = round(entry + g1 + g2, 2)
                    balance    = round(balance - total_loss, 2)
                    s_pnl      = round(s_pnl - total_loss, 2)
                    s_losses  += 1
                    max_loss_per_sig = max(max_loss_per_sig, total_loss)

                min_bal   = min(min_bal, balance)
                dr.min_bal = min(dr.min_bal, balance)
                peak       = max(peak, balance)
                max_dd     = max(max_dd, round(peak - balance, 2))

            if s_wins >= cfg.w_needed:
                stop_reason = "target_wins"
                w_streak += 1
                l_streak  = 0
            else:
                l_streak += 1
                w_streak  = 0
            best_streak  = max(best_streak, w_streak)
            worst_streak = max(worst_streak, l_streak)

            sr = SessionRecord(
                date=day_str, idx=s_idx,
                wins=s_wins, losses=s_losses, signals_used=s_used,
                pnl=round(s_pnl, 2), start_bal=start_bal, end_bal=round(balance, 2),
                stop_reason=stop_reason,
            )
            all_sessions.append(sr)
            dr.sessions += 1
            dr.pnl       = round(dr.pnl + s_pnl, 2)

        dr.end_bal  = round(balance, 2)
        dr.meta_40  = dr.pnl >= 40.0
        dr.meta_60  = dr.pnl >= 60.0
        all_days.append(dr)

    max_dd_pct = round(max_dd / cfg.base * 100, 2)

    return SimResult(
        cfg=cfg, balance=round(balance, 2), peak=round(peak, 2),
        min_bal=round(min_bal, 2), max_dd=round(max_dd, 2), max_dd_pct=max_dd_pct,
        sessions=all_sessions, days=all_days,
        days_meta_40=sum(1 for d in all_days if d.meta_40),
        days_meta_60=sum(1 for d in all_days if d.meta_60),
        days_total=len(all_days),
        sessions_total=len(all_sessions),
        sessions_won=sum(1 for s in all_sessions if s.stop_reason == "target_wins"),
        sessions_guarded=sum(1 for s in all_sessions if s.stop_reason == "max_losses"),
        sigs_used=sigs_used,
        max_loss_per_sig=round(max_loss_per_sig, 2),
        worst_streak=worst_streak,
        best_streak=best_streak,
    )


# ═══════════════════════════════════════════════════════════════════════════
# 5. HELPERS EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def _fill(h: str) -> PatternFill:
    return PatternFill("solid", fgColor=h)

def _border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _hdr(cell: Any, bg: str = AH, color: str = WH, size: int = 10) -> None:
    cell.fill = _fill(bg); cell.font = Font(bold=True, color=color, size=size)
    cell.alignment = _center(); cell.border = _border()

def _body(cell: Any, even: bool, bold: bool = False, color: str = "000000") -> None:
    cell.fill = _fill(GR if even else WH)
    cell.font = Font(size=10, bold=bold, color=color)
    cell.alignment = _center(); cell.border = _border()

def _color_pnl(v: float) -> str:
    return GH if v > 0 else (RH if v < 0 else "000000")

def _set_widths(ws: Any, widths: list[float]) -> None:
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


# ═══════════════════════════════════════════════════════════════════════════
# 6. HOJA 1 — COMPARATIVA
# ═══════════════════════════════════════════════════════════════════════════

def sheet_comparativa(wb: Any, results: list[SimResult], outcomes: list[Outcome]) -> None:
    ws = wb.active
    ws.title = "Comparativa"
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [38, 20, 20, 20, 10])

    # Título
    ws.merge_cells("A1:E1")
    ws["A1"].value = "AUDITORÍA DEL BOT — Comparativa 3 Sistemas sobre datos reales (ejemplo.md)"
    _hdr(ws["A1"], bg=AH, size=13)
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:E2")
    ws["A2"].value = (
        f"2,655 señales reales · 52 días (17/03–07/05/2026) · "
        f"WR real=87.87% · Payout=92% · Base=$300 · Cap=10%"
    )
    _hdr(ws["A2"], bg=AM, size=10)
    ws.row_dimensions[2].height = 20

    cols_hdr = ["KPI", "Sistema A\nBug cap (12/4)", "Sistema B\nFix cap (12/4)", "Sistema C\nÓptimo (6/2+guard)"]
    ws.row_dimensions[4].height = 36
    for c, t in enumerate(cols_hdr, 1):
        _hdr(ws.cell(row=4, column=c), bg=AH if c == 1 else SYSTEMS[c - 2].color)
    for c, t in enumerate(cols_hdr, 1):
        ws.cell(row=4, column=c).value = t

    kpis = [
        ("Balance final",        lambda r: f"${r.balance:,.2f}"),
        ("ROI total",            lambda r: f"{(r.balance - r.cfg.base) / r.cfg.base * 100:+.2f}%"),
        ("P&L total",            lambda r: f"${r.balance - r.cfg.base:+,.2f}"),
        ("Max DrawDown ($)",     lambda r: f"${r.max_dd:,.2f}"),
        ("Max DrawDown (%)",     lambda r: f"{r.max_dd_pct:.2f}%"),
        ("Pérdida máx/señal",    lambda r: f"${r.max_loss_per_sig:,.2f}"),
        ("Cap real por señal",   lambda r: f"${r.cfg.base * r.cfg.cap_pct:,.2f}  ({'✔' if r.cfg.cap_mode == 'total' else '✖ BUG'})"),
        ("Señales ejecutadas",   lambda r: str(r.sigs_used)),
        ("Sesiones totales",     lambda r: str(r.sessions_total)),
        ("Sesiones ganadas",     lambda r: f"{r.sessions_won} ({r.sessions_won / r.sessions_total * 100:.0f}%)"),
        ("Sesiones por guard",   lambda r: str(r.sessions_guarded)),
        ("Días meta ≥$40",       lambda r: f"{r.days_meta_40}/{r.days_total} ({r.days_meta_40 / r.days_total * 100:.0f}%)"),
        ("Días meta ≥$60",       lambda r: f"{r.days_meta_60}/{r.days_total} ({r.days_meta_60 / r.days_total * 100:.0f}%)"),
        ("P&L promedio/día",     lambda r: f"${(r.balance - r.cfg.base) / r.days_total:+.2f}"),
        ("Racha ganadora máx",   lambda r: str(r.best_streak)),
        ("Racha perdedora máx",  lambda r: str(r.worst_streak)),
        ("Config N_OPS/W_NEEDED",lambda r: f"{r.cfg.n_ops}/{r.cfg.w_needed}"),
        ("Max losses guard",     lambda r: str(r.cfg.max_losses) if r.cfg.max_losses > 0 else "— desactivado"),
    ]

    for row_i, (kpi_name, fn) in enumerate(kpis, start=5):
        even = (row_i % 2 == 0)
        ws.cell(row=row_i, column=1).value = kpi_name
        ws.cell(row=row_i, column=1).font  = Font(bold=True, size=10)
        ws.cell(row=row_i, column=1).fill  = _fill(GR if even else WH)
        ws.cell(row=row_i, column=1).border = _border()
        ws.cell(row=row_i, column=1).alignment = _left()
        for col_i, res in enumerate(results, start=2):
            val = fn(res)
            cell = ws.cell(row=row_i, column=col_i, value=val)
            _body(cell, even)
            # Resaltar bugs
            if kpi_name == "Cap real por señal" and res.cfg.cap_mode == "per_step":
                cell.fill = _fill(RB)
                cell.font = Font(bold=True, size=10, color=RH)
            if kpi_name.startswith("Pérdida") and res.cfg.cap_mode == "per_step":
                cell.fill = _fill(RB)
                cell.font = Font(bold=True, size=10, color=RH)
        ws.row_dimensions[row_i].height = 20

    # ── Mini tabla para gráfica ─────────────────────────────────────────
    chart_row = 5 + len(kpis) + 2
    ws.cell(row=chart_row, column=1).value = "Sistema"
    ws.cell(row=chart_row, column=2).value = "Balance final"
    ws.cell(row=chart_row, column=3).value = "Días meta ≥$60"
    for c in range(1, 4):
        _hdr(ws.cell(row=chart_row, column=c), bg=AM)
    for i, res in enumerate(results, start=1):
        r = chart_row + i
        ws.cell(row=r, column=1).value = f"Sistema {chr(64+i)}"
        ws.cell(row=r, column=2).value = res.balance
        ws.cell(row=r, column=3).value = res.days_meta_60
        for c in range(1, 4):
            _body(ws.cell(row=r, column=c), i % 2 == 0)

    bar = BarChart(); bar.type = "col"; bar.style = 10
    bar.title = "Balance final por sistema"
    bar.height = 12; bar.width = 18
    data = Reference(ws, min_col=2, min_row=chart_row, max_row=chart_row + len(results))
    cats = Reference(ws, min_col=1, min_row=chart_row + 1, max_row=chart_row + len(results))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, f"F4")


# ═══════════════════════════════════════════════════════════════════════════
# 7. HOJA DE DETALLE (DÍAS + SESIONES)
# ═══════════════════════════════════════════════════════════════════════════

def sheet_detalle(wb: Any, res: SimResult, label: str) -> None:
    ws = wb.create_sheet(f"Detalle {label}")
    ws.sheet_view.showGridLines = False

    cols_day = ["Fecha", "P&L día", "Bal inicio", "Bal fin", "Bal mín", "Sesiones", "≥$40", "≥$60", "DD intradia"]
    _set_widths(ws, [13, 12, 12, 12, 12, 10, 8, 8, 12, 6,
                     10, 8, 8, 12, 12, 12, 14])

    ws.merge_cells(f"A1:{get_column_letter(len(cols_day) + 8)}1")
    ws["A1"].value = f"DETALLE DÍA A DÍA + SESIONES — {res.cfg.name}"
    _hdr(ws["A1"], bg=res.cfg.color, size=11)
    ws.row_dimensions[1].height = 28

    for c, t in enumerate(cols_day, 1):
        _hdr(ws.cell(row=2, column=c), bg=AH)
        ws.cell(row=2, column=c).value = t

    for r, d in enumerate(res.days, start=3):
        even = (r % 2 == 0)
        dd_int = round(d.start_bal - d.min_bal, 2)
        vals   = [d.date, d.pnl, d.start_bal, d.end_bal, d.min_bal, d.sessions,
                  "✔" if d.meta_40 else "—", "✔" if d.meta_60 else "—", dd_int]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if d.meta_60:
                cell.fill = _fill(GB)
            elif d.pnl < 0:
                cell.fill = _fill(RB)
            else:
                _body(cell, even)
            cell.border = _border(); cell.alignment = _center()
            cell.font   = Font(size=10, bold=(c == 2),
                               color=(_color_pnl(d.pnl) if c == 2 else "000000"))
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A3"

    # Sesiones en columnas separadas
    ses_col_start = len(cols_day) + 2
    ses_cols = ["Fecha", "Ses#", "Wins", "Losses", "Sigs", "P&L ses", "Bal ini", "Bal fin", "Stop"]
    for c, t in enumerate(ses_cols, start=ses_col_start):
        _hdr(ws.cell(row=2, column=c), bg=AM)
        ws.cell(row=2, column=c).value = t

    for r, s in enumerate(res.sessions, start=3):
        even = (r % 2 == 0)
        vals = [s.date, s.idx, s.wins, s.losses, s.signals_used,
                s.pnl, s.start_bal, s.end_bal, s.stop_reason]
        for c, v in enumerate(vals, start=ses_col_start):
            cell = ws.cell(row=r, column=c, value=v)
            _body(cell, even, bold=(c == ses_col_start + 5),
                  color=(_color_pnl(s.pnl) if c == ses_col_start + 5 else "000000"))
            if s.stop_reason == "max_losses":
                if c >= ses_col_start:
                    cell.fill = _fill(OB)
        ws.row_dimensions[r].height = 18


# ═══════════════════════════════════════════════════════════════════════════
# 8. HOJA EQUITY CURVES
# ═══════════════════════════════════════════════════════════════════════════

def sheet_equity_curves(wb: Any, results: list[SimResult]) -> None:
    ws = wb.create_sheet("Equity Curves")
    ws.sheet_view.showGridLines = False

    # Construir series de balance diario acumulado
    all_dates = sorted({d.date for res in results for d in res.days},
                       key=lambda x: datetime.strptime(x, "%d/%m/%Y"))

    ws.cell(row=1, column=1).value = "Fecha"
    _hdr(ws.cell(row=1, column=1), bg=AH)
    for c, res in enumerate(results, start=2):
        ws.cell(row=1, column=c).value = res.cfg.name
        _hdr(ws.cell(row=1, column=c), bg=res.cfg.color)
        ws.column_dimensions[get_column_letter(c)].width = 28
    ws.column_dimensions["A"].width = 14

    day_map: list[dict[str, float]] = []
    balances = {res.cfg.name: res.cfg.base for res in results}

    for r, date in enumerate(all_dates, start=2):
        ws.cell(row=r, column=1).value = date
        _body(ws.cell(row=r, column=1), r % 2 == 0)
        for c, res in enumerate(results, start=2):
            d_match = next((d for d in res.days if d.date == date), None)
            if d_match:
                balances[res.cfg.name] = round(balances[res.cfg.name] + d_match.pnl, 2)
            val = balances[res.cfg.name]
            cell = ws.cell(row=r, column=c, value=val)
            cell.number_format = '"$"#,##0.00'
            _body(cell, r % 2 == 0)
        ws.row_dimensions[r].height = 18

    n_rows = len(all_dates)
    chart = LineChart()
    chart.title  = "Equity Curve — Balance diario acumulado"
    chart.style  = 10; chart.height = 16; chart.width = 32
    chart.y_axis.title = "Balance ($)"
    chart.x_axis.title = "Día"

    for c, res in enumerate(results, start=2):
        data = Reference(ws, min_col=c, min_row=1, max_row=1 + n_rows)
        chart.add_data(data, titles_from_data=True)
        chart.series[c - 2].graphicalProperties.line.solidFill = res.cfg.color
        chart.series[c - 2].graphicalProperties.line.width = 20000

    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_rows)
    chart.set_categories(cats)
    ws.add_chart(chart, "F2")


# ═══════════════════════════════════════════════════════════════════════════
# 9. HOJA AUDITORÍA DE RIESGO
# ═══════════════════════════════════════════════════════════════════════════

def sheet_auditoria_riesgo(wb: Any, results: list[SimResult], outcomes: list[Outcome]) -> None:
    ws = wb.create_sheet("Auditoria Riesgo")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [36, 24, 24, 24, 6])

    ws.merge_cells("A1:E1")
    ws["A1"].value = "AUDITORÍA DE RIESGO — Exposición máxima y distribución de pérdidas"
    _hdr(ws["A1"], bg=AH, size=12); ws.row_dimensions[1].height = 30

    hdr_cols = ["Métrica de riesgo", "Sistema A (bug)", "Sistema B (fix)", "Sistema C (óptimo)"]
    for c, t in enumerate(hdr_cols, 1):
        ws.cell(row=3, column=c).value = t
        _hdr(ws.cell(row=3, column=c), bg=(AH if c == 1 else SYSTEMS[c - 2].color))

    rows_data = [
        ("Cap correcto aplicado",
         "❌ cap per-paso ($30 x3)",
         "✔ cap total ($30 total)",
         "✔ cap total ($30 total)"),
        ("Pérdida máxima teórica (L)",
         f"${BASE * CAP_PCT:.0f} x3 pasos = ${BASE * CAP_PCT * TOTAL_MULT / (1 + G1_MULT + G2_MULT) * TOTAL_MULT:.2f}",
         f"${BASE * CAP_PCT:.2f}",
         f"${BASE * CAP_PCT:.2f}"),
        ("Pérdida máxima registrada",
         *[f"${r.max_loss_per_sig:.2f}" for r in results]),
        ("Max DD ($)",
         *[f"${r.max_dd:.2f}" for r in results]),
        ("Max DD (% de base)",
         *[f"{r.max_dd_pct:.2f}%" for r in results]),
        ("Balance mínimo alcanzado",
         *[f"${r.min_bal:.2f}" for r in results]),
        ("Señales totales ejecutadas",
         *[str(r.sigs_used) for r in results]),
        ("Señales promedio/día",
         *[f"{r.sigs_used / r.days_total:.1f}" for r in results]),
        ("Racha perdedora de sesiones",
         *[str(r.worst_streak) for r in results]),
        ("Racha ganadora de sesiones",
         *[str(r.best_streak) for r in results]),
        ("Sesiones con guard (max_losses)",
         *[str(r.sessions_guarded) for r in results]),
    ]

    for r, row_data in enumerate(rows_data, start=4):
        even = (r % 2 == 0)
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            _body(cell, even, bold=(c == 1))
            cell.alignment = _left() if c == 1 else _center()
            if c == 2 and "❌" in str(val):
                cell.fill = _fill(RB); cell.font = Font(bold=True, size=10, color=RH)
        ws.row_dimensions[r].height = 22

    # Nota explicativa del bug
    last = 4 + len(rows_data)
    ws.merge_cells(f"A{last}:E{last+3}")
    cell = ws[f"A{last}"]
    cell.value = (
        "🔴 BUG Sistema A — Explicación técnica:\n"
        "  min(entry,$30), min(G1,$30), min(G2,$30) → cada paso individualmente cappado\n"
        "  Si entry=$14.70: G1=$30.68, G2=$64.03 → total exposición si pierde todo = $14.70+$30+$30 = $74.70 (24.9% del capital)\n"
        "  Con entry_raw alto: G1=$30, G2=$30 → total = $90 (30% del capital en 1 señal)\n\n"
        "✅ Fix aplicado (Sistemas B y C):\n"
        "  entry_max = $300 × 10% / 7.44 = $4.03\n"
        "  entry=$4.03, G1=$8.41, G2=$17.56 → total = $30.00 (10% exacto sin importar cuántos gales caigan)"
    )
    cell.fill = _fill(YB); cell.font = Font(size=9)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[last].height = 80


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main() -> None:
    print("Leyendo señales reales de ejemplo.md...")
    outcomes = parse_outcomes(HISTORY)
    wd = sum(1 for o in outcomes if o.result == "WD")
    g1 = sum(1 for o in outcomes if o.result == "G1")
    g2 = sum(1 for o in outcomes if o.result == "G2")
    l  = sum(1 for o in outcomes if o.result == "L")
    wr = (wd + g1 + g2) / len(outcomes) * 100
    print(f"  {len(outcomes)} señales  WD={wd}  G1={g1}  G2={g2}  L={l}  WR={wr:.2f}%")
    print()

    results: list[SimResult] = []
    for cfg in SYSTEMS:
        print(f"Simulando: {cfg.name} ...")
        res = simulate(outcomes, cfg)
        roi = (res.balance - cfg.base) / cfg.base * 100
        m60 = res.days_meta_60 / res.days_total * 100 if res.days_total else 0
        print(f"  Balance={res.balance:.2f}  ROI={roi:+.1f}%  MaxDD=${res.max_dd:.2f}({res.max_dd_pct:.1f}%)  "
              f"Días≥$60={res.days_meta_60}/{res.days_total}({m60:.0f}%)  "
              f"PérdMáx/señal=${res.max_loss_per_sig:.2f}")
        results.append(res)

    print()
    print("Generando Excel de auditoría...")
    wb = openpyxl.Workbook()
    sheet_comparativa(wb, results, outcomes)
    for res, lbl in zip(results, ["A", "B", "C"]):
        sheet_detalle(wb, res, lbl)
    sheet_equity_curves(wb, results)
    sheet_auditoria_riesgo(wb, results, outcomes)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\n✔  Informe guardado en: {OUTPUT}")

    print()
    print("=" * 65)
    print("VEREDICTO DE AUDITORÍA")
    print("=" * 65)
    for res in results:
        roi = (res.balance - res.cfg.base) / res.cfg.base * 100
        m60 = res.days_meta_60 / res.days_total * 100 if res.days_total else 0
        print(f"  {res.cfg.name[:38]:<38}  "
              f"ROI={roi:+6.1f}%  DD={res.max_dd_pct:5.1f}%  "
              f"Pérd/señal=${res.max_loss_per_sig:.2f}  "
              f"Meta60={m60:.0f}%")


if __name__ == "__main__":
    main()
